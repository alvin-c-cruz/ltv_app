from flask import Blueprint, render_template, g, request, send_file, abort
from datetime import date, timedelta
from ...tz import ph_today
import io
from .. transactions import get_balance, get_transactions
from .. auth import login_required

bp = Blueprint('bank', __name__, template_folder='pages', url_prefix='/bank')


def _fmt_date(d):
    return date.fromisoformat(d).strftime('%d %b %Y') if d else ''

def _fmt_int(value):
    if value < 0:
        return '({:,.0f})'.format(abs(value))
    return '{:,.0f}'.format(value)

def _fmt_num(value, decimals=2):
    if value < 0:
        return '({:,.{}f})'.format(abs(value), decimals)
    return '{:,.{}f}'.format(value, decimals)


@bp.route('/')
@login_required
def home():
    return render_template('bank/home.html')


@bp.route('/<bank_id>')
@login_required
def account_position(bank_id):
    from .. database import get_db
    db = get_db()
    row = db.execute("SELECT ref_num, bank_name FROM tbl_bank_account WHERE bank_id=?;", (bank_id,)).fetchone()
    if row is None:
        abort(404)  # bank_id is a string code ('DBPe'), so no <int:> converter can catch a bad one
    bank_ref, bank_name = row['ref_num'], row['bank_name']
    trade_date = str(ph_today())
    position = gather_position(db=db, trade_date=trade_date, bank_ref=bank_ref)

    return render_template('bank/account.html', position=position, bank_name=bank_name, bank_id=bank_id)


# _dict = {
#     "ccy": {
#         "int(bank_ref)": {
#             "beginning": {
#                 "int(code_ref)": {
#                     "quantity": 0,
#                     "cost_to_date": 0,
#                     "average": 0
#                 }
#             },
#         }
#     }
# }
def gather_position(db, trade_date, bank_ref):
    dict_position = {}

    for stock in g.stocks:
        code_ref = stock['ref_num']
        quantity, cost_to_date = get_balance(db=db, bank_ref=bank_ref, code_ref=code_ref, trade_date=trade_date)
        if quantity:
            ccy = db.execute('SELECT ccy_id FROM tbl_currency '
                             'INNER JOIN tbl_code ON tbl_code.ccy_ref = tbl_currency.ref_num '
                             'WHERE tbl_code.ref_num=?;', (code_ref, )).fetchone()[0]

            if ccy not in dict_position:
                dict_position[ccy] = {}

            dict_position[ccy][code_ref] = {
                "code": stock["code"],
                "name": stock["stock_name"],
                "quantity": '{:,.0f}'.format(quantity),
                "average": '{:,.4f}'.format(cost_to_date / quantity),
            }

    return dict_position


def gather_position_bulk(db, bank_ref):
    """Faster alternative to gather_position — fetches all transactions for
    a bank in one query instead of one query per stock."""

    transaction_basis = db.execute(
        "SELECT transaction_basis FROM tbl_bank_account WHERE ref_num=?", (bank_ref,)
    ).fetchone()[0]

    rows = db.execute(
        "SELECT t.code_ref, t.quantity, t.price, "
        "t.brokerage, t.commission, t.foreign_charge, t.stamp_duty, t.misc "
        "FROM tbl_transaction t "
        "INNER JOIN tbl_transaction_type tt ON tt.transaction_type = t.transaction_type "
        f"WHERE t.bank_ref=? "
        f"ORDER BY t.code_ref, t.{transaction_basis}, tt.priority",
        (bank_ref,)
    ).fetchall()

    # replicate the balance logic from get_balance for each stock in one pass
    balances = {}  # code_ref -> [balance, cost_to_date]
    for row in rows:
        code_ref = row['code_ref']
        if code_ref not in balances:
            balances[code_ref] = [0, 0.0]

        balance, cost_to_date = balances[code_ref]
        quantity = row['quantity']
        charges = (row['brokerage'] + row['commission'] + row['foreign_charge']
                   + row['stamp_duty'] + row['misc'])
        amount = quantity * row['price'] + charges

        if quantity > 0:
            if balance > 0:
                cost_to_date += amount
            elif balance == 0:
                cost_to_date += amount
            elif balance < 0:
                if balance + quantity == 0:
                    cost_to_date = 0
                elif balance + quantity < 0:
                    cost_to_date = 0
                else:
                    cost_to_date = (balance + quantity) / quantity * amount
        else:
            if balance > 0:
                if balance - abs(quantity) > 0:
                    cost_to_date -= cost_to_date * abs(quantity) / balance
                else:
                    cost_to_date = 0
            else:
                cost_to_date = 0

        balances[code_ref] = [balance + quantity, cost_to_date]

    # keep non-zero balances (includes short/negative positions)
    balances = {cr: (bal, cost) for cr, (bal, cost) in balances.items() if bal != 0}
    if not balances:
        return {}

    # batch-fetch currencies for all remaining stocks in one query
    placeholders = ','.join('?' * len(balances))
    ccy_rows = db.execute(
        "SELECT tbl_code.ref_num, tbl_currency.ccy_id "
        "FROM tbl_currency "
        "INNER JOIN tbl_code ON tbl_code.ccy_ref = tbl_currency.ref_num "
        f"WHERE tbl_code.ref_num IN ({placeholders})",
        list(balances.keys())
    ).fetchall()
    ccy_map = {row['ref_num']: row['ccy_id'] for row in ccy_rows}

    stock_info = {s['ref_num']: s for s in g.stocks}
    dict_position = {}
    for code_ref, (quantity, cost_to_date) in balances.items():
        stock = stock_info.get(code_ref)
        ccy = ccy_map.get(code_ref)
        if not stock or not ccy:
            continue
        if ccy not in dict_position:
            dict_position[ccy] = {}
        qty_fmt = ('({:,.0f})'.format(abs(quantity)) if quantity < 0
                   else '{:,.0f}'.format(quantity))
        avg_fmt = ('—' if quantity <= 0
                   else '{:,.4f}'.format(cost_to_date / quantity))
        dict_position[ccy][code_ref] = {
            "code": stock["code"],
            "name": stock["stock_name"],
            "quantity": qty_fmt,
            "average": avg_fmt,
        }

    for ccy in dict_position:
        dict_position[ccy] = dict(
            sorted(dict_position[ccy].items(), key=lambda item: item[1]['code'])
        )

    return dict_position


def gather_short_position_bulk(db, bank_ref):
    """Net short positions for a bank from tbl_transaction_short."""
    rows = db.execute(
        "SELECT t.code_ref, SUM(t.quantity) as net_qty "
        "FROM tbl_transaction_short t "
        "WHERE t.bank_ref=? "
        "GROUP BY t.code_ref "
        "HAVING net_qty != 0",
        (bank_ref,)
    ).fetchall()

    if not rows:
        return {}

    code_refs = [r['code_ref'] for r in rows]
    placeholders = ','.join('?' * len(code_refs))
    ccy_rows = db.execute(
        "SELECT tbl_code.ref_num, tbl_currency.ccy_id "
        "FROM tbl_currency "
        "INNER JOIN tbl_code ON tbl_code.ccy_ref = tbl_currency.ref_num "
        f"WHERE tbl_code.ref_num IN ({placeholders})",
        code_refs
    ).fetchall()
    ccy_map = {row['ref_num']: row['ccy_id'] for row in ccy_rows}

    stock_info = {s['ref_num']: s for s in g.stocks}
    dict_position = {}
    for row in rows:
        code_ref = row['code_ref']
        quantity = row['net_qty']
        stock = stock_info.get(code_ref)
        ccy = ccy_map.get(code_ref)
        if not stock or not ccy:
            continue
        if ccy not in dict_position:
            dict_position[ccy] = {}
        qty_fmt = ('({:,.0f})'.format(abs(quantity)) if quantity < 0
                   else '{:,.0f}'.format(quantity))
        dict_position[ccy][code_ref] = {
            "code": stock["code"],
            "name": stock["stock_name"],
            "quantity": qty_fmt,
            "is_short": quantity < 0,
        }

    for ccy in dict_position:
        dict_position[ccy] = dict(
            sorted(dict_position[ccy].items(), key=lambda item: item[1]['code'])
        )

    return dict_position


def _compute_transactions(db, bank_ref, code_ref, date_from, date_to):
    date_before = str(date.fromisoformat(date_from) - timedelta(days=1))
    start_qty, start_cost = get_balance(db=db, bank_ref=bank_ref, code_ref=code_ref, trade_date=date_before)
    start_avg = start_cost / start_qty if start_qty else 0.0

    db_data = get_transactions(db=db, date_from=date_from, date_to=date_to, bank_ref=bank_ref, code_ref=code_ref)
    balance = start_qty
    cost_to_date = start_cost
    result = []

    for data in db_data:
        qty = data['quantity']
        price = data['price']
        charges = data['charges']
        amount = qty * price + charges

        gain_loss = 0.0
        if qty < 0 and balance > 0:
            gain_loss = abs(qty) * price - charges - (cost_to_date / balance) * abs(qty)

        if qty > 0:
            if balance >= 0:
                cost_to_date += amount
            elif balance < 0:
                if balance + qty <= 0:
                    cost_to_date = 0
                else:
                    cost_to_date = (balance + qty) / qty * amount
        else:
            if balance > 0:
                if balance - abs(qty) > 0:
                    cost_to_date -= cost_to_date * abs(qty) / balance
                else:
                    cost_to_date = 0
            else:
                cost_to_date = 0

        balance += qty
        average = cost_to_date / balance if balance > 0 else 0.0

        data['amount'] = amount
        data['gain_loss'] = gain_loss
        data['balance'] = balance
        data['average'] = average
        result.append(data)

    return start_qty, start_avg, result


@bp.route('/<bank_id>/<code>', methods=['GET', 'POST'])
@login_required
def transaction_list(bank_id, code):
    from .. database import get_db
    db = get_db()
    bank_row = db.execute("SELECT ref_num, bank_name FROM tbl_bank_account WHERE bank_id=?;", (bank_id,)).fetchone()
    if bank_row is None:
        abort(404)
    bank_ref, bank_name = bank_row['ref_num'], bank_row['bank_name']
    code_row = db.execute("SELECT ref_num, stock_name FROM tbl_code WHERE code=?;", (code,)).fetchone()
    if code_row is None:
        abort(404)
    code_ref, stock_name = code_row['ref_num'], code_row['stock_name']

    if request.method == 'POST':
        date_from = request.form.get('date_from', '')
        date_to = request.form.get('date_to', '')
    else:
        today = ph_today()
        date_from = str(today.replace(month=1, day=1))
        date_to = str(today.replace(month=12, day=31))

    start_qty, start_avg, transactions = _compute_transactions(
        db=db, bank_ref=bank_ref, code_ref=code_ref, date_from=date_from, date_to=date_to
    )

    for t in transactions:
        t['trade_date_fmt'] = _fmt_date(t['trade_date'])
        t['value_date_fmt'] = _fmt_date(t['value_date'])
        t['quantity_fmt'] = _fmt_int(t['quantity'])
        t['price_fmt'] = '{:,.4f}'.format(t['price'])
        t['charges_fmt'] = _fmt_num(t['charges'], 2)
        t['amount_fmt'] = _fmt_num(t['amount'], 2)
        t['gain_loss_fmt'] = _fmt_num(t['gain_loss'], 2) if t['gain_loss'] != 0 else ''
        t['balance_fmt'] = _fmt_int(t['balance'])
        t['average_fmt'] = '{:,.4f}'.format(t['average']) if t['average'] else '—'

    form = {
        'bank_name': bank_name,
        'bank_id': bank_id,
        'code': code,
        'stock_name': stock_name,
        'date_from': date_from,
        'date_to': date_to,
        'start_quantity': _fmt_int(start_qty) if start_qty else '—',
        'start_average': '{:,.4f}'.format(start_avg) if start_avg else '—',
        'transactions': transactions,
    }

    return render_template('bank/transactions.html', form=form)


@bp.route('/<bank_id>/<code>/short', methods=['GET', 'POST'])
@login_required
def short_transaction_list(bank_id, code):
    from .. database import get_db
    db = get_db()
    bank_row = db.execute("SELECT ref_num, bank_name FROM tbl_bank_account WHERE bank_id=?;", (bank_id,)).fetchone()
    if bank_row is None:
        abort(404)
    bank_ref, bank_name = bank_row['ref_num'], bank_row['bank_name']
    code_row = db.execute("SELECT ref_num, stock_name FROM tbl_code WHERE code=?;", (code,)).fetchone()
    if code_row is None:
        abort(404)
    code_ref, stock_name = code_row['ref_num'], code_row['stock_name']

    if request.method == 'POST':
        date_from = request.form.get('date_from', '')
        date_to = request.form.get('date_to', '')
    else:
        today = ph_today()
        date_from = str(today.replace(month=1, day=1))
        date_to = str(today.replace(month=12, day=31))

    rows = db.execute(
        "SELECT ref_num, trade_date, value_date, transaction_type, quantity, price, "
        "brokerage, commission, foreign_charge, stamp_duty, misc "
        "FROM tbl_transaction_short "
        "WHERE bank_ref=? AND code_ref=? AND trade_date>=? AND trade_date<=? "
        "ORDER BY trade_date",
        (bank_ref, code_ref, date_from, date_to)
    ).fetchall()

    # compute balance and average cost forwarded from before date_from
    prior_rows = db.execute(
        "SELECT quantity, price, brokerage, commission, foreign_charge, stamp_duty, misc "
        "FROM tbl_transaction_short "
        "WHERE bank_ref=? AND code_ref=? AND trade_date<? "
        "ORDER BY trade_date",
        (bank_ref, code_ref, date_from)
    ).fetchall()

    def _update_short_cost(balance, cost_to_date, qty, price, charges):
        amount = abs(qty) * price + charges
        if qty < 0:  # opening/deepening short
            if balance <= 0:
                cost_to_date += amount
        else:  # closing short
            if balance < 0:
                if abs(balance) - qty > 0:
                    cost_to_date -= cost_to_date * qty / abs(balance)
                else:
                    cost_to_date = 0.0
        return cost_to_date

    balance = 0
    cost_to_date = 0.0
    for r in prior_rows:
        charges = r['brokerage'] + r['commission'] + r['foreign_charge'] + r['stamp_duty'] + r['misc']
        cost_to_date = _update_short_cost(balance, cost_to_date, r['quantity'], r['price'], charges)
        balance += r['quantity']

    start_qty = balance
    start_avg = cost_to_date / abs(balance) if balance != 0 else 0.0

    transactions = []
    for row in rows:
        charges = (row['brokerage'] + row['commission'] + row['foreign_charge']
                   + row['stamp_duty'] + row['misc'])
        cost_to_date = _update_short_cost(balance, cost_to_date, row['quantity'], row['price'], charges)
        balance += row['quantity']
        avg = cost_to_date / abs(balance) if balance != 0 else 0.0
        transactions.append({
            'ref_num': row['ref_num'],
            'trade_date_fmt': _fmt_date(row['trade_date']),
            'value_date_fmt': _fmt_date(row['value_date']),
            'description': row['transaction_type'],
            'quantity': row['quantity'],
            'quantity_fmt': _fmt_int(row['quantity']),
            'price_fmt': '{:,.4f}'.format(row['price']),
            'charges_fmt': _fmt_num(charges, 2),
            'amount_fmt': _fmt_num(row['quantity'] * row['price'] + charges, 2),
            'balance_fmt': _fmt_int(balance),
            'average_fmt': '{:,.4f}'.format(avg) if balance != 0 else '—',
        })

    form = {
        'bank_name': bank_name,
        'bank_id': bank_id,
        'code': code,
        'stock_name': stock_name,
        'date_from': date_from,
        'date_to': date_to,
        'start_quantity': _fmt_int(start_qty) if start_qty else '—',
        'start_average': '{:,.4f}'.format(start_avg) if start_qty else '—',
        'transactions': transactions,
    }

    return render_template('bank/short_transactions.html', form=form)


@bp.route('/<bank_id>/<code>/download')
@login_required
def download_transactions(bank_id, code):
    from .. database import get_db
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    db = get_db()
    bank_row = db.execute("SELECT ref_num, bank_name FROM tbl_bank_account WHERE bank_id=?;", (bank_id,)).fetchone()
    if bank_row is None:
        abort(404)
    bank_ref, bank_name = bank_row['ref_num'], bank_row['bank_name']
    code_row = db.execute("SELECT ref_num, stock_name FROM tbl_code WHERE code=?;", (code,)).fetchone()
    if code_row is None:
        abort(404)
    code_ref, stock_name = code_row['ref_num'], code_row['stock_name']

    today = ph_today()
    date_from = request.args.get('date_from', str(today.replace(month=1, day=1)))
    date_to = request.args.get('date_to', str(today.replace(month=12, day=31)))

    start_qty, start_avg, transactions = _compute_transactions(
        db=db, bank_ref=bank_ref, code_ref=code_ref, date_from=date_from, date_to=date_to
    )

    wb = Workbook()
    ws = wb.active
    ws.title = code

    # Format date range: same month → "January 2026"
    #                    same year  → "From January 1 to December 31, 2026"
    #                    diff years → "From January 1, 2025 to December 31, 2026"
    d1 = date.fromisoformat(date_from)
    d2 = date.fromisoformat(date_to)
    if d1.year == d2.year and d1.month == d2.month:
        date_label = d1.strftime('%B %Y')
    elif d1.year == d2.year:
        date_label = f"From {d1.strftime('%B')} {d1.day} to {d2.strftime('%B')} {d2.day}, {d1.year}"
    else:
        date_label = (f"From {d1.strftime('%B')} {d1.day}, {d1.year} "
                      f"to {d2.strftime('%B')} {d2.day}, {d2.year}")

    num_cols = 10

    thin  = Side(style='thin',   color="BBBBBB")
    med   = Side(style='medium', color="888888")
    thin_border  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    header_border = Border(left=thin, right=thin,  top=thin,  bottom=med)

    # ── Header rows ───────────────────────────────────────────────
    for r in range(1, 4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)

    ws['A1'] = bank_name
    ws['A1'].font = Font(bold=True, size=18)
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 34

    ws['A2'] = f"{stock_name} ({code})"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = Alignment(vertical='center')
    ws.row_dimensions[2].height = 26

    ws['A3'] = date_label
    ws['A3'].font = Font(size=12, italic=True, color="555555")
    ws['A3'].alignment = Alignment(vertical='center')
    ws.row_dimensions[3].height = 22

    # ── Column headers ────────────────────────────────────────────
    header_row = 5
    headers = ['Trade Date', 'Value Date', 'Description', 'Quantity', 'Price',
               'Charges', 'Amount', 'Gain / Loss', 'Balance', 'Average']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(fill_type='solid', fgColor="0C1A2E")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border
    ws.row_dimensions[header_row].height = 24

    num_fmt_int  = '#,##0;(#,##0)'
    num_fmt_2dp  = '#,##0.00;(#,##0.00)'
    num_fmt_4dp  = '#,##0.0000'
    date_fmt     = 'DD MMM YYYY'

    # ── Balance Forwarded row ─────────────────────────────────────
    bf_row = header_row + 1
    ws.row_dimensions[bf_row].height = 22
    bf_desc = ws.cell(row=bf_row, column=3, value='Balance Forwarded')
    bf_desc.font = Font(italic=True, size=12, color="888888")
    bf_desc.alignment = Alignment(horizontal='center')
    c9 = ws.cell(row=bf_row, column=9, value=start_qty)
    c9.number_format = num_fmt_int; c9.font = Font(size=12)
    c10 = ws.cell(row=bf_row, column=10, value=round(start_avg, 4))
    c10.number_format = num_fmt_4dp; c10.font = Font(size=12)
    for col in range(1, num_cols + 1):
        ws.cell(row=bf_row, column=col).border = thin_border

    # ── INDIRECT formulas — same string for every row ─────────────
    # D = Quantity (col 4), E = Price (col 5), F = Charges (col 6)
    # I = Balance (col 9), J = Average (col 10)
    #
    # Balance:  previous balance + this row's quantity
    BAL = '=INDIRECT("I"&ROW()-1)+INDIRECT("D"&ROW())'
    #
    # Average:
    #   if new balance = 0  → 0
    #   if qty > 0 (buy)    → weighted average with previous position
    #   if qty ≤ 0 (sell)   → unchanged from previous row
    AVG = ('=IF(INDIRECT("I"&ROW())=0,0,'
           'IF(INDIRECT("D"&ROW())>0,'
           '(INDIRECT("I"&ROW()-1)*INDIRECT("J"&ROW()-1)'
           '+INDIRECT("D"&ROW())*INDIRECT("E"&ROW()))'
           '/INDIRECT("I"&ROW()),'
           'INDIRECT("J"&ROW()-1)))')
    #
    # Gain/Loss: only when qty < 0 (sell or transfer out)
    #   gain = qty × (prev_avg − sell_price)   [qty is negative, so result is positive when sold above cost]
    GL  = ('=IF(INDIRECT("D"&ROW())<0,'
           'INDIRECT("D"&ROW())*(INDIRECT("J"&ROW()-1)-INDIRECT("E"&ROW())),'
           '"")' )

    # ── Data rows ─────────────────────────────────────────────────
    data_font = Font(size=12)
    for i, t in enumerate(transactions):
        r = bf_row + 1 + i
        ws.row_dimensions[r].height = 22
        c1 = ws.cell(row=r, column=1, value=date.fromisoformat(t['trade_date']))
        c1.number_format = date_fmt; c1.font = data_font; c1.border = thin_border
        c2 = ws.cell(row=r, column=2, value=date.fromisoformat(t['value_date']))
        c2.number_format = date_fmt; c2.font = data_font; c2.border = thin_border
        c3 = ws.cell(row=r, column=3, value=t['description'])
        c3.font = data_font; c3.border = thin_border
        c4 = ws.cell(row=r, column=4, value=t['quantity'])
        c4.number_format = num_fmt_int; c4.font = data_font; c4.border = thin_border
        c5 = ws.cell(row=r, column=5, value=round(t['price'], 4))
        c5.number_format = num_fmt_4dp; c5.font = data_font; c5.border = thin_border
        c6 = ws.cell(row=r, column=6, value=round(t['charges'], 2))
        c6.number_format = num_fmt_2dp; c6.font = data_font; c6.border = thin_border
        c7 = ws.cell(row=r, column=7, value=round(t['amount'], 2))
        c7.number_format = num_fmt_2dp; c7.font = data_font; c7.border = thin_border
        # H — Gain/Loss formula
        c8 = ws.cell(row=r, column=8, value=GL)
        c8.number_format = num_fmt_2dp; c8.border = thin_border
        # I — Balance formula
        c9 = ws.cell(row=r, column=9, value=BAL)
        c9.number_format = num_fmt_int; c9.font = data_font; c9.border = thin_border
        # J — Average formula
        c10 = ws.cell(row=r, column=10, value=AVG)
        c10.number_format = num_fmt_4dp; c10.font = data_font; c10.border = thin_border

    # ── Conditional formatting: colour Gain/Loss red/green ────────
    last_row = bf_row + len(transactions) + 200   # buffer for manual additions
    gl_range = f"H{bf_row + 1}:H{last_row}"
    ws.conditional_formatting.add(gl_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                   font=Font(size=12, color='1D7A45')))
    ws.conditional_formatting.add(gl_range,
        CellIsRule(operator='lessThan',    formula=['0'],
                   font=Font(size=12, color='A82828')))

    col_widths = [14, 14, 38, 13, 13, 13, 15, 15, 13, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{bank_id}_{code}_{date_from}_{date_to}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
