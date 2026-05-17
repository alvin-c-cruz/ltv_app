from dataclasses import dataclass
from pathlib import Path
import openpyxl


@dataclass
class CreateFile:
    instance_path: str = None
    data: any = None
    
    def __post_init__(self):
        wb = openpyxl.Workbook()
        self.write_data(wb)
        wb.save(self.filename)
        wb.close()
        
    def write_data(self, wb):
        ws = wb.active
        
        row = 1
        for bank_name, stocks in self.data.items():
            ws[f"A{row}"].value = bank_name
            row += 1


            ws[f"A{row}"].value = "Stock"
            ws[f"B{row}"].value = "Shares"
            ws[f"C{row}"].value = "Cost to Date"
            ws[f"D{row}"].value = "Unblocked Shares"
            ws[f"F{row}"].value = "Blocked Shares"
            ws[f"E{row}"].value = "Shares to Block"

            row += 1
            
            for stock_name, info in stocks.items():
                ws[f"A{row}"].value = stock_name
                ws[f"B{row}"].value = info['shares_on_hand']
                ws[f"C{row}"].value = info['cost-to-date']
                ws[f"D{row}"].value = info['unblocked_shares']
                ws[f"F{row}"].value = info['block_shares']
                ws[f"E{row}"].value = info['shares_to_block']
                
                row += 1
            
            row += 1
            
    
    @property
    def filename(self):
        path = Path(self.instance_path) / "temp" / "blocked_shares.xlsx"
        return path