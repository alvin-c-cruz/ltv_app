from flask import Blueprint, send_file, render_template
from flask_login import login_required
import openpyxl
import os
import tempfile

from ..database import get_db
from ...tz import ph_today

bp = Blueprint("stock_position", __name__, template_folder="pages", url_prefix="/stock-position")


@bp.route("/", methods=["GET"])
@login_required
def home():
    return render_template('stock_position/home.html')


@bp.route("/download", methods=["GET"])
@login_required
def download():
    """Generate and download stock balance report"""
    db = get_db()

    # Get stock positions for all bank accounts
    stock_positions = get_stock_positions(db)

    # Get blocked/unblocked shares from active DECU contracts
    to_block = get_blocked_unblocked(db)

    # Split each position into blocked/unblocked
    for bank_id in stock_positions:
        for code in stock_positions[bank_id]:
            pos = stock_positions[bank_id][code]
            blocked, unblocked = split_blocked_unblocked(
                pos['shares'], to_block.get(bank_id, {}).get(code, 0)
            )
            pos['blocked'] = blocked
            pos['unblocked'] = unblocked

    # Create Excel file
    excel_file = create_excel_report(stock_positions, db)

    # Send file
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=f'stock_balance_{ph_today().strftime("%Y%m%d")}.xlsx'
    )


def get_stock_positions(db):
    """
    Calculate LONG-side stock positions for all bank accounts, using the same
    weighted-average cost engine as block_unblock/transactions (accumulate_position).
    Short-side activity ('Sell (Short)', 'Buy (Pay Short)') is excluded — legacy
    only ever reports long positions (short tracking is permanently disabled
    there too), and mixing the two into one SUM corrupts both.
    Returns dict: {bank_id: {code: {shares, average_cost, total_cost}}}
    """
    from ..transactions.models import accumulate_position

    pairs = db.execute("""
        SELECT DISTINCT b.ref_num AS bank_ref, b.bank_id, c.ref_num AS code_ref, c.code
        FROM tbl_transaction t
        INNER JOIN tbl_bank_account b ON b.ref_num = t.bank_ref
        INNER JOIN tbl_code c ON c.ref_num = t.code_ref
        WHERE t.transaction_type NOT IN ('Sell (Short)', 'Buy (Pay Short)')
        ORDER BY b.priority, c.code
    """).fetchall()

    trade_date = str(ph_today())

    positions = {}
    for pair in pairs:
        transaction_basis = db.execute(
            "SELECT transaction_basis FROM tbl_bank_account WHERE ref_num=?",
            (pair['bank_ref'],)
        ).fetchone()[0]

        transactions = db.execute(
            "SELECT * FROM tbl_transaction "
            "INNER JOIN tbl_transaction_type "
            "ON tbl_transaction_type.transaction_type = tbl_transaction.transaction_type "
            "WHERE tbl_transaction.bank_ref=? AND tbl_transaction.code_ref=? "
            "AND tbl_transaction.transaction_type NOT IN ('Sell (Short)', 'Buy (Pay Short)') "
            f"AND tbl_transaction.{transaction_basis}<=? "
            f"ORDER BY tbl_transaction.{transaction_basis}, tbl_transaction_type.priority",
            (pair['bank_ref'], pair['code_ref'], trade_date)
        ).fetchall()

        shares, cost_to_date, average = accumulate_position(transactions)

        if shares == 0:
            continue

        bank_id = pair['bank_id']
        code = pair['code']

        if bank_id not in positions:
            positions[bank_id] = {}

        positions[bank_id][code] = {
            'shares': shares,
            'average_cost': abs(average),
            'total_cost': abs(cost_to_date)
        }

    return positions


def get_blocked_unblocked(db):
    """
    Total DECU shares still to be delivered per bank/code, same formula as
    ltv_app's block_unblock blueprint: daily_shares * remaining_days,
    doubled when leveraged. Caller clamps this against the actual share
    balance to split it into blocked/unblocked (legacy: blocked_shares class).
    """
    from ..term_sheet import StockContract

    sql = """
        SELECT tbl_stock_contract.ref_num
        FROM tbl_stock_contract
        INNER JOIN tbl_bank_account ON tbl_bank_account.ref_num = tbl_stock_contract.bank_ref
        INNER JOIN tbl_code ON tbl_code.ref_num = tbl_stock_contract.code_ref
        WHERE status="active"
            AND transaction_type="DECU"
    """
    active_decu_refs = [row['ref_num'] for row in db.execute(sql).fetchall()]

    to_block = {}
    for contract_ref in active_decu_refs:
        ts = StockContract(db=db)
        ts.get(ref_num=contract_ref)
        ts.get_schedules()

        bank_id = ts.bank_id
        code = ts.code

        if bank_id not in to_block:
            to_block[bank_id] = {}
        if code not in to_block[bank_id]:
            to_block[bank_id][code] = 0

        if ts.leveraged == 'Yes':
            to_block[bank_id][code] += ts.daily_shares * ts.remaining_days * 2
        else:
            to_block[bank_id][code] += ts.daily_shares * ts.remaining_days

    return to_block


def split_blocked_unblocked(shares_balance, to_block):
    """Clamp total_blocked against the actual balance — mirrors legacy's
    blocked_shares: if the amount still owed exceeds what's on hand, treat
    the whole balance as blocked rather than reporting negative unblocked."""
    if to_block >= shares_balance:
        return shares_balance, 0
    return to_block, shares_balance - to_block


def get_currency_from_code(code):
    """Determine currency from stock code"""
    if code.endswith('JP'):
        return 'JPY'
    elif ':' in code:
        suffix = code.split(':')[1]
        return suffix + 'D'  # e.g., HK -> HKD, SG -> SGD
    return 'HKD'  # Default


def annotate_decu_strikes(db, wb, sheet_name):
    """Port of legacy Stock_Balance.get_decu() (localhost/modules/stock_balance.py).

    Preserves its exact row offsets (H{row-2} for closing price, P{row-1} for the
    stock code cell, L{row} for the strike annotation itself — confirmed against
    both the legacy source and the live values already baked into
    excel_templates/stock_summary.xlsx, e.g. sheet "3" rows 9/11/12 and sheet "1"
    row 17, which line up with the bank-id row, not two rows above it) and its
    code-variable-carries-over-from-last-CBSG-row quirk on non-CBSG rows — this
    matches legacy behavior as-is, not a bug fix. See task-4 brief/report for
    details.
    """
    date_now = str(ph_today())

    ws = wb[sheet_name]
    counter = 0
    row_num = 0
    code = None

    while counter < 20:
        row_num += 1
        bank_id = ws[f"O{row_num}"].value
        if not bank_id:
            counter += 1
            continue
        else:
            counter = 0

        if bank_id == "CBSG":
            raw_code = ws[f"P{row_num - 1}"].value
            raw_code = raw_code[:len(raw_code) - 3]
            code = "{:0>4}".format(raw_code)

            price_row = db.execute(
                "SELECT tbl_stock_price.closing_price "
                "FROM tbl_stock_price "
                "INNER JOIN tbl_code ON tbl_code.ref_num = tbl_stock_price.code_ref "
                "WHERE tbl_stock_price.trade_date=? AND tbl_code.code=?",
                (date_now, code)
            ).fetchone()
            ws[f"H{row_num - 2}"].value = price_row['closing_price'] if price_row else 0

        # Note: on non-CBSG rows, `code` is whatever the last CBSG row set it to
        # (legacy scoping quirk, preserved deliberately — see docstring above).
        contract_refs = [row['ref_num'] for row in db.execute(
            "SELECT c.ref_num "
            "FROM tbl_stock_contract as c "
            "INNER JOIN tbl_bank_account as account ON account.ref_num = c.bank_ref "
            "INNER JOIN tbl_code as stock ON stock.ref_num = c.code_ref "
            "WHERE c.transaction_type='DECU' "
            "  AND c.status='active' "
            "  AND account.bank_id=? "
            "  AND stock.code=?",
            (bank_id, code)
        ).fetchall()]

        decus = []
        for contract_ref in contract_refs:
            ts_row = db.execute(
                "SELECT strike_rate, spot FROM tbl_stock_contract WHERE ref_num=?",
                (contract_ref,)
            ).fetchone()
            strike_value = ts_row['spot'] * ts_row['strike_rate'] / 100

            # "Still open" = at least one period not yet received. Verified against
            # the live DB: tbl_stock_contract_period.received stores "" (empty
            # string) for not-yet-received periods, never NULL (typeof() check
            # across the whole table returned only 'integer' and 'text', no
            # 'null') — matches legacy's `received != ""` / ts.footer['remaining']
            # logic and ltv_app/blueprints/term_sheet/models.py's own
            # `if period['received']:` truthiness check. Guard both NULL and ''
            # to be safe.
            remaining = db.execute(
                "SELECT COUNT(*) AS remaining FROM tbl_stock_contract_period "
                "WHERE contract_ref=? AND (received IS NULL OR received='')",
                (contract_ref,)
            ).fetchone()['remaining']

            if remaining != 0:
                decus.append(strike_value)

        if decus:
            dq_list = "with Decu Strike " + "; ".join('{:,.4f}'.format(x) for x in decus)
        else:
            dq_list = None

        ws[f"L{row_num}"].value = dq_list


def create_excel_report(positions, db):
    """Create Excel file with stock positions using template"""

    # Load the template
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'excel_templates', 'stock_summary.xlsx')
    wb = openpyxl.load_workbook(template_path)

    # Update date in ALL sheet
    wb["ALL"]["A1"].value = f'As of {ph_today().strftime("%B %d, %Y")}'

    # Populate Download sheet with data
    ws_download = wb['Download']
    download_row = 2

    for bank_id in sorted(positions.keys()):
        for code in sorted(positions[bank_id].keys()):
            pos = positions[bank_id][code]
            ccy = get_currency_from_code(code)

            # Write to Download sheet - this feeds the formulas in ALL sheet
            ws_download[f'A{download_row}'] = bank_id
            ws_download[f'B{download_row}'].value = f'=A{download_row}&C{download_row}'
            ws_download[f'C{download_row}'] = code
            ws_download[f'D{download_row}'] = f'{int(pos["shares"]):,} shares'
            ws_download[f'E{download_row}'] = f'{ccy} ={pos["total_cost"]}/{pos["shares"]}'
            ws_download[f'G{download_row}'] = pos['shares']
            ws_download[f'H{download_row}'] = round(pos['average_cost'], 4)

            download_row += 1

        download_row += 1  # Add spacing between banks

    # Update Stocks sheet with stock names
    ws_stocks = wb['Stocks']
    stocks_row = 1

    # Get all unique stock codes
    all_codes = set()
    for bank_id in positions:
        all_codes.update(positions[bank_id].keys())

    # Write stock code to name mapping
    for code in sorted(all_codes):
        stock_name_row = db.execute(
            "SELECT stock_name FROM tbl_code WHERE code = ?", (code,)
        ).fetchone()
        stock_name = stock_name_row['stock_name'] if stock_name_row else code

        ws_stocks[f'A{stocks_row}'] = code
        ws_stocks[f'B{stocks_row}'] = stock_name
        stocks_row += 1

    # Annotate the per-bank sheets with DECU strike-list info (port of legacy's
    # Stock_Balance.get_decu())
    for sheet_name in ("1", "2", "3", "4"):
        annotate_decu_strikes(db, wb, sheet_name)

    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    wb.close()

    return temp_file.name
