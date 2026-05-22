from flask import Blueprint, jsonify, send_file, current_app
from dataclasses import dataclass
from openpyxl import Workbook
import os

from .. auth import login_required
from .. database import get_db
from .. term_sheet import StockContract

bp = Blueprint('hkd_margin', __name__, template_folder='pages', url_prefix='/hkd-margin')


@bp.route('/')
@login_required
def home():
    db = get_db()
    hkd_margin = {}
    sql = '''SELECT tbl_stock_contract.ref_num 
        FROM tbl_stock_contract 
        INNER JOIN tbl_bank_account ON tbl_bank_account.ref_num = tbl_stock_contract.bank_ref
        WHERE tbl_stock_contract.status="active" 
        ORDER BY tbl_bank_account.priority, tbl_stock_contract.trade_date
        '''
    contract_refs = [row['ref_num'] for row in db.execute(sql).fetchall()]
    for contract_ref in contract_refs:
        ts = StockContract(
            db=db
        )
        ts.get(ref_num=contract_ref)
        ts.__post_init__()
        if ts.next_date == "DONE":
            continue  # Skip finished term sheet

        if ts.ccy_id not in hkd_margin:
            hkd_margin[ts.ccy_id] = {}

        if ts.bank_id not in hkd_margin[ts.ccy_id]:
            hkd_margin[ts.ccy_id][ts.bank_id] = {
                "ACCU": [],
                "DECU": []
            }

        hkd_margin[ts.ccy_id][ts.bank_id][ts.transaction_type].append(ts.as_dict())
    # Analyze records for single / double
    hkd_margin = single_double(db=db, hkd_margin=hkd_margin)

    # Generate excel file
    filename = os.path.join(current_app.instance_path, 'temp', 'hkd_margin.xlsx')
    HKDMarginToExcel(data=hkd_margin, filename=filename)

    return send_file('{}'.format(filename), as_attachment=True)


@dataclass
class HKDMarginToExcel:
    data: any
    filename: str = None
    accounts = ['CB1', 'CB2', 'CB3', 'CBBH', 'CBSG', 'DBPe', 'SHK', 'SHK2', 'MST2', 'BOS', 'DBPL', 'MST1', 'MSPL']

    def __post_init__(self):
        wb = Workbook()
        del wb['Sheet']

        self.create_sheets(wb)
        self.write_entries(wb)

        wb.save(self.filename)
        wb.close()

    def create_sheets(self, wb):
        for ccy in self.data:
            for _type in ('ACCU', 'DECU'):
                wb.create_sheet(f'{_type}-{ccy}')

    def write_entries(self, wb):
        for ccy in self.data:
            for _type in ('ACCU', 'DECU'):
                ws = wb[f'{_type}-{ccy}']
                row_num = 1
                operator = ">" if _type == "ACCU" else "<"
                for account in self.accounts:
                    if not self.data[ccy].get(account):
                        continue

                    term_sheets = self.data[ccy][account][_type]
                    if term_sheets:
                        # Write Header
                        ws[f'A{row_num}'].value = account
                        row_num += 1

                        cols = [
                            "",
                            _type.upper() + "MULATOR",
                            "CODE",
                            "Reference",
                            "SHARES / DAY",
                            "",
                            "",
                            "SPOT PRICE",
                            "STRIKE PRICE",
                            "K/O PRICE",
                            "RCVD DAYS",
                            "REM DAYS",
                            "Total DAYS",
                            "START DATE",
                            "END DATE",
                            "",
                            "Estimated Total Shares Remaining",
                            "Estimated Amount",
                            "",
                            "Estimated Amount per month",
                        ]
                        for col_num, value in enumerate(cols):
                            col_num += 1
                            cell = ws.cell(row=row_num, column=col_num)
                            cell.value = value

                        ws.merge_cells(f"E{row_num}:G{row_num}")

                        row_num += 1

                        # Write entries
                        for ts in term_sheets:
                            cols = [
                                '=INDIRECT("A"&row()-1)+1',
                                ts["stock_name"],
                                ts["code"],
                                ts["reference"],
                                ts["single"],
                                "/",
                                ts["double"],
                                ts["spot"],
                                ts["strike"],
                                ts["ko"],
                                ts["received_days"],
                                ts["remaining_days"],
                                ts["total_days"],
                                ts["start_date"],
                                ts["end_date"],
                                "",
                                f"=E{row_num}*L{row_num}*AD{row_num}",
                                f"=Q{row_num}*I{row_num}",
                                "",
                                f"=E{row_num}*I{row_num}*20*AD{row_num}",
                                "",
                                ts['price_1'],
                                ts['price_2'],
                                ts['price_3'],
                                "",
                                f"=IF(I{row_num}{operator}V{row_num},2,1)",
                                f"=IF(I{row_num}{operator}W{row_num},2,1)",
                                f"=IF(I{row_num}{operator}X{row_num},2,1)",
                                "",
                                f"=IF(E{row_num}=G{row_num},1,IF(SUM(Z{row_num}:AB{row_num})>4,2,1))"
                            ]

                            for col_num, value in enumerate(cols):
                                cell = ws.cell(row=row_num, column=col_num+1)
                                cell.value = value

                            row_num += 1

                        # Write totals

                        row_num += 2


def single_double(db, hkd_margin):
    def stock_price(code):
        sql = "SELECT p.closing_price FROM tbl_stock_price as p " \
              "INNER JOIN tbl_code as c ON c.ref_num = p.code_ref " \
              "WHERE c.code = ?" \
              "ORDER BY p.trade_date desc;"
        data = db.execute(sql, (code, )).fetchall()

        # Ensure we have at least 3 price records, fill with None if not available
        if len(data) >= 3:
            prices = (data[2][0], data[1][0], data[0][0])
        elif len(data) == 2:
            prices = (data[1][0], data[1][0], data[0][0])
        elif len(data) == 1:
            prices = (data[0][0], data[0][0], data[0][0])
        else:
            prices = (None, None, None)

        return prices

    for ccy in hkd_margin:
        for account in hkd_margin[ccy]:
            for _type in ('ACCU', 'DECU'):
                term_sheets = hkd_margin[ccy][account][_type]
                for ts in term_sheets:
                    code = ts['code']
                    ts['price_1'], ts['price_2'], ts['price_3'] = stock_price(code)

    return hkd_margin
