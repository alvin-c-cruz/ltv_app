from openpyxl import Workbook
import os
from dataclasses import dataclass, field


class CreateExcel:
    def __init__(self, path: str, report_date: str, db):
        self.filename = os.path.join(path, f"{report_date} portfolio.xlsx")
        self.wb = Workbook()

        self.data = GetData(db=db, report_date=report_date).position
        self.create()

        self.wb.save(self.filename)
        self.wb.close()

    def create(self):
        for bank_id, bank_data in self.data.items():
            self.wb.create_sheet(bank_id)
            ws = self.wb[bank_id]

            row_num = 1
            row_num = self.write_header(ws, bank_data, row_num)
            row_num = self.write_details(ws, bank_data, row_num)
            row_num += 2

    def write_header(self, ws, bank_data, row_num):
        cell = ws[f"A{row_num}"]
        cell.value = bank_data["bank_name"][0]

        row_num += 1

        cols = {
            "A": "Stock Name",
            "B": "Code",
            "C": "Shares",
            "D": "Cost-to-date",
            "E": "Average"
        }

        for column_letter, value in cols.items():
            cell = ws[f"{column_letter}{row_num}"]
            cell.value = value

        row_num += 1
        return row_num

    def write_details(self, ws, bank_data, row_num):
        for code, details in bank_data["position"].items():
            if not details["shares"]:
                continue

            cols = {
                "A": details["stock_name"],
                "B": code,
                "C": details["shares"],
                "D": details["cost_to_date"],
                "E": f"=D{row_num}/C{row_num}"
            }

            for column_letter, value in cols.items():
                cell = ws[f"{column_letter}{row_num}"]
                cell.value = value

            row_num += 1

        return row_num


@dataclass
class GetData:
    db: any
    report_date: str

    def __post_init__(self):
        self.position = {}

        self.gather_data()

    def gather_data(self):
        for account in self.bank_accounts():
            bank_id = account["bank_id"]
            bank_name = account["bank_name"]
            bank_ref = account["ref_num"]
            transaction_basis = account["transaction_basis"]

            self.position[bank_id] = {}
            self.position[bank_id]["bank_name"] = bank_name,
            self.position[bank_id]["position"] = {}
            self.get_position(
                bank_id,
                self.get_transactions(bank_ref, transaction_basis)
            )

    def bank_accounts(self):
        sql = "SELECT ref_num, bank_id, bank_name, transaction_basis FROM tbl_bank_account ORDER BY priority;"
        data = self.db.execute(sql).fetchall()
        return data

    def get_transactions(self, bank_ref, transaction_basis):
        sql = f"SELECT " \
              f"    tbl_code.code, " \
              f"    tbl_code.stock_name, " \
              f"    tbl_transaction.transaction_type, " \
              f"    tbl_transaction.quantity, " \
              f"    tbl_transaction.price, " \
              f"    tbl_transaction.brokerage, " \
              f"    tbl_transaction.commission, " \
              f"    tbl_transaction.foreign_charge, " \
              f"    tbl_transaction.stamp_duty, " \
              f"    tbl_transaction.misc " \
              f"FROM tbl_transaction " \
              f"INNER JOIN tbl_code ON tbl_code.ref_num = tbl_transaction.code_ref " \
              f"INNER JOIN tbl_transaction_type ON tbl_transaction_type.transaction_type" \
              f"     = tbl_transaction.transaction_type " \
              f"WHERE tbl_transaction.{transaction_basis} <= ? AND tbl_transaction.bank_ref = ?" \
              f"ORDER BY " \
              f"    tbl_code.code, " \
              f"    tbl_transaction.{transaction_basis}, " \
              f"    tbl_transaction_type.priority " \
              f";"

        data = self.db.execute(sql, (self.report_date, bank_ref)).fetchall()
        return data

    def get_position(self, bank_id, transactions):
        for transaction in transactions:
            code = transaction["code"]
            stock_name = transaction["stock_name"]

            if code not in self.position[bank_id]["position"]:
                self.position[bank_id]["position"][code] = {
                    "stock_name": stock_name,
                    "shares": 0,
                    "cost_to_date": 0
                }

            balance = self.position[bank_id]["position"][code]["shares"]
            cost_to_date = self.position[bank_id]["position"][code]["cost_to_date"]

            quantity = transaction['quantity']
            price = transaction['price']
            charges = transaction['brokerage'] + transaction['commission'] \
                      + transaction['foreign_charge'] + transaction['stamp_duty'] + transaction['misc']
            amount = quantity * price + charges

            if quantity > 0:
                if balance > 0:
                    cost_to_date += amount
                elif balance == 0:
                    cost_to_date += amount
                elif balance < 0:
                    if balance + quantity == 0:
                        cost_to_date = 0
                    elif balance + quantity < 0:
                        cost_to_date = 0
                    else:
                        cost_to_date = (balance + quantity) / quantity * amount

            else:
                if balance > 0:
                    if balance - abs(quantity) > 0:
                        cost_to_date -= cost_to_date * abs(quantity) / balance
                    else:
                        cost_to_date = 0
                else:
                    cost_to_date = 0

            balance += quantity

            self.position[bank_id]["position"][code]["shares"] = balance
            self.position[bank_id]["position"][code]["cost_to_date"] = cost_to_date
