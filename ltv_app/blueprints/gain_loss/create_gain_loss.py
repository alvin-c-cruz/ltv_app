import io
import os
import openpyxl
from datetime import datetime, timedelta
from flask import current_app, g, url_for

from .. database import get_db
from .. transactions import get_balance, get_transactions, get_short_balance, get_short_transactions
from .. transactions.models import (
    accumulate_position, accumulate_short_position,
    restructure_transactions, restructure_short_transactions,
)


def _refs():
    """Reference tables, loaded once per request.

    These are tiny (16 banks, 94 codes, 5 currencies) but were previously
    re-queried per use -- get_code/get_stock_name alone fired once per written
    workbook row, which is most of this report's query count.
    """
    if 'gain_loss_refs' not in g:
        db = get_db()
        # dict preserves insertion order, so 'ORDER BY priority' is retained
        # by the *_order lists below.
        banks = {r['ref_num']: r for r in
                 db.execute("SELECT * FROM tbl_bank_account ORDER BY priority;").fetchall()}
        codes = {r['ref_num']: r for r in
                 db.execute("SELECT * FROM tbl_code;").fetchall()}
        currencies = db.execute("SELECT * FROM tbl_currency ORDER BY priority;").fetchall()

        g.gain_loss_refs = {
            'banks': banks,
            'codes': codes,
            'bank_ref_by_id': {r['bank_id']: ref for ref, r in banks.items()},
            'ccy_id_by_ref': {r['ref_num']: r['ccy_id'] for r in currencies},
            'ccy_ref_by_id': {r['ccy_id']: r['ref_num'] for r in currencies},
            'ccy_order': [r['ccy_id'] for r in currencies],
            'bank_id_order': [r['bank_id'] for r in banks.values()],
            'code_refs': list(codes.keys()),
        }
    return g.gain_loss_refs


def list_currency():
    return list(_refs()['ccy_order'])


def list_bank_accounts():
    return list(_refs()['bank_id_order'])


def bank_ref_nums():
    return list(_refs()['banks'].keys())


def get_bank_ref(bank_id):
    return _refs()['bank_ref_by_id'][bank_id]


def get_bank_id(bank_ref):
    return _refs()['banks'][bank_ref]['bank_id']


def get_bank_name(bank_ref):
    return _refs()['banks'][bank_ref]['bank_name']


def get_code(code_ref):
    return _refs()['codes'][code_ref]['code']


def get_stock_name(code_ref):
    return _refs()['codes'][code_ref]['stock_name']


def get_ccy_ref(ccy_id):
    return _refs()['ccy_ref_by_id'][ccy_id]


def get_ccy(code_ref):
    refs = _refs()
    return refs['ccy_id_by_ref'][refs['codes'][code_ref]['ccy_ref']]


def get_stock_ref():
    return list(_refs()['code_refs'])


def create_file(date_from, date_to, bank_ref, code_ref):

    wb = openpyxl.load_workbook(os.path.join(current_app.instance_path, 'excel_templates', 'gain_loss.xlsx'))

    wb['SUMMARY']['A3'] = f'As of {datetime.strftime(datetime.strptime(date_to, "%Y-%m-%d"), "%B %d, %Y")}'

    head_line = {}
    short_head_line = {}

    if date_from[:4] == date_to[:4]:
        if date_from[5:7] == date_to[5:7]:
            if date_from[-2:] == date_to[-2:]:
                period_covered = f'=TEXT("{date_from}","mmmm  d, yyyy")'
            else:
                period_covered = f'="From "&TEXT("{date_from}","mmmm  d")&" to "&TEXT("{date_to}","d, yyyy")'
        else:
            period_covered = f'="From "&TEXT("{date_from}","mmmm  d")&" to "&TEXT("{date_to}","mmmm  d, yyyy")'
    else:
        period_covered = f'="From "&TEXT("{date_from}","mmmm  d, yyyy")&" to "&TEXT("{date_to}","mmmm  d, yyyy")'

    period_covered = '=INDIRECT("SUMMARY!A3")'

    year = int(date_from[:4])
    month = date_from[5:7]
    prev_month = {
        "01": ("December", year - 1),
        "02": ("January", year),
        "03": ("February", year),
        "04": ("March", year),
        "05": ("April", year),
        "06": ("May", year),
        "07": ("June", year),
        "08": ("July", year),
        "09": ("August", year),
        "10": ("September", year),
        "11": ("October", year),
        "12": ("November", year),
    }

    end_of = f'End of {prev_month[month][0]} {prev_month[month][1]}'

    download_name, bank_refs, code_refs = _get_refs_and_name(date_from, date_to, bank_ref, code_ref)
    dict_gain_loss = gather_gain_loss(date_from, date_to, bank_refs, code_refs)

    for ccy in list_currency():
        for bank_id in list_bank_accounts():
            if ccy in dict_gain_loss:
                if get_bank_ref(bank_id) not in dict_gain_loss[ccy]:
                    if f'{bank_id}-{ccy}' in wb.sheetnames:
                        ws = wb[f'{bank_id}-{ccy}']
                        ws.sheet_state = 'hidden'
            else:
                if f'{bank_id}-{ccy}' in wb.sheetnames:
                    ws = wb[f'{bank_id}-{ccy}']
                    ws.sheet_state = 'hidden'

    for ccy in dict_gain_loss:
        head_line[ccy] = {}

        for bank_ref in dict_gain_loss[ccy]:
            head_line[ccy][bank_ref] = {}
            short_head_line.setdefault(ccy, {})[bank_ref] = {}
            bank_id = get_bank_id(bank_ref)
            bank_name = get_bank_name(bank_ref)
            sheet_name = f'{bank_id}-{ccy}'
            if sheet_name not in wb.sheetnames: wb.create_sheet(sheet_name)
            ws = wb[sheet_name]

            # Headers

            write_title(ws, 'A1', 'Avg. Price vs Selling Price', 12, 'left')
            write_title(ws, 'A2', period_covered, 12, 'left')
            write_title(ws, 'A4', bank_name, 16, 'left')

            for cell_name in ("A4", "B4", "C4"):
                fill_color(ws, cell_name, "00FFFF00")

            rate = {
                "HKD": "=hkd_per_usd",
                "JPY": "=jpy_per_usd",
                "AUD": "=aud_per_usd",
                "USD": 1,
                "SGD": "=sgd_per_usd",
            }

            write_title(ws, 'D6', 'USD RATE', 11, 'left')
            write(ws, 'E6', rate[ccy], 11, 'right', 'currency_2', True)

            write_title(ws, 'D8', end_of, 11, 'left')
            ws.merge_cells('D8:G8')
            write_title(ws, 'D9', 'Average Price', 11, 'left')
            ws.merge_cells('D9:G9')

            # Beginning Balance Headers
            write_title(ws, 'A10', '', 11, 'center', True)
            write_title(ws, 'B10', 'Stock', 11, 'center', True)
            ws.merge_cells('B10:C10')
            write_title(ws, 'D10', 'Code', 11, 'center', True)
            write_title(ws, 'E10', 'Quantity', 11, 'center', True)
            write_title(ws, 'F10', 'Avg Price', 11, 'center', True)
            write_title(ws, 'G10', 'Value', 11, 'center', True)
            write_title(ws, 'H10', 'USD  Value', 11, 'center', True)
            ws.merge_cells('H10:I10')

            write_title(ws, 'K9', 'Gain Loss', 11, 'center', True)
            ws.merge_cells('K9:M9')
            write_title(ws, 'K10', 'Net', 11, 'center', True)
            write_title(ws, 'L10', 'Stock', 11, 'center', True)
            ws.merge_cells('L10:M10')

            # Beginning Balance Details (long + short interleaved)
            row_num = 11
            start_total = row_num

            long_beg = dict_gain_loss[ccy][bank_ref]['beginning']
            short_beg = dict_gain_loss[ccy][bank_ref].get('short_beginning', {})

            # Collect all code_refs from both books, sorted by stock code
            all_code_refs = set(long_beg.keys()) | set(short_beg.keys())
            ordered_code_refs = sorted(all_code_refs, key=lambda cr: get_code(cr))

            for code_ref in ordered_code_refs:
                code = get_code(code_ref)
                stock_name = get_stock_name(code_ref)

                if code_ref in long_beg:
                    head_line[ccy][bank_ref][code_ref] = row_num
                    beginning_balance = long_beg[code_ref]['quantity']
                    average = long_beg[code_ref]['average']

                    write_box_normal(ws, f'A{row_num}', f'=INDIRECT("A"&row()-1)+1', "integer")
                    write_box_normal(ws, f'B{row_num}', stock_name)
                    ws.merge_cells(f'B{row_num}:C{row_num}')
                    write_box_normal(ws, f'D{row_num}', code, "code")
                    write_box_normal(ws, f'E{row_num}', beginning_balance, "shares")
                    write_box_normal(ws, f'F{row_num}', average, "currency_4")
                    write_box_normal(ws, f'G{row_num}', f'=E{row_num}*F{row_num}', "currency_2")
                    write_box_normal(ws, f'H{row_num}', f'=G{row_num}/hkd_per_usd', "currency_2")
                    ws.merge_cells(f'H{row_num}:I{row_num}')
                    write_box_normal(ws, f'K{row_num}', '')
                    write_box_normal(ws, f'L{row_num}', f'=B{row_num}')
                    ws.merge_cells(f'L{row_num}:M{row_num}')
                    row_num += 1

                if code_ref in short_beg:
                    short_head_line[ccy][bank_ref][code_ref] = row_num
                    s_qty = short_beg[code_ref]['quantity']
                    s_avg = short_beg[code_ref]['average']

                    write_box_normal(ws, f'A{row_num}', f'=INDIRECT("A"&row()-1)+1', "integer")
                    write_box_normal(ws, f'B{row_num}', stock_name + ' - Short')
                    ws.merge_cells(f'B{row_num}:C{row_num}')
                    write_box_normal(ws, f'D{row_num}', code, "code")
                    write_box_normal(ws, f'E{row_num}', s_qty, "shares")
                    write_box_normal(ws, f'F{row_num}', s_avg, "currency_4")
                    write_box_normal(ws, f'G{row_num}', f'=E{row_num}*F{row_num}', "currency_2")
                    write_box_normal(ws, f'H{row_num}', f'=G{row_num}/hkd_per_usd', "currency_2")
                    ws.merge_cells(f'H{row_num}:I{row_num}')
                    write_box_normal(ws, f'K{row_num}', 0, "currency_2")
                    write_box_normal(ws, f'L{row_num}', f'=B{row_num}')
                    ws.merge_cells(f'L{row_num}:M{row_num}')
                    row_num += 1

            end_total = row_num - 1

            # Total Line
            write_box_normal(ws, f'A{row_num}', '')
            write_box_normal(ws, f'B{row_num}', '')
            ws.merge_cells(f'B{row_num}:C{row_num}')
            write_box_normal(ws, f'D{row_num}', 'TOTAL', 'string', True)
            write_box_normal(ws, f'E{row_num}', f'=SUM(E{start_total}:E{end_total})', 'shares', True)
            write_box_normal(ws, f'F{row_num}', '')
            write_box_normal(ws, f'G{row_num}', f'=SUM(G{start_total}:G{end_total})', 'currency_2', True)
            write_box_normal(ws, f'H{row_num}', f'=SUM(H{start_total}:H{end_total})', 'currency_2', True)
            ws.merge_cells(f'H{row_num}:I{row_num}')

            ws[f'G{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
            ws[f'H{row_num}'].number_format = f'[$USD] #,##0.00_);([$USD] #,##0.00)'

            write_box_normal(ws, f'K{row_num}', f'=SUM(K{start_total}:K{end_total})', 'currency_2', True)
            write_box_normal(ws, f'L{row_num}', 'TOTAL')
            ws.merge_cells(f'L{row_num}:M{row_num}')

            ws[f'K{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'

            row_num += 2

            # Detail sections: one pass over all code_refs sorted by stock code,
            # writing the long section then the short section for each.
            all_detail_code_refs = set(dict_gain_loss[ccy][bank_ref]['trades']) | \
                                   set(dict_gain_loss[ccy][bank_ref].get('short_trades', {}))
            for code_ref in sorted(all_detail_code_refs, key=lambda cr: get_code(cr)):
                code = get_code(code_ref)
                name = get_stock_name(code_ref)
                trades      = dict_gain_loss[ccy][bank_ref]['trades'].get(code_ref, [])
                short_trades = dict_gain_loss[ccy][bank_ref].get('short_trades', {}).get(code_ref, [])

                # ── Long detail section ──────────────────────────────────────────
                if trades:
                    stock_name = f'{code}:{ccy[:2]} - {name}'
                    write_title(ws, f'A{row_num}', stock_name, 12, 'left')
                    for cell_name in (f'A{row_num}', f'B{row_num}', f'C{row_num}'):
                        fill_color(ws, cell_name, "00FFFF00")

                    head_row_num = head_line[ccy][bank_ref][code_ref]
                    write(ws, f'E{row_num}', f'=E{head_row_num}', 11, 'right', 'shares', True)
                    write(ws, f'F{row_num}', f'=F{head_row_num}', 11, 'right', 'currency_4', True)
                    start_balance = row_num
                    ws[f"O{row_num}"].value = f"=E{row_num}"
                    row_num += 1

                    write_title(ws, f'A{row_num}', '', 11, 'center', True)
                    write_title(ws, f'B{row_num}', 'Trade  Date', 11, 'center', True)
                    write_title(ws, f'C{row_num}', 'Settlement Date', 11, 'center', True)
                    write_title(ws, f'D{row_num}', 'Transaction', 11, 'center', True)
                    write_title(ws, f'E{row_num}', 'Shares', 11, 'center', True)
                    write_title(ws, f'F{row_num}', 'Price', 11, 'center', True)
                    write_title(ws, f'G{row_num}', 'Gross Total', 11, 'center', True)
                    write_title(ws, f'H{row_num}', 'Charges', 11, 'center', True)
                    write_title(ws, f'I{row_num}', 'Net Total', 11, 'center', True)
                    write_title(ws, f'J{row_num}', 'Avg. Price', 11, 'center', True)
                    write_title(ws, f'K{row_num}', 'Avg. Cost', 11, 'center', True)
                    write_title(ws, f'L{row_num}', 'Net Gain Loss', 11, 'center', True)
                    write_title(ws, f'M{row_num}', '%', 11, 'center', True)
                    ws[f"O{row_num}"].value = f"=O{row_num-1}"
                    ws[f"P{row_num}"].value = f"=E{start_balance}"
                    ws[f"Q{row_num}"].value = f"=R{row_num}/P{row_num}"
                    ws[f"R{row_num}"].value = f"=E{start_balance}*F{start_balance}"
                    row_num += 1
                    g_l = False

                    for record in trades:
                        write_box_normal(ws, f'A{row_num}', f'=INDIRECT("A"&row()-1)+1', "integer")
                        write_box_normal(ws, f'B{row_num}', datetime.strptime(record['trade_date'], "%Y-%m-%d"), "date")
                        write_box_normal(ws, f'C{row_num}', datetime.strptime(record['value_date'], "%Y-%m-%d"), "date")
                        write_box_normal(ws, f'D{row_num}', record['description'])
                        write_box_normal(ws, f'E{row_num}', record['quantity'], "shares")
                        write_box_normal(ws, f'F{row_num}', record['price'], "currency_4")
                        write_box_normal(ws, f'G{row_num}', f'=E{row_num}*F{row_num}', "currency_2")
                        write_box_normal(ws, f'H{row_num}', record['charges'], "currency_2")
                        write_box_normal(ws, f'I{row_num}', f'=G{row_num}+H{row_num}', "currency_2")
                        ws[f"P{row_num}"].value = f"=P{row_num-1}+E{row_num}"
                        ws[f"Q{row_num}"].value = f"=IFERROR(R{row_num}/P{row_num},Q{row_num-1})"
                        ws[f"R{row_num}"].value = f"=IF(E{row_num}>0,R{row_num-1}+I{row_num},R{row_num-1}+Q{row_num-1}*E{row_num})"
                        write_box_normal(ws, f'J{row_num}', f'=Q{row_num}', "currency_4")

                        if record['description'][:4] == 'Sell':
                            write_box_normal(ws, f'K{row_num}', f'=E{row_num}*J{row_num}', "currency_2")
                            write_box_normal(ws, f'L{row_num}', f'=K{row_num}-I{row_num}', "currency_2")
                            write_box_normal(ws, f'M{row_num}', f'=(I{row_num}/K{row_num})-1', "percentage")
                            g_l = True
                        else:
                            write_box_normal(ws, f'K{row_num}', '', "currency_2")
                            write_box_normal(ws, f'L{row_num}', '', "currency_2")
                            write_box_normal(ws, f'M{row_num}', '', "percentage")

                        ws[f"O{row_num}"].value = f'=INDIRECT("O"&row()-1)+E{row_num}'
                        ws[f"O{row_num}"].number_format = "#,##0;[RED]-#,##0"
                        cell = ws[f"D{row_num}"]
                        if 'Accu' in record['description'] or 'Decu' in record['description']:
                            cell.hyperlink = url_for('fixings.edit', ref_num=record['ref_num'], _external=True)
                        else:
                            cell.hyperlink = url_for('transactions.edit', ref_num=record['ref_num'], _external=True)
                        row_num += 1

                    write_box_normal(ws, f'A{row_num}', "")
                    write_box_normal(ws, f'B{row_num}', "")
                    write_box_normal(ws, f'C{row_num}', "")
                    write_box_normal(ws, f'D{row_num}', "TOTAL", "string", True)
                    write_box_normal(ws, f'E{row_num}', f'=SUM(E{start_balance}:E{row_num - 1})', "shares", True)
                    write_box_normal(ws, f'F{row_num}', "")
                    write_box_normal(ws, f'G{row_num}', f'=SUM(G{start_balance}:G{row_num - 1})', "currency_2", True)
                    ws[f'G{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
                    write_box_normal(ws, f'H{row_num}', "")
                    write_box_normal(ws, f'I{row_num}', f'=SUM(I{start_balance}:I{row_num - 1})', "currency_2", True)
                    ws[f'I{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
                    write_box_normal(ws, f'J{row_num}', "")
                    write_box_normal(ws, f'K{row_num}', f'=SUM(K{start_balance}:K{row_num - 1})', "currency_2", True)
                    ws[f'K{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
                    write_box_normal(ws, f'L{row_num}', f'=SUM(L{start_balance}:L{row_num - 1})', "currency_2", True)
                    ws[f'L{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
                    if g_l:
                        fill_color(ws, f'L{row_num}', "00FF9900")
                    write_box_normal(ws, f'M{row_num}', "")
                    write_box_normal(ws, f'K{head_row_num}', f'=L{row_num}', "currency_2")
                    row_num += 2

                # ── Short detail section ─────────────────────────────────────────
                if short_trades:
                    stock_name = f'{code}:{ccy[:2]} - {name} - Short'
                    write_title(ws, f'A{row_num}', stock_name, 12, 'left')
                    for cell_name in (f'A{row_num}', f'B{row_num}', f'C{row_num}'):
                        fill_color(ws, cell_name, "00FFFF00")

                    short_head_row_num = short_head_line.get(ccy, {}).get(bank_ref, {}).get(code_ref)
                    write(ws, f'E{row_num}', f'=E{short_head_row_num}', 11, 'right', 'shares', True)
                    write(ws, f'F{row_num}', f'=F{short_head_row_num}', 11, 'right', 'currency_4', True)
                    start_balance = row_num
                    ws[f"O{row_num}"].value = f"=E{row_num}"
                    row_num += 1

                    write_title(ws, f'A{row_num}', '', 11, 'center', True)
                    write_title(ws, f'B{row_num}', 'Trade  Date', 11, 'center', True)
                    write_title(ws, f'C{row_num}', 'Settlement Date', 11, 'center', True)
                    write_title(ws, f'D{row_num}', 'Transaction', 11, 'center', True)
                    write_title(ws, f'E{row_num}', 'Shares', 11, 'center', True)
                    write_title(ws, f'F{row_num}', 'Price', 11, 'center', True)
                    write_title(ws, f'G{row_num}', 'Gross Total', 11, 'center', True)
                    write_title(ws, f'H{row_num}', 'Charges', 11, 'center', True)
                    write_title(ws, f'I{row_num}', 'Net Total', 11, 'center', True)
                    write_title(ws, f'J{row_num}', 'Avg. Price', 11, 'center', True)
                    write_title(ws, f'K{row_num}', 'Avg. Cost', 11, 'center', True)
                    write_title(ws, f'L{row_num}', 'Net Gain Loss', 11, 'center', True)
                    write_title(ws, f'M{row_num}', '%', 11, 'center', True)
                    ws[f"O{row_num}"].value = f"=O{row_num-1}"
                    ws[f"P{row_num}"].value = f"=E{start_balance}"
                    ws[f"Q{row_num}"].value = f"=R{row_num}/P{row_num}"
                    ws[f"R{row_num}"].value = f"=E{start_balance}*F{start_balance}"
                    row_num += 1
                    g_l = False

                    for record in short_trades:
                        qty = record['quantity']
                        write_box_normal(ws, f'A{row_num}', f'=INDIRECT("A"&row()-1)+1', "integer")
                        write_box_normal(ws, f'B{row_num}', datetime.strptime(record['trade_date'], "%Y-%m-%d"), "date")
                        write_box_normal(ws, f'C{row_num}', datetime.strptime(record['value_date'], "%Y-%m-%d"), "date")
                        write_box_normal(ws, f'D{row_num}', record['description'])
                        write_box_normal(ws, f'E{row_num}', record['quantity'], "shares")
                        write_box_normal(ws, f'F{row_num}', record['price'], "currency_4")
                        write_box_normal(ws, f'G{row_num}', f'=E{row_num}*F{row_num}', "currency_2")
                        write_box_normal(ws, f'H{row_num}', record['charges'], "currency_2")
                        write_box_normal(ws, f'I{row_num}', f'=G{row_num}+H{row_num}', "currency_2")
                        ws[f"P{row_num}"].value = f"=P{row_num-1}+E{row_num}"
                        ws[f"Q{row_num}"].value = f"=IFERROR(R{row_num}/P{row_num},Q{row_num-1})"
                        ws[f"R{row_num}"].value = f"=IF(E{row_num}<0,R{row_num-1}+I{row_num},R{row_num-1}+Q{row_num-1}*E{row_num})"
                        write_box_normal(ws, f'J{row_num}', f'=Q{row_num}', "currency_4")

                        if qty > 0:
                            write_box_normal(ws, f'K{row_num}', f'=E{row_num}*J{row_num}', "currency_2")
                            write_box_normal(ws, f'L{row_num}', f'=K{row_num}-I{row_num}', "currency_2")
                            write_box_normal(ws, f'M{row_num}', f'=(I{row_num}/K{row_num})-1', "percentage")
                            g_l = True
                        else:
                            write_box_normal(ws, f'K{row_num}', '', "currency_2")
                            write_box_normal(ws, f'L{row_num}', '', "currency_2")
                            write_box_normal(ws, f'M{row_num}', '', "percentage")

                        ws[f"O{row_num}"].value = f'=INDIRECT("O"&row()-1)+E{row_num}'
                        ws[f"O{row_num}"].number_format = "#,##0;[RED]-#,##0"
                        ws[f"D{row_num}"].hyperlink = url_for('transactions.edit_short', ref_num=record['ref_num'], _external=True)
                        row_num += 1

                    write_box_normal(ws, f'A{row_num}', "")
                    write_box_normal(ws, f'B{row_num}', "")
                    write_box_normal(ws, f'C{row_num}', "")
                    write_box_normal(ws, f'D{row_num}', "TOTAL", "string", True)
                    write_box_normal(ws, f'E{row_num}', f'=SUM(E{start_balance}:E{row_num - 1})', "shares", True)
                    write_box_normal(ws, f'F{row_num}', "")
                    write_box_normal(ws, f'G{row_num}', f'=SUM(G{start_balance}:G{row_num - 1})', "currency_2", True)
                    ws[f'G{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
                    write_box_normal(ws, f'H{row_num}', "")
                    write_box_normal(ws, f'I{row_num}', f'=SUM(I{start_balance}:I{row_num - 1})', "currency_2", True)
                    ws[f'I{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
                    write_box_normal(ws, f'J{row_num}', "")
                    write_box_normal(ws, f'K{row_num}', f'=SUM(K{start_balance}:K{row_num - 1})', "currency_2", True)
                    ws[f'K{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
                    write_box_normal(ws, f'L{row_num}', f'=SUM(L{start_balance}:L{row_num - 1})', "currency_2", True)
                    ws[f'L{row_num}'].number_format = f'[${ccy}] #,##0.00_);([${ccy}] #,##0.00)'
                    if g_l:
                        fill_color(ws, f'L{row_num}', "00FF9900")
                    write_box_normal(ws, f'M{row_num}', "")
                    if short_head_row_num:
                        write_box_normal(ws, f'K{short_head_row_num}', f'=L{row_num}', "currency_2")
                    row_num += 2

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    return output, download_name


def create_file_name(date_from, date_to, bank_ref, code_ref):
    if bank_ref == 0:
        bank_refs = bank_ref_nums()
        if code_ref == 0:
            code_refs = get_stock_ref()
            filename = os.path.join(current_app.instance_path, 'temp', f'{date_from}-to-{date_to}-gain_loss.xlsx')
        else:
            code_refs = [code_ref]
            code = get_code(code_ref)
            filename = os.path.join(current_app.instance_path, 'temp', f'{date_from}-to-{date_to}-gain_loss-{code}.xlsx')
    else:
        bank_refs = [bank_ref]
        bank_id = get_bank_id(bank_ref)

        if code_ref == 0:
            code_refs = get_stock_ref()
            filename = os.path.join(current_app.instance_path, 'temp',
                                    f'{date_from}-to-{date_to}-gain_loss-{bank_id}.xlsx')
        else:
            code_refs = [code_ref]
            code = get_code(code_ref)
            filename = os.path.join(current_app.instance_path, 'temp',
                                    f'{date_from}-to-{date_to}-gain_loss-{bank_id}-{code}.xlsx')

    return filename, bank_refs, code_refs


def _get_refs_and_name(date_from, date_to, bank_ref, code_ref):
    if bank_ref == 0:
        bank_refs = bank_ref_nums()
        if code_ref == 0:
            code_refs = get_stock_ref()
            name = f'{date_from}-to-{date_to}-gain_loss.xlsx'
        else:
            code_refs = [code_ref]
            name = f'{date_from}-to-{date_to}-gain_loss-{get_code(code_ref)}.xlsx'
    else:
        bank_refs = [bank_ref]
        bank_id = get_bank_id(bank_ref)
        if code_ref == 0:
            code_refs = get_stock_ref()
            name = f'{date_from}-to-{date_to}-gain_loss-{bank_id}.xlsx'
        else:
            code_refs = [code_ref]
            name = f'{date_from}-to-{date_to}-gain_loss-{bank_id}-{get_code(code_ref)}.xlsx'
    return name, bank_refs, code_refs


# _dict = {
#     "ccy": {
#         "int(bank_ref)": {
#             "beginning": {
#                 "int(code_ref)": {
#                     "quantity": 0,
#                     "cost_to_date": 0,
#                     "average": 0
#                 }
#             },
#             "trades": {
#                 "int(code_ref)": [
#                     {"SELECT * FROM tbl_transaction"}
#                 ]
#             }
#         }
#     }
# }
def _banks_by_basis(bank_refs):
    """Group the requested banks by transaction_basis.

    The date column to filter and sort on differs per bank ('trade_date' vs
    'value_date'), so batched queries are issued one per distinct basis --
    two, in practice -- rather than one per bank/code pair.
    """
    refs = _refs()
    by_basis = {}
    for bank_ref in bank_refs:
        basis = refs['banks'][bank_ref]['transaction_basis']
        by_basis.setdefault(basis, []).append(bank_ref)
    return by_basis


def _grouped_rows(db, table, columns, bank_refs, code_refs, date_to, order_by_priority):
    """All rows of `table` up to date_to, grouped by (bank_ref, code_ref).

    Each group comes back in exactly the order the per-pair queries produced:
    ordered by the bank's transaction_basis (then transaction-type priority for
    the long book). Rows are fetched with no lower date bound because the
    opening balance replays the position from inception.
    """
    grouped = {}
    for basis, basis_banks in _banks_by_basis(bank_refs).items():
        bank_marks = ",".join("?" * len(basis_banks))
        code_marks = ",".join("?" * len(code_refs))

        if order_by_priority:
            sql = (f"SELECT {columns}, {table}.{basis} AS _basis "
                   f"FROM {table} "
                   "INNER JOIN tbl_transaction_type "
                   f"ON tbl_transaction_type.transaction_type = {table}.transaction_type "
                   f"WHERE {table}.bank_ref IN ({bank_marks}) "
                   f"AND {table}.code_ref IN ({code_marks}) "
                   f"AND {table}.{basis}<=? "
                   f"ORDER BY {table}.bank_ref, {table}.code_ref, "
                   f"{table}.{basis}, tbl_transaction_type.priority;")
        else:
            sql = (f"SELECT {columns}, {table}.{basis} AS _basis "
                   f"FROM {table} "
                   f"WHERE {table}.bank_ref IN ({bank_marks}) "
                   f"AND {table}.code_ref IN ({code_marks}) "
                   f"AND {table}.{basis}<=? "
                   f"ORDER BY {table}.bank_ref, {table}.code_ref, {table}.{basis};")

        params = (*basis_banks, *code_refs, date_to)
        for row in db.execute(sql, params).fetchall():
            grouped.setdefault((row['bank_ref'], row['code_ref']), []).append(row)

    return grouped


def gather_gain_loss(date_from, date_to, bank_refs, code_refs):
    db = get_db()
    refs = _refs()
    dict_gain_loss = {}

    beginning_date = str(datetime.strptime(date_from, "%Y-%m-%d") - timedelta(days=1))[:10]

    def _ensure(ccy, bank_ref):
        if ccy not in dict_gain_loss:
            dict_gain_loss[ccy] = {}
        if bank_ref not in dict_gain_loss[ccy]:
            dict_gain_loss[ccy][bank_ref] = {
                "beginning": {}, "trades": {},
                "short_beginning": {}, "short_trades": {},
            }

    def bank_id_of(ref_num):
        return refs['banks'][ref_num]['bank_id']

    # Two bulk queries per book (one per transaction_basis) replace the
    # 4 x len(bank_refs) x len(code_refs) per-pair queries this used to issue.
    long_rows = _grouped_rows(db, "tbl_transaction", "tbl_transaction.*",
                              bank_refs, code_refs, date_to, order_by_priority=True)
    short_rows = _grouped_rows(
        db, "tbl_transaction_short",
        "tbl_transaction_short.ref_num, tbl_transaction_short.bank_ref, "
        "tbl_transaction_short.code_ref, tbl_transaction_short.trade_date, "
        "tbl_transaction_short.value_date, tbl_transaction_short.transaction_type, "
        "tbl_transaction_short.quantity, tbl_transaction_short.price, "
        "tbl_transaction_short.brokerage, tbl_transaction_short.commission, "
        "tbl_transaction_short.foreign_charge, tbl_transaction_short.stamp_duty, "
        "tbl_transaction_short.misc",
        bank_refs, code_refs, date_to, order_by_priority=False)

    for bank_ref in bank_refs:
        for code_ref in code_refs:
            long_group = long_rows.get((bank_ref, code_ref), ())
            short_group = short_rows.get((bank_ref, code_ref), ())
            if not long_group and not short_group:
                continue

            #  1. Long book balance
            opening = [r for r in long_group if r['_basis'] <= beginning_date]
            quantity, cost_to_date, _ = accumulate_position(opening)
            if quantity:
                ccy = get_ccy(code_ref)
                _ensure(ccy, bank_ref)
                dict_gain_loss[ccy][bank_ref]["beginning"][code_ref] = {
                    "quantity": quantity,
                    "cost_to_date": cost_to_date,
                    "average": cost_to_date / quantity,
                }

            #  2. Long book transactions
            period = [r for r in long_group if date_from <= r['_basis'] <= date_to]
            transactions = restructure_transactions(period, bank_id_of) if period else []
            if transactions:
                ccy = get_ccy(code_ref)
                _ensure(ccy, bank_ref)
                if code_ref not in dict_gain_loss[ccy][bank_ref]["beginning"]:
                    dict_gain_loss[ccy][bank_ref]["beginning"][code_ref] = {
                        "quantity": 0, "cost_to_date": 0, "average": 0,
                    }
                dict_gain_loss[ccy][bank_ref]["trades"][code_ref] = transactions

            #  3. Short book balance
            short_opening = [r for r in short_group if r['_basis'] <= beginning_date]
            short_qty, short_cost = accumulate_short_position(short_opening)
            if short_qty:
                ccy = get_ccy(code_ref)
                _ensure(ccy, bank_ref)
                dict_gain_loss[ccy][bank_ref]["short_beginning"][code_ref] = {
                    "quantity": short_qty,
                    "cost_to_date": short_cost,
                    "average": short_cost / abs(short_qty),
                }

            #  4. Short book transactions
            short_period = [r for r in short_group if date_from <= r['_basis'] <= date_to]
            short_transactions = restructure_short_transactions(short_period) if short_period else []
            if short_transactions:
                ccy = get_ccy(code_ref)
                _ensure(ccy, bank_ref)
                if code_ref not in dict_gain_loss[ccy][bank_ref]["short_beginning"]:
                    dict_gain_loss[ccy][bank_ref]["short_beginning"][code_ref] = {
                        "quantity": 0, "cost_to_date": 0, "average": 0,
                    }
                dict_gain_loss[ccy][bank_ref]["short_trades"][code_ref] = short_transactions

    return dict_gain_loss


def write_box_normal(ws, cell_name, cell_value, value_format="string", bold=False):
    cell = ws[cell_name]
    cell.value = cell_value
    if bold:
        cell.font = openpyxl.styles.Font(size="12", name='Arial', bold=True)
    else:
        cell.font = openpyxl.styles.Font(size="12", name='Arial')
    cell.border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='thin'),
        right=openpyxl.styles.borders.Side(style='thin'),
        top=openpyxl.styles.borders.Side(style='thin'),
        bottom=openpyxl.styles.borders.Side(style='thin')
    )

    if value_format == "integer":
        cell.number_format = '#,##0_);(#,##0)'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "date":
        cell.number_format = 'd-mmm-yyyy'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "shares":
        cell.number_format = '#,##0_);(#,##0)'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='right',
            vertical='center',
            indent=0)
    elif value_format == "currency_4":
        cell.number_format = '#,##0.0000'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "currency_2":
        cell.number_format = '#,##0.00_);(#,##0.00)'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='right',
            vertical='center',
            indent=0)
    elif value_format == "percentage":
        cell.number_format = '0.00%;( 0.00% )'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "code":
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "string":
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='left',
            vertical='center',
            indent=0)


def write_title(ws, cell_name, cell_value, font_size, h_align, border=False):
    cell = ws[cell_name]
    cell.value = cell_value
    cell.font = openpyxl.styles.Font(size=font_size, bold=True, name='Arial')
    cell.alignment = openpyxl.styles.Alignment(
        horizontal=h_align,
        vertical='center')

    if border:
        cell.border = openpyxl.styles.borders.Border(
            left=openpyxl.styles.borders.Side(style='thin'),
            right=openpyxl.styles.borders.Side(style='thin'),
            top=openpyxl.styles.borders.Side(style='thin'),
            bottom=openpyxl.styles.borders.Side(style='thin')
        )


def write(ws, cell_name, cell_value, font_size, h_align, value_format, bold=False):
    cell = ws[cell_name]
    cell.value = cell_value
    if bold:
        cell.font = openpyxl.styles.Font(size=font_size, bold=True, name='Arial')
    else:
        cell.font = openpyxl.styles.Font(size=font_size)

    if value_format == "integer":
        cell.number_format = '#,##0_);(#,##0)'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "date":
        cell.number_format = 'd-mmm-yyyy'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "shares":
        cell.number_format = '#,##0_);(#,##0)'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='right',
            vertical='center',
            indent=0)
    elif value_format == "currency_4":
        cell.number_format = '#,##0.0000'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "currency_2":
        cell.number_format = '#,##0.00_);(#,##0.00)'
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='right',
            vertical='center',
            indent=0)
    elif value_format == "code":
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center')
    elif value_format == "string":
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='left',
            vertical='center',
            indent=0)


def fill_color(ws, cell_name, cell_color):
    cell = ws[cell_name]
    cell.fill = openpyxl.styles.fills.PatternFill(
        patternType='solid',
        fgColor=cell_color
    )

