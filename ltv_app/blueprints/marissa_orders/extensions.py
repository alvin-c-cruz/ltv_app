import os
from datetime import datetime, timedelta

from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side


def _instance_temp_path(filename):
    return os.path.join(current_app.instance_path, 'temp', filename)


def _posting_transactions(db, start_date, end_date):
    """Port of DownloadPosting.transactions (methods.py)."""
    sql = """
      SELECT
        tbl_bank_account.bank_id,
        tbl_bank_account.bank_name,
        tbl_code.code,
        tbl_code.stock_name,
        tbl_transaction.trade_date,
        tbl_transaction.value_date,
        tbl_transaction.transaction_type,
        tbl_transaction.quantity,
        tbl_transaction.price,
        tbl_transaction.brokerage,
        tbl_transaction.commission,
        tbl_transaction.foreign_charge,
        tbl_transaction.stamp_duty,
        tbl_transaction.misc,
        tbl_transaction.counter_bank_ref,
        tbl_currency.ccy_id
      FROM tbl_transaction
      INNER JOIN tbl_bank_account on tbl_bank_account.ref_num = tbl_transaction.bank_ref
      INNER JOIN tbl_code on tbl_code.ref_num = tbl_transaction.code_ref
      INNER JOIN tbl_currency on tbl_currency.ref_num = tbl_code.ccy_ref
      INNER JOIN tbl_transaction_type on tbl_transaction_type.transaction_type = tbl_transaction.transaction_type
      WHERE tbl_transaction.trade_date >= ?
        AND tbl_transaction.trade_date <= ?
      ORDER BY tbl_bank_account.priority, tbl_currency.priority, tbl_transaction.trade_date, tbl_transaction_type.priority
      ;
    """

    result = db.execute(sql, (start_date, end_date)).fetchall()

    dict_transaction = {}

    for transaction in result:
        bank_id = transaction['bank_id']
        trade_date = transaction['trade_date']
        trade_date = datetime(
            int(trade_date[:4]),
            int(trade_date[5:7]),
            int(trade_date[-2:])
        )
        value_date = transaction['value_date']
        value_date = datetime(
            int(value_date[:4]),
            int(value_date[5:7]),
            int(value_date[-2:])
        )
        code = transaction['code']
        transaction_type = transaction['transaction_type']
        quantity = transaction['quantity']
        price = transaction['price']
        brokerage = transaction['brokerage']
        commission = transaction['commission']
        foreign_charge = transaction['foreign_charge']
        stamp_duty = transaction['stamp_duty']
        misc = transaction['misc']
        counter_bank_ref = transaction['counter_bank_ref']
        ccy_id = transaction['ccy_id']

        if bank_id not in dict_transaction: dict_transaction[bank_id] = {}

        if ccy_id not in dict_transaction[bank_id]: dict_transaction[bank_id][ccy_id] = []

        dict_transaction[bank_id][ccy_id].append(
            {
                'trade_date': trade_date,
                'value_date': value_date,
                'code': code,
                'transaction_type': transaction_type,
                'quantity': quantity,
                'price': price,
                'brokerage': brokerage,
                'commission': commission,
                'foreign_charge': foreign_charge,
                'stamp_duty': stamp_duty,
                'misc': misc,
                'counter_bank_ref': counter_bank_ref
            }
        )

    return dict_transaction


def _posting_write_sheet(db, ws, transactions, date_ref):
    """Port of DownloadPosting.write_sheet (methods.py)."""
    date_label = 'Trade Date' if date_ref == 'trade_date' else 'Value Date'

    row_num = 1
    cols = {
        'A': date_label,
        'B': 'Transaction',
        'C': 'Type',
        'D': 'Code',
        'E': 'Description',
        'F': 'Quantity',
        'G': 'Price',
        'H': 'Particulars',
        'I': 'Amount',
        'J': 'Charges'
    }

    for col in cols:
        cell = ws[f'{col}{row_num}']
        cell.value = cols[col]

    row_num = 2
    for transaction in transactions:
        trade_date = transaction['trade_date']
        value_date = transaction['value_date']
        code = transaction['code']
        transaction_type = transaction['transaction_type']
        quantity = transaction['quantity']
        price = transaction['price']
        brokerage = transaction['brokerage']
        commission = transaction['commission']
        foreign_charge = transaction['foreign_charge']
        stamp_duty = transaction['stamp_duty']
        misc = transaction['misc']
        counter_bank_ref = transaction['counter_bank_ref']

        _date = trade_date if date_ref == 'trade_date' else value_date

        if 'Buy' in transaction_type:
            side = 'PURCHASE'
        elif 'Sell' in transaction_type:
            side = 'SOLD'
        else:
            side = transaction_type

        if 'Spot' in transaction_type:
            _type = 'SPOT'
        elif 'Accu' in transaction_type:
            _type = 'ACCU'
        elif 'Decu' in transaction_type:
            _type = 'DECU'
        else:
            if 'Transfer' in transaction_type:
                transfer_account = db.execute("SELECT bank_name FROM tbl_bank_account WHERE ref_num = ?", (counter_bank_ref, )).fetchone()[0]
                _type = transaction_type + " - " + transfer_account
            else:
                _type = transaction_type

        try:
            code = int(code)
        except:
            pass

        charges = brokerage + commission + foreign_charge + stamp_duty + misc

        cols = {
            'A': _date,
            'B': side,
            'C': _type,
            'D': code,
            'E': f'=INDEX(INDIRECT("CODES!A:B"),MATCH($D{row_num},INDIRECT("CODES!A:A"),),2)&" - "&C{row_num}',
            'F': quantity,
            'G': price,
            'H': f'=B{row_num}&" - "&E{row_num}&" - "&C{row_num}&" - "&TEXT(F{row_num},"#,###,##0")&" shares @ "&TEXT(G{row_num},"#,###,##0.0000")',
            'I': f'=((F{row_num}*G{row_num})+J{row_num})*-1',
            'J': charges
        }

        for col in cols:
            cell = ws[f'{col}{row_num}']
            cell.value = cols[col]

            if col == 'A':
                cell.number_format = 'dd-mmm-yyyy'
            elif col == 'F':
                cell.number_format = '#,##0'
            elif col == 'G':
                cell.number_format = '#,##0.0000'
            elif col == 'I':
                cell.number_format = '#,##0.00'
            elif col == 'J':
                cell.number_format = '#,##0.00'

        row_num += 1


def download_posting(db, posting_month, posting_year):
    """Port of localhost/libraries/Marissa_Orders/methods.py::DownloadPosting."""
    date_ref = {
        'CB1': 'trade_date',  #  Changes from value_date
        'CB2': 'trade_date',  #  Changes from value_date
        'CB3': 'trade_date',  #  Changes from value_date
        'CBBH': 'trade_date',  #  Changes from value_date
        'CBBH2': 'trade_date',  #  Changes from value_date
        'CBSG': 'trade_date',  #  Changes from value_date
        'BOS': 'trade_date',  #  Changes from value_date
        'DBPe': 'trade_date',  #  Changes from value_date
        'DBPL': 'trade_date',  #  Changes from value_date
        'SC': 'trade_date',  #  Changes from value_date
        'SHK': 'trade_date',  #  Changes from value_date
        'SHK2': 'trade_date',  #  Changes from value_date
        'MST1': 'trade_date',
        'MST2': 'trade_date',
        'MSPL': 'trade_date',
        'NSG': 'trade_date',
    }

    start_date = str(datetime(posting_year, posting_month, 1))[:10]
    end_date = str(datetime(
        posting_year if posting_month != 12 else posting_year + 1,
        posting_month + 1 if posting_month != 12 else 1,
        1) - timedelta(days=1)
    )[:10]

    wb = Workbook()
    del wb['Sheet']

    transactions = _posting_transactions(db, start_date, end_date)

    for bank_id in transactions:
        for ccy_id in transactions[bank_id]:
            wb.create_sheet(f'{bank_id}-{ccy_id}')
            ws = wb[f'{bank_id}-{ccy_id}']
            _posting_write_sheet(db, ws, transactions[bank_id][ccy_id], date_ref[bank_id])

    path = _instance_temp_path(f'stock_posting_{posting_year}_{posting_month:02d}.xlsx')
    wb.save(path)
    wb.close()
    return path


def download_daily_transactions(db, trade_date):
    """Port of DownloadDailyTransactions — single date, single sheet,
    columns Trade Date/Account/Transaction/Stock/Quantity/Price/Amount(formula)."""
    sql = """
      SELECT
        tbl_bank_account.bank_id,
        tbl_bank_account.bank_name,
        tbl_code.code,
        tbl_code.stock_name,
        tbl_transaction.trade_date,
        tbl_transaction.transaction_type,
        tbl_transaction.quantity,
        tbl_transaction.price,
        tbl_transaction.counter_bank_ref,
        tbl_currency.ccy_id
      FROM tbl_transaction
      INNER JOIN tbl_bank_account on tbl_bank_account.ref_num = tbl_transaction.bank_ref
      INNER JOIN tbl_code on tbl_code.ref_num = tbl_transaction.code_ref
      INNER JOIN tbl_currency on tbl_currency.ref_num = tbl_code.ccy_ref
      INNER JOIN tbl_transaction_type on tbl_transaction_type.transaction_type = tbl_transaction.transaction_type
      WHERE tbl_transaction.trade_date = ?
      ORDER BY tbl_currency.priority, tbl_bank_account.priority, tbl_code.code, tbl_transaction_type.priority
      ;
    """

    result = db.execute(sql, (trade_date, )).fetchall()

    list_transaction = []

    for transaction in result:
        bank_name = transaction['bank_name']
        t_date = transaction['trade_date']
        t_date = datetime(
            int(t_date[:4]),
            int(t_date[5:7]),
            int(t_date[-2:])
        )
        code = transaction['code']
        transaction_type = transaction['transaction_type']
        quantity = transaction['quantity']
        price = transaction['price']
        counter_bank_ref = transaction['counter_bank_ref']
        ccy_id = transaction['ccy_id']

        list_transaction.append(
            {
                'ccy_id': ccy_id,
                'trade_date': t_date,
                'bank_name': bank_name,
                'transaction_type': transaction_type,
                'code': code,
                'quantity': quantity,
                'price': price,
                'counter_bank_ref': counter_bank_ref
            }
        )

    wb = Workbook()
    ws = wb.active

    row_num = 1
    cols = {
        'A': 'Trade Date',
        'B': 'Account',
        'C': 'Transaction',
        'D': 'Stock',
        'E': 'Quantity',
        'F': 'Price',
        'G': 'Amount',
    }

    for col in cols:
        cell = ws[f'{col}{row_num}']
        cell.value = cols[col]

    row_num = 2
    for transaction in list_transaction:
        t_date = transaction['trade_date']
        bank_name = transaction['bank_name']
        code = transaction['code']
        transaction_type = transaction['transaction_type']
        quantity = transaction['quantity']
        price = transaction['price']

        cols = {
            'A': t_date,
            'B': bank_name,
            'C': transaction_type,
            'D': code,
            'E': quantity,
            'F': price,
            'G': f'=E{row_num}*F{row_num}',
        }

        for col in cols:
            cell = ws[f'{col}{row_num}']
            cell.value = cols[col]

            if col == 'A':
                cell.number_format = 'dd-mmm-yyyy'
            elif col == 'E':
                cell.number_format = '#,##0'
            elif col == 'F':
                cell.number_format = '#,##0.0000'
            elif col == 'G':
                cell.number_format = '#,##0.00'

        row_num += 1

    path = _instance_temp_path(f'daily_transactions_{trade_date}.xlsx')
    wb.save(path)
    wb.close()
    return path


def download_transaction_range(db, start_date, end_date):
    """Port of DownloadTransactionRange — date range, single sheet, same
    columns as daily transactions plus explicit column widths + thin
    borders on every cell."""
    sql = """
      SELECT
        tbl_bank_account.bank_name,
        tbl_code.code,
        tbl_transaction.trade_date,
        tbl_transaction.transaction_type,
        tbl_transaction.quantity,
        tbl_transaction.price,
        tbl_transaction.counter_bank_ref,
        tbl_currency.ccy_id
      FROM tbl_transaction
      INNER JOIN tbl_bank_account on tbl_bank_account.ref_num = tbl_transaction.bank_ref
      INNER JOIN tbl_code on tbl_code.ref_num = tbl_transaction.code_ref
      INNER JOIN tbl_currency on tbl_currency.ref_num = tbl_code.ccy_ref
      INNER JOIN tbl_transaction_type on tbl_transaction_type.transaction_type = tbl_transaction.transaction_type
      WHERE tbl_transaction.trade_date >= ? AND tbl_transaction.trade_date <= ?
      ORDER BY tbl_transaction.trade_date, tbl_currency.priority, tbl_bank_account.priority, tbl_code.code, tbl_transaction_type.priority
      ;
    """
    result = db.execute(sql, (start_date, end_date)).fetchall()
    list_transaction = []
    for transaction in result:
        trade_date = transaction['trade_date']
        trade_date = datetime(int(trade_date[:4]), int(trade_date[5:7]), int(trade_date[-2:]))
        list_transaction.append({
            'trade_date': trade_date,
            'bank_name': transaction['bank_name'],
            'transaction_type': transaction['transaction_type'],
            'code': transaction['code'],
            'quantity': transaction['quantity'],
            'price': transaction['price'],
            'counter_bank_ref': transaction['counter_bank_ref'],
        })

    wb = Workbook()
    ws = wb.active

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions['A'].width = 11.43
    ws.column_dimensions['B'].width = 29.0
    ws.column_dimensions['C'].width = 14.0
    ws.column_dimensions['D'].width = 5.71
    ws.column_dimensions['E'].width = 8.71
    ws.column_dimensions['F'].width = 8.57
    ws.column_dimensions['G'].width = 12.43

    row_num = 1
    headers = {'A': 'Trade Date', 'B': 'Account', 'C': 'Transaction', 'D': 'Stock', 'E': 'Quantity', 'F': 'Price', 'G': 'Amount'}
    for col, label in headers.items():
        cell = ws[f'{col}{row_num}']
        cell.value = label
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for t in list_transaction:
        ws[f'A{row_num}'].value = t['trade_date']
        ws[f'A{row_num}'].number_format = r'dd\-mmm\-yyyy'
        ws[f'B{row_num}'].value = t['bank_name']
        ws[f'C{row_num}'].value = t['transaction_type']
        ws[f'D{row_num}'].value = t['code']
        ws[f'E{row_num}'].value = t['quantity']
        ws[f'E{row_num}'].number_format = '#,##0'
        ws[f'F{row_num}'].value = t['price']
        ws[f'F{row_num}'].number_format = '#,##0.0000'
        ws[f'G{row_num}'].value = f'=E{row_num}*F{row_num}'
        ws[f'G{row_num}'].number_format = '#,##0.00'
        for col in 'ABCDEFG':
            ws[f'{col}{row_num}'].border = border
        row_num += 1

    path = _instance_temp_path(f'transaction_range_{start_date}_{end_date}.xlsx')
    wb.save(path)
    wb.close()
    return path
