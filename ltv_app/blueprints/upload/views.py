from flask import Blueprint, render_template, request, current_app, redirect, jsonify, send_from_directory, abort, url_for
from werkzeug.utils import secure_filename
import os
from datetime import datetime
from openpyxl import load_workbook
from dataclasses import dataclass

from ..auth import login_required, superuser_required
from ..transactions import Transaction
from ..database import get_db

bp = Blueprint('upload', __name__, template_folder="pages", url_prefix="/upload")


@bp.route('/downloads')
@login_required
def downloads():
    uploads_dir = os.path.join(current_app.instance_path, 'uploads')
    try:
        files = sorted(os.listdir(uploads_dir), reverse=True)
    except FileNotFoundError:
        files = []
    return render_template('upload/downloads.html', files=files)


@bp.route('/downloads/<path:filename>')
@login_required
def download_file(filename):
    uploads_dir = os.path.join(current_app.instance_path, 'uploads')
    safe = os.path.normpath(os.path.join(uploads_dir, filename))
    if not safe.startswith(os.path.normpath(uploads_dir)):
        abort(404)
    return send_from_directory(uploads_dir, filename, as_attachment=True)


def _inspect_dir():
    """Folder where files uploaded for Claude inspection are stored."""
    path = current_app.config.get(
        'INSPECT_UPLOAD_DIR',
        os.path.join(current_app.instance_path, 'uploads', 'inspect'),
    )
    os.makedirs(path, exist_ok=True)
    return path


@bp.route('/inspect', methods=["GET", "POST"])
@superuser_required
def inspect():
    inspect_dir = _inspect_dir()

    if request.method == "POST":
        for file in request.files.getlist("file[]"):
            filename = secure_filename(file.filename)
            if filename:
                file.save(os.path.join(inspect_dir, filename))
        return redirect(url_for('upload.inspect'))

    files = []
    for name in sorted(os.listdir(inspect_dir)):
        full_path = os.path.join(inspect_dir, name)
        if not os.path.isfile(full_path):
            continue
        stat = os.stat(full_path)
        files.append({
            'name': name,
            'size_kb': f"{stat.st_size / 1024:,.1f}",
            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
        })
    return render_template('upload/inspect.html', files=files)


@bp.route('/inspect/clear', methods=["POST"])
@superuser_required
def inspect_clear():
    inspect_dir = _inspect_dir()
    for name in os.listdir(inspect_dir):
        full_path = os.path.join(inspect_dir, name)
        if os.path.isfile(full_path):
            os.remove(full_path)
    return redirect(url_for('upload.inspect'))


@bp.route('/heroku', methods=["GET", "POST"])
@login_required
def heroku():
    if request.method == "POST":
        root_path = os.path.join(current_app.instance_path, 'heroku')
        # remove old files
        for file in os.listdir(root_path):
            os.remove(os.path.join(root_path, file))

        files = request.files.getlist("file[]")
        for file in files:
            filename = os.path.join(root_path, secure_filename(file.filename))
            file.save(filename)
        fetched_data = FetchData(os.path.join(root_path))
        SaveData(data=fetched_data.data, db=get_db())
        return redirect("/")
    return render_template('upload/heroku.html')


class FetchData:
    data = []

    def __init__(self, path):
        files = os.listdir(path)
        for i, file in enumerate(files):
            filename = os.path.join(path, file)
            self.data += self.open_file(filename)

    def open_file(self, filename):
        list_records = []
        account = self.get_account(filename)
        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            code = self.get_code(sheet_name)
            side = 'short' if sheet_name[-2:] == '-S' else 'long'
            list_records += self.get_records(ws, code, account, side)

        wb.close()
        return list_records

    def get_account(self, filename):
        if 'bank-of-singapore-bos' in filename:
            return 'BOS'
        if 'citibank-account-1' in filename:
            return 'CB1'
        if 'citibank-account-2' in filename:
            return 'CB2'
        if 'citibank-account-3' in filename:
            return 'CB3'
        if 'citibank-bh' in filename:
            return 'CBBH'
        if 'citibank-bh-2' in filename:
            return 'CBBH2'
        if 'deutsche-bank-personal' in filename:
            return 'DBPe'
        if 'deutsche-bank-pl' in filename:
            return 'DBPL'
        if 'standard-chartered-stocks' in filename:
            return 'SC'
        if 'shk-private-shkis' in filename:
            return 'SHK'
        if 'morgan-stanley-t1' in filename:
            return 'MST1'
        if 'morgan-stanley-t2' in filename:
            return 'MST2'
        if 'morgan-stanley-pl' in filename:
            return 'MSPL'

    def get_code(self, sheet_name):
        i = sheet_name.find('-')
        code = sheet_name[:i]
        if code.isnumeric():
            code = code.rjust(4, '0')

        return code

    def get_transaction(self, transaction):
        if 'TRANSFER' in transaction:
            if 'TRANSFER OUT' in transaction:
                return 'Transfer-Out'
            if 'TRANSFER IN' in transaction:
                return 'Transfer-In'
        else:
            heroku_transaction = {
                'INITIALIZE': 'INITIALIZE',
                'PURCHASE - ACCU': 'Buy (Accu)',
                'PURCHASE - ACCU KO': 'Buy (Accu-KO)',
                'SOLD - DECU': 'Sell (Decu)',
                'SOLD - DECU KO': 'Sell (Decu-KO)',
                'PURCHASE - SPOT': 'Buy (Spot)',
                'SOLD - SPOT': 'Sell (Spot)',
                'DIVIDEND': 'Stock Dividend',
                'SHORT SELL': 'Sell (Short)',
                'PAY SHORT (From Account)': 'From Account (Pay Short)',
                'PAY SHORT': 'Buy (Pay Short)',
            }

            if transaction in heroku_transaction:
                return heroku_transaction[transaction]
            else:
                print('Transaction: ' + transaction)
                return transaction

    def get_counter_account(self, transaction):
        if 'TRANSFER' not in transaction:
            return None
        heroku_transactions = {
            'Citibank - Account 1 (1)': 'CB1',
            'Citibank - Account 2 (2)': 'CB2',
            'Citibank - Account 3 (4)': 'CB3',
            'Citibank - BH (3)': "CBBH",
            'Citibank - BH 2 (5)': "CBBH2",
            'Bank of Singapore - BOS (10)': 'BOS',
            'Deutsche Bank - Personal (4)': 'DBPe',
            'Deutsche Bank - PL (5)': 'DBPL',
            'Standard Chartered - Stocks (6)': 'SC',
            'SHK Private - SHKIS (SC Bank HK)': 'SHK',
            'Morgan Stanley - T1 (7)': 'MST1',
            'Morgan Stanley - T2 (8)': 'MST2',
            'Morgan Stanley - PL (9)': 'MSPL',
        }
        counter_transaction = transaction[transaction.find("-") + 2:]
        if counter_transaction in heroku_transactions:
            return heroku_transactions[counter_transaction]
        else:
            print('Counter: ' + counter_transaction)
            return transaction

    def get_records(self, ws, code, account, side):
        list_record = []
        start_row = 5
        for row in range(start_row, ws.max_row):
            trade_date = str(ws.cell(row=row, column=1).value)[:10]
            value_date = str(ws.cell(row=row, column=2).value)[:10]
            transaction = ws.cell(row=row, column=3).value
            quantity = ws.cell(row=row, column=4).value
            price = ws.cell(row=row, column=5).value
            brokerage_fee = ws.cell(row=row, column=6).value
            commission = ws.cell(row=row, column=7).value
            foreign_charge = ws.cell(row=row, column=8).value
            stamp_duty = ws.cell(row=row, column=9).value
            misc = ws.cell(row=row, column=10).value

            if transaction in ("SUMMARY", None):
                continue

            list_record.append(
                {
                    'account': account,
                    'code': code,
                    'side': side,
                    'trade_date': trade_date,
                    'value_date': value_date,
                    'transaction': self.get_transaction(transaction),
                    'quantity': quantity,
                    'price': price,
                    'brokerage_fee': brokerage_fee,
                    'commission': commission,
                    'foreign_charge': foreign_charge,
                    'stamp_duty': stamp_duty,
                    'misc': misc,
                    'counter_account': self.get_counter_account(transaction)
                }
            )

        return list_record


@dataclass
class SaveData:
    data: any = None
    db: any = None

    def __post_init__(self):
        self.clear_tables()
        for entry in self.data:
            self.save_record(entry)
        self.db.commit()

    def clear_tables(self):
        self.db.execute("DELETE FROM tbl_transaction;")

    def save_record(self, entry):
        if entry['side'] == "long":
            table_name = "tbl_transaction"
        else:
            return None

        bank_ref = self.db.execute("SELECT ref_num FROM tbl_bank_account WHERE bank_id=?;",
                                           (entry['account'],)).fetchone()[0]
        code_ref = self.db.execute("SELECT ref_num FROM tbl_code WHERE code=?;",
                                            (entry['code'],)).fetchone()[0]
        trade_date = entry['trade_date']
        value_date = entry['value_date']
        transaction = entry['transaction']
        quantity = entry['quantity']
        price = entry['price']
        brokerage_fee = entry['brokerage_fee']
        commission = entry['commission']
        foreign_charge = entry['foreign_charge']
        stamp_duty = entry['stamp_duty']
        misc = entry['misc']
        if entry['counter_account']:
            counter_bank_ref = self.db.execute("SELECT ref_num FROM tbl_bank_account WHERE bank_id=?;",
                                                       (entry['counter_account'],)).fetchone()[0]
        else:
            counter_bank_ref = None

        sql = f"""
            INSERT INTO {table_name}
            (bank_ref, code_ref, trade_date, value_date, transaction_type, quantity, price, brokerage, commission, 
            foreign_charge, stamp_duty, misc, counter_bank_ref)
            VALUES
            (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ;
        """

        self.db.execute(sql,
                        (bank_ref, code_ref, trade_date, value_date, transaction,
                            quantity, price,
                            brokerage_fee, commission,
                            foreign_charge, stamp_duty, misc, counter_bank_ref)
                        )

