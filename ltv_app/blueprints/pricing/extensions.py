import io

BANKS = {
    "CB": "Citibank",
    "BOS": "Bank of Singapore",
    "DB": "Deutsche Bank",
    "GS": "Goldman Sachs",
    "SHK": "Sun Hung Kai",
    "MS": "Morgan Stanley",
    "NSG": "Nomura"
}

HEADER = {
    "BOS": [
        {
          "CP": "",
          "Daycount": "",
          "Guaranteed": "gtd",
          "KO Barrier": "ko",
          "Normal/Leveraged": "leverage",
          "Product": "product",
          "RFQ ID": "",
          "Set Freq": "frequency",
          "Strike": "strike",
          "Tenor": "tenor",
          "Underlying": "code"
        }
    ],
    "SHK": [
        {
          "GUARANTEED PERIODS": "gtd",
          "Issuer": "",
          "KO (%)": "ko",
          "KO+1 SETTLEMENT CYCLE": "",
          "NORMAL / LEVERAGED": "leverage",
          "PRODUCT": "product",
          "Remark": "remark",
          "SETTLEMENT FREQUENCY": "frequency",
          "Strike": "strike",
          "TENOR (M)": "tenor",
          "Tradable Size": "tradable_size",
          "UNDERLYING (BLOOMBERG)": "code"
        }
    ],
    "GS": [
        {
          "GUARANTEED PERIODS": "gtd",
          "Issuer": "",
          "KO (%)": "ko",
          "KO+1 SETTLEMENT CYCLE": "",
          "NORMAL / LEVERAGED": "leverage",
          "PRODUCT": "product",
          "Remark": "remark",
          "SETTLEMENT FREQUENCY": "frequency",
          "Strike (%)": "strike",
          "TENOR (M)": "tenor",
          "Tradable Size": "tradable_size",
          "UNDERLYING (BLOOMBERG)": "code"
        }
    ],
}

ACCUMULATORS = (
    "Accum",
)

DECUMULATORS = (
    "Decum",
)

LEVERAGE_SINGLE = (
    "Normal",
)

LEVERAGE_DOUBLE = (
    "Leveraged",
)

GTD_NO = (
    "0m",
)

GTD_1 = (
    "1m",
)


def transform_email(textarea_data, bank_code, form):
    list_data = transform_raw_email(textarea_data, bank_code, form)
    analyzed_data = analyze_data(list_data)
    return analyzed_data


def transform_raw_email(textarea_data, bank_code, form):
    list_data = []
    for i, row in enumerate(textarea_data.split("\n")):
        if i == 0:
            email_headers = list([i.strip() for i in row.replace("\r", "").split("\t")])
            sorted_header = list([i.strip() for i in row.replace("\r", "").split("\t")]).sort()

            headers = [list(header.keys()) for header in HEADER[bank_code]]
            if headers:
                for i, header in enumerate(headers):
                    if sorted_header == header.sort():
                        use_header = HEADER[bank_code][i]
                        break
                else:
                    form["not_in_record"] = email_headers if email_headers[0] != "" else []
                    break
            else:
                form["not_in_record"] = email_headers if email_headers[0] != "" else []
                use_header = []
        else:
            row_data = row.replace("\r", "").split("\t")
            dict_row = {}
            for key, value in zip(email_headers, row_data):
                if value and use_header:
                    use_key = use_header[key]
                    if use_key:
                        dict_row[use_key] = value

            if dict_row:
                dict_row["bank_code"] = bank_code
                list_data.append(dict_row)

    return list_data


def analyze_data(list_data):
    def decision(key, _dict, *args):
        for choices, value in args:
            if _dict[key] in choices:
                _dict[key] = value
                break
        else:
            _dict[key] += " not in record"

    for row in list_data:
        decision("product", row, (ACCUMULATORS, "ACCU"), (DECUMULATORS, "DECU"))
        decision("leverage", row, (LEVERAGE_DOUBLE, "Yes"), (LEVERAGE_SINGLE, "No"))
        decision("gtd", row, (GTD_NO, "No"), (GTD_1, "1m"))

    analyzed_data = list_data
    return analyzed_data


def parse_rate(val):
    """Parse rate as decimal: '85.45%' → 0.8545, '0.8545' → 0.8545"""
    val = str(val).strip()
    if not val:
        return 0.0
    if val.endswith('%'):
        return float(val[:-1]) / 100.0
    return float(val)


def parse_pasted_data(textarea_data, bank_code):
    """Parse tab-separated pasted pricing data into a list of row dicts."""
    if bank_code not in HEADER:
        return []

    header_map = HEADER[bank_code][0]
    email_headers = None
    rows = []

    for raw_line in textarea_data.replace('\r\n', '\n').replace('\r', '\n').split('\n'):
        if not raw_line.strip():
            continue
        cells = raw_line.split('\t')

        if email_headers is None:
            email_headers = [c.strip() for c in cells]
            continue

        row_data = {}
        for col_name, cell_val in zip(email_headers, cells):
            mapped = header_map.get(col_name.strip(), '')
            if mapped:
                row_data[mapped] = cell_val.strip()

        if not row_data.get('product'):
            continue

        product = row_data['product']
        if product in ACCUMULATORS:
            row_data['product'] = 'ACCU'
        elif product in DECUMULATORS:
            row_data['product'] = 'DECU'

        rows.append(row_data)

    return rows


def generate_pricing_excel(rows, stock_name, spot_price, date_str):
    """Generate a formatted pricing Excel with ACC and DECU sections."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = str(stock_name)[:31]

    spot = float(spot_price) if spot_price else 0.0

    # ── Styles ───────────────────────────────────────────────────────────────
    dark_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    acc_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    decu_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    white_bold = Font(color="FFFFFF", bold=True, size=10)
    bold10     = Font(bold=True, size=10)
    normal10   = Font(size=10)

    thin_s = Side(style='thin',   color="BFBFBF")
    med_s  = Side(style='medium', color="8EA9C1")
    data_border  = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    sec_border   = Border(left=med_s,  right=med_s,  top=med_s,  bottom=med_s)

    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    COLS = [
        # (display_header,   data_key,        width)
        ('PRODUCT',          'product',        12),
        ('NORMAL/LEV.',      'leverage',       11),
        ('GTD PERIODS',      'gtd',            11),
        ('KO (%)',           'ko',              9),
        ('KO PRICE',         '_ko_price',      11),
        ('TENOR (M)',        'tenor',           9),
        ('STRIKE (%)',       'strike',          9),
        ('STRIKE PRICE',     '_strike_price',  12),
        ('TRADABLE SIZE',    'tradable_size',  14),
        ('REMARK',           'remark',         24),
    ]
    NUM_COLS = len(COLS)

    # ── Column widths ────────────────────────────────────────────────────────
    for col_idx, (_, _, width) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Row 1: title / spot ──────────────────────────────────────────────────
    ws['A1'] = stock_name
    ws['A1'].font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells('A1:C1')

    ws['D1'] = date_str
    ws['D1'].font = Font(size=10, color="595959")
    ws['D1'].alignment = center

    ws['F1'] = 'SPOT'
    ws['F1'].font = Font(bold=True, size=10)
    ws['F1'].alignment = center

    ws['G1'] = spot
    ws['G1'].number_format = '#,##0.000'
    ws['G1'].font = Font(bold=True, size=11, color="C00000")
    ws['G1'].alignment = center

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 6   # blank spacer

    current_row = [3]

    def write_section(label, section_rows, hdr_fill):
        r = current_row[0]

        # Section label
        cell = ws.cell(row=r, column=1, value=label)
        cell.fill = hdr_fill
        cell.font = bold10
        cell.alignment = left
        cell.border = sec_border
        for c in range(2, NUM_COLS + 1):
            bc = ws.cell(row=r, column=c)
            bc.fill = hdr_fill
            bc.border = sec_border
        ws.row_dimensions[r].height = 18
        r += 1

        # Column headers
        for c, (hdr_label, _, _) in enumerate(COLS, start=1):
            cell = ws.cell(row=r, column=c, value=hdr_label)
            cell.fill = dark_fill
            cell.font = white_bold
            cell.alignment = center
            cell.border = data_border
        ws.row_dimensions[r].height = 28
        r += 1

        # Data rows
        for row_data in section_rows:
            for c, (_, key, _) in enumerate(COLS, start=1):
                if key == '_ko_price':
                    try:
                        val = round(spot * parse_rate(str(row_data.get('ko', 0))), 3) if spot else ''
                    except Exception:
                        val = ''
                elif key == '_strike_price':
                    try:
                        val = round(spot * parse_rate(str(row_data.get('strike', 0))), 3) if spot else ''
                    except Exception:
                        val = ''
                elif key in ('ko', 'strike'):
                    try:
                        val = parse_rate(str(row_data.get(key, '')))
                    except Exception:
                        val = row_data.get(key, '')
                else:
                    val = row_data.get(key, '')

                cell = ws.cell(row=r, column=c, value=val)
                cell.border = data_border
                cell.font = normal10

                if key in ('ko', 'strike') and isinstance(val, float):
                    cell.number_format = '0.00%'
                    cell.alignment = center
                elif key in ('_ko_price', '_strike_price') and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.000'
                    cell.alignment = center
                elif key in ('tenor', 'product', 'leverage', 'gtd'):
                    cell.alignment = center
                else:
                    cell.alignment = left

            ws.row_dimensions[r].height = 16
            r += 1

        current_row[0] = r + 1  # blank spacer after section

    acc_rows  = [r for r in rows if r.get('product') == 'ACCU']
    decu_rows = [r for r in rows if r.get('product') == 'DECU']

    if acc_rows:
        write_section('ACCUMULATOR', acc_rows, acc_fill)
    if decu_rows:
        write_section('DECUMULATOR', decu_rows, decu_fill)

    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
