from flask import current_app
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.fills import PatternFill, Color
from openpyxl.styles.borders import Border, Side
from datetime import datetime

from .trades_done_average import TradesDoneAverage

HEADER_HEIGHT = 18.75
DETAIL_HEIGHT = 45
DIVIDER_HEIGHT = 6
DIVIDER_COLOR = "00000000"

CODE_TO_SHRINK = ('0857', '0981', '0939', '1099', '1288', '2196', '3993')

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# double_rule_border = Border(bottom=Side(style='double'))


class DownloadTradesDone:
    def __init__(self, db, trade_date, trade_summary):
        self.db = db
        self.trade_date = trade_date
        self.trade_summary = trade_summary

        self.template_file = os.path.join(current_app.instance_path, "excel_templates", "trades_done.xlsx")
        self.filename = os.path.join(current_app.instance_path, "temp", f"{trade_date} Trades Done.xlsx")
        self.create_file()

    def create_file(self):
        wb = load_workbook(self.template_file)
        ws = wb["Sheet"]
        ws.title = f"{int(self.trade_date[-2:])}"

        ws["A2"].value = "Date: " + datetime.strptime(self.trade_date, "%Y-%m-%d").strftime("%B %d, %Y")

        row_num = 5
        if self.trade_summary.accus:
            row_num = self.write_term_sheet(ws, row_num, self.trade_summary.accus, "ACCU")
            row_num += 2

        if self.trade_summary.decus:
            row_num = self.write_term_sheet(ws, row_num, self.trade_summary.decus, "DECU")
            row_num += 2

        if self.trade_summary.transactions:
            row_num = self.write_transactions(ws, row_num, self.trade_summary.transactions)
            row_num += 2

        if self.trade_summary.transfers:
            row_num = self.write_transfers(ws, row_num, self.trade_summary.transfers)
            row_num += 2

        ws.print_area = f"A1:K{row_num}"
        wb.save(self.filename)
        wb.close()

    def write_term_sheet(self, ws, row_num, term_sheets, _type):
        # Header
        cols = {
            "A": "Accu" if _type == "ACCU" else "Decu",
            "B": "",
            "C": "Stock",
            "D": "Shares",
            "E": "Spot",
            "F": "Strike Rate",
            "G": "Strike",
            "H": "KO",
            "I": "KO Rate",
            "J": "Tenor",
            "K": "GTD"
        }
        for col, label in cols.items():
            cell = ws[f"{col}{row_num}"]
            cell.value = label

            if col == "A":
                cell.font = Font(size=14, bold=True)
                cell.alignment = Alignment(horizontal="center")
                if _type == "ACCU":
                    cell.fill = PatternFill(patternType="solid", fill_type="solid", fgColor=Color("0099CC00"))
                else:
                    cell.fill = PatternFill(patternType="solid", fill_type="solid", fgColor=Color("00FF8080"))
            else:
                cell.font = Font(size=11, bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
        ws.row_dimensions[row_num].height = HEADER_HEIGHT
        row_num += 1

        # Details
        for ccy in term_sheets:
            for ts in term_sheets[ccy]:
                cols = {
                    "A": ts.bank_name,
                    "B": ts.code,
                    "C": ts.stock_name,
                    "D": ts.shares,
                    "E": ts.spot,
                    "F": ts.strike_rate,
                    "G": f'=TEXT(E{row_num}*F{row_num}/100,"0.0000")&CHAR(10)&"("&TEXT(F{row_num},"0.00")&"%)"',
                    "H": f'=TEXT(E{row_num}*I{row_num}/100,"0.0000")&CHAR(10)&"("&TEXT(I{row_num},"0.00")&"%)"',
                    "I": ts.ko_rate,
                    "J": ts.tenor,
                    "K": ts.gtd
                }

                for col, label in cols.items():
                    cell = ws[f"{col}{row_num}"]
                    cell.value = label

                    if ts.code in CODE_TO_SHRINK:
                        font_size = 9
                    else:
                        font_size = 11

                    if col == "A":
                        cell.font = Font(size=11, bold=True)
                    elif col == "C":
                        cell.font = Font(size=font_size)
                    else:
                        cell.font = Font(size=11)

                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

                ws.row_dimensions[row_num].height = DETAIL_HEIGHT
                row_num += 1

        return row_num

    def write_transactions(self, ws, row_num, transactions):
        for ccy in transactions:
            for transaction_type in transactions[ccy]:
                if "Transfer" in transaction_type:
                    continue

                start_row = row_num
                trade_total = False
                # Headers
                cols = {
                    "A": "BUY" if "Buy" in transaction_type else "SELL",
                    "D": ccy,
                    "E": "Price",
                    "G": "Shares",
                    "H": "Total"
                }

                if "Buy" in transaction_type:
                    cols["J"] = "Average"

                for col, label in cols.items():
                    cell = ws[f"{col}{row_num}"]
                    cell.value = label

                    if col in ("A", "D"):
                        cell.font = Font(size=14, bold=True)
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.font = Font(size=11, bold=True)
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = thin_border

                if "Buy" in transaction_type:
                    ws.merge_cells(f"J{row_num}:K{row_num}")

                ws.row_dimensions[row_num].height = HEADER_HEIGHT
                row_num += 1

                # Details
                bank_name_count = len(transactions[ccy][transaction_type])
                for i, bank_name in enumerate(transactions[ccy][transaction_type]):
                    for stock_name in transactions[ccy][transaction_type][bank_name]:
                        trade_count = len(transactions[ccy][transaction_type][bank_name][stock_name])
                        if trade_count > 1:
                            trade_total = True
                        for i_trade, trade in enumerate(transactions[ccy][transaction_type][bank_name][stock_name]):
                            quantity = float(trade["quantity"].replace(",", ""))
                            price = float(trade["price"].replace(",", ""))
                            float_check = str(price)
                            loc = float_check.find(".")
                            length = len(float_check) - loc - 1
                            price_decimal = len(float_check[-length:])

                            code = trade["code"]
                            bank_id = trade["bank_id"]

                            cols = {
                                "A": bank_name,
                                "B": f'=A{row_num}&D{row_num}&MID(C{row_num},FIND("(",C{row_num})+1,4)',
                                "C": stock_name,
                                "D": transaction_type,
                                "E": price,
                                "G": quantity,
                                "H": f'=E{row_num}*G{row_num}'
                            }

                            if "Buy" in transaction_type:
                                if i_trade + 1 == trade_count:
                                    average = TradesDoneAverage(db=self.db, trade_date=self.trade_date,
                                                                code=code, bank_id=bank_id).average
                                else:
                                    average = None

                                cols["J"] = average

                            for col, value in cols.items():
                                cell = ws[f"{col}{row_num}"]
                                cell.value = value

                                cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                                cell.border = thin_border

                                if col == "A":
                                    cell.font = Font(size=11, bold=True)
                                else:
                                    cell.font = Font(size=9) if col == "C" and code in CODE_TO_SHRINK else Font(size=11)

                                if col == "E":
                                    if price_decimal < 3:
                                        cell.number_format = "#,##0.00"
                                    else:
                                        cell.number_format = "#,##0.0000"
                                elif col == "G":
                                    cell.number_format = "#,##0"
                                elif col == "H":
                                    cell.number_format = f"[${ccy}] #,##0.00"
                                elif col == "J":
                                    cell.number_format = "#,##0.0000"

                            if "Buy" in transaction_type:
                                average = None
                                cols["J"] = average
                                ws.merge_cells(f"J{row_num}:K{row_num}")

                            ws.row_dimensions[row_num].height = DETAIL_HEIGHT
                            row_num += 1

                    if i != bank_name_count-1:
                        # Bank divider
                        cols = ["A", "B", "C", "D", "E", "F", "G", "H"]
                        if "Buy" in transaction_type:
                            cols += ["J", "K"]

                        for col in cols:
                            cell = ws[f"{col}{row_num}"]
                            cell.fill = PatternFill(patternType="solid", fill_type="solid", fgColor=DIVIDER_COLOR)
                        ws.row_dimensions[row_num].height = DIVIDER_HEIGHT
                        row_num += 1

                # Footes
                if bank_name_count > 1 or trade_total:
                    end_row = row_num - 1
                    cols = {
                        "E": "TOTAL",
                        "G": f"=SUM(G{start_row}:G{end_row})",
                        "H": f"=SUM(H{start_row}:H{end_row})"
                    }

                    for col, value in cols.items():
                        cell = ws[f"{col}{row_num}"]
                        cell.value = value
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center")
                        cell.font = Font(size=11)

                        if col == "G":
                            cell.number_format = "#,##0"
                        elif col == "H":
                            cell.number_format = f"[${ccy}] #,##0.00"

                row_num += 2

        return row_num

    def write_transfers(self, ws, row_num, transfers):
        # Section header
        cell = ws[f"A{row_num}"]
        cell.value = "Transfer of Stocks"
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal="center")

        cell = ws[f"J{row_num}"]
        cell.value = "Average"
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.merge_cells(f"J{row_num}:K{row_num}")

        ws.row_dimensions[row_num].height = HEADER_HEIGHT
        row_num += 1

        for i, t in enumerate(transfers):
            # Stock name line
            cell = ws[f"A{row_num}"]
            cell.value = f"{t['seq']}) {t['stock_name']} ({t['code']})"
            cell.font = Font(size=11, bold=True)
            ws.row_dimensions[row_num].height = HEADER_HEIGHT
            row_num += 1

            # From / to detail line — average of the receiving account after the transfer
            average = TradesDoneAverage(
                db=self.db, trade_date=self.trade_date,
                code=t['code'], bank_id=t['counter_bank_id'],
                include_transfers=True,
            ).average

            cell = ws[f"A{row_num}"]
            cell.value = f"from {t['bank_name']} to {t['counter_bank_name']}"
            cell.font = Font(size=11)

            cell = ws[f"H{row_num}"]
            cell.value = f"{t['quantity']:,.0f} shares"
            cell.font = Font(size=11)
            cell.alignment = Alignment(horizontal="center")

            cell = ws[f"J{row_num}"]
            cell.value = average if average else None
            cell.font = Font(size=11)
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "#,##0.0000"
            ws.merge_cells(f"J{row_num}:K{row_num}")

            ws.row_dimensions[row_num].height = HEADER_HEIGHT
            row_num += 1

            # Blank separator between items
            if i < len(transfers) - 1:
                row_num += 1

        return row_num
