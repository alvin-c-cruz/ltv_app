import datetime

from flask import current_app
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import get_column_letter

from .fixing_average import FixingAverage

FILL_COLOR = {
    "ACCU": "0099CC00",
    "DECU": "00FF8080",
}

GREEN_FONT = "00008000"
RED_FONT = "00FF0000"

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


class DownloadFixings:
    def __init__(self, db, trade_date, fixing_data):
        self.db = db

        self.knockouts = []

        self.trade_date = trade_date
        self.fixing_data = fixing_data

        self.template_file = os.path.join(current_app.instance_path, "excel_templates", "fixings.xlsx")
        self.filename = os.path.join(current_app.instance_path, "temp", f"{trade_date} Fixings.xlsx")

        self.account_ccy = ['Sheet']

        self.top_sheet_col_num = 1

        self.create_file()

    def create_file(self):
        wb = load_workbook(self.template_file)

        self.process_fixings(wb)
        WriteKnockouts(self.db, wb, self.knockouts)

        wb.save(self.filename)
        wb.close()

    def process_fixings(self, wb):
        for ccy, accounts in self.fixing_data.items():
            for account, fixings in accounts.items():
                sheet_name = f"{account}-{ccy}"
                self.account_ccy.append(sheet_name)
                ws = wb[sheet_name]
                self.write_fixings(ws, account, fixings)
                self.write_table(wb['Sheet'], account, fixings)

        for sheet_name in wb.sheetnames:
            if sheet_name not in self.account_ccy:
                wb[sheet_name].sheet_state = 'hidden'

    def separate_accu_decu(self, fixings):
        result = {
            "ACCU": [],
            "DECU": []
        }

        for _type in ("ACCU", "DECU"):
            for fixing in fixings:
                if fixing["transaction_type"] == _type:
                    result[_type].append(fixing)

        return result

    def write_fixings(self, ws, account_code, fixings):
        fixings = self.separate_accu_decu(fixings)

        # Write trade_date
        trade_date = datetime.datetime.strptime(self.trade_date, "%Y-%m-%d")
        cell = ws["D3"]
        cell.value = trade_date
        cell.number_format = "mmmm d, yyyy"

        row_num = 5
        code_row = {}
        for _type in ("ACCU", "DECU"):
            if fixings[_type]:
                #  Header
                cell = ws[f"A{row_num}"]
                cell.value = _type
                cell.font = Font(size=14, bold=True)
                cell.fill = PatternFill(patternType="solid", fill_type="solid", fgColor=FILL_COLOR[_type])

                row_num += 1

                cols = {
                    "A": "Underlying",
                    "B": "Code",
                    "C": "Daily Shares",
                    "D": "Spot",
                    "E": "Strike",
                    "F": "KO",
                    "G": f"Shares {_type.lower()}",
                    "H": "Settlement Amount",
                    "I": "Total Days",
                    "J": "Remaining Shares",
                    "K": "Settlement Date",
                    "L": "Next",
                    "M": "Days Double"
                }
                if _type == "ACCU":
                    cols["N"] = "Average"

                for col, value in cols.items():
                    cell = ws[f"{col}{row_num}"]
                    cell.value = value
                    cell.font = Font(size=11, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                    cell.border = thin_border

                # Extensions
                cols = {
                    "O": "",
                    "P": "Control",
                    "Q": "Beginning",
                    "R": "Average",
                    "S": "Shares",
                    "T": "Price",
                    "U": "Ending",
                    "V": "Average",
                    "W": "LTV Entry"
                }

                for col, value in cols.items():
                    cell = ws[f'{col}{row_num}']
                    cell.value = value

                ws.row_dimensions[row_num].height = 30
                row_num += 1

                #  Details
                for fixing in fixings[_type]:
                    #  Gather knockouts
                    if 'KO' in fixing["reference"]:
                        fixing["account_code"] = account_code
                        self.knockouts.append(fixing)

                    cols = {
                        "A": fixing["reference"],
                        "B": fixing["code"],
                        "C": fixing["shares"],
                        "D": fixing["spot"],
                        "E": fixing["strike"],
                        "F": fixing["ko"],
                        "G": '{:,.0f}'.format(fixing["shares_fixing"])
                             + " + "
                             + '{:,.0f}'.format(fixing["shares_indicative"])
                             + " = "
                             + '{:,.0f}'.format(fixing["shares_fixing"]
                                                + fixing["shares_indicative"]) if fixing["shares_indicative"]
                                                else '{:,.0f}'.format(fixing["shares_fixing"]),
                        "H": f"=ABS(S{row_num}*T{row_num})",
                        "I": '{:,.0f}'.format(fixing["days_fixing"])
                             + " + "
                             + '{:,.0f}'.format(fixing["days_indicative"])
                             + " = "
                             + '{:,.0f}'.format(fixing["days_fixing"]
                                                + fixing["days_indicative"]) if fixing["days_indicative"]
                                                else '{:,.0f}'.format(fixing["days_fixing"]),
                        "J": "",
                        "K": datetime.datetime.strptime(fixing["value_date"], "%Y-%m-%d"),
                        "L": fixing["next_date"] if fixing["next_date"] in ("Done", "KO")
                                    else datetime.datetime.strptime(fixing["next_date"], "%Y-%m-%d"),
                        "M": fixing["days_double"]
                    }
                    if _type == "ACCU":
                        cols["N"] = None

                    for col, value in cols.items():
                        cell = ws[f"{col}{row_num}"]
                        cell.value = value
                        cell.border = thin_border

                        # Font
                        if col in ("A", "B", "D", "E", "F", "H"):
                            cell.font = Font(size=11, bold=True)
                        elif col in ("L", ):
                            if value == "Done":
                                cell.font = Font(size=11, color=GREEN_FONT, bold=True)
                            elif value == "KO":
                                cell.font = Font(size=11, color=RED_FONT, bold=True)
                            else:
                                cell.font = Font(size=11)
                        else:
                            cell.font = Font(size=11)

                        # Alignment
                        if col in ("A", "C", "G", "I"):
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                        # Number format
                        if col in ("D", "E", "F", "N"):
                            cell.number_format = "#,##0.0000"
                        elif col in ("H",):
                            cell.number_format = "#,##0.00"
                        elif col in ("J",):
                            cell.number_format = "#,##0; [RED](#,##0)"
                        elif col in ("K", "L"):
                            cell.number_format = "d-mmm-yy"
                        else:
                            cell.number_format = "General"

                    # Extensions
                    cols = {
                        "O": f'=INDIRECT("O"&ROW()-1)+1',
                        "P": '',
                        "R": '',
                        "S": f'=IF(VALUE(D{row_num})>VALUE(E{row_num}),1,-1)*IF(ISNUMBER(VALUE(G{row_num})),'
                             f'VALUE(G{row_num}),VALUE(RIGHT(G{row_num},LEN(G{row_num})-FIND("=",G{row_num})-1)))',
                        "T": f'=E{row_num}',
                        "U": f'=Q{row_num}+S{row_num}',
                        "V": "",
                        "W": f'="("&TEXT($D$3,"m/d")&") "&IF(INDIRECT("A"&ROW()-O{row_num}-1)="ACCU",'
                             f'IF(L{row_num}="KO","Buy (Accu-KO) ","Buy (Accu) "),IF(L{row_num}="KO",'
                             f'"Sell (Decu-KO) ","Sell (Decu) "))&'
                             f'IF(ISNUMBER(VALUE(G{row_num})),TEXT(VALUE(G{row_num}),"#,###,##0"),'
                             f'TEXT(VALUE(RIGHT(G{row_num},LEN(G{row_num})-FIND("=",G{row_num}))),"#,###,##0"))&'
                             f'" @ "&TEXT(E{row_num},"#,##0.0000")&" = "&TEXT(J{row_num},"#,###,##0")'
                    }

                    if fixing["code"] not in code_row:
                        code_row[fixing["code"]] = {
                            "last_row": 0
                        }
                        cols["Q"] = get_balance(self.db, account_code, fixing["code"], self.trade_date)
                        if _type == "ACCU":
                            code_row[fixing["code"]]["accus"] = []
                            code_row[fixing["code"]]["accus"].append({
                                "quantity": fixing["shares_fixing"] + fixing["shares_indicative"],
                                "price": float(fixing["strike"])
                            })
                            cols["N"] = FixingAverage(self.db, self.trade_date, fixing["code"],
                                                      account_code, code_row[fixing["code"]]["accus"]).average()
                    else:
                        ws[f"J{code_row[fixing['code']]['last_row']}"].value = None  # Previous row with balance
                        cols["Q"] = f"=U{code_row[fixing['code']]['last_row']}"
                        if _type == "ACCU":
                            ws[f"N{code_row[fixing['code']]['last_row']}"].value = None  # Previous row with balance
                            code_row[fixing["code"]]["accus"].append({
                                "quantity": fixing["shares_fixing"] + fixing["shares_indicative"],
                                "price": float(fixing["strike"])
                            })
                            cols["N"] = FixingAverage(self.db, self.trade_date, fixing["code"],
                                                      account_code, code_row[fixing["code"]]["accus"]).average()

                    code_row[fixing['code']]['last_row'] = row_num
                    ws[f"J{row_num}"].value = f"=U{row_num}"

                    for col, value in cols.items():
                        cell = ws[f'{col}{row_num}']
                        cell.value = value

                    ws.row_dimensions[row_num].height = 60
                    row_num += 1

                row_num += 2

        ws.print_area = f"A1:N{row_num}"

    def write_table(self, ws, account_code, fixings):
        col_num = self.top_sheet_col_num

        for fixing in fixings:
            headers = [
                {"label": "Account", "value": account_code},
                {"label": "Product", "value": fixing["transaction_type"]},
                {"label": "Reference", "value": fixing["reference"]},
                {"label": "Code", "value": fixing["code"]},
                {"label": "Spot", "value": fixing["spot"]},
                {"label": "Strike", "value": fixing["strike"]},
                {"label": "KO", "value": fixing["ko"]},
                {"label": "Single", "value": fixing["single"]},
                {"label": "Double", "value": fixing["double"]},
            ]

            for row, header in enumerate(headers, 1):
                label = ws.cell(column=col_num, row=row)
                label.value = header["label"]

                value = ws.cell(column=col_num+1, row=row)
                value.value = header["value"]

                if value.value in ("ACCU", "DECU"):
                    value.fill = PatternFill(patternType="solid", fgColor=FILL_COLOR[value.value])

                if label.value in ("Spot", "Strike", "KO"):
                    value.number_format = "#,##0.0000"
                elif label.value in ("Single", "Double"):
                    value.number_format = "#,##0"

                for cell in (label, value):
                    cell.font = Font(size=12)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left")

            start_row = row_num = len(headers) + 2

            # Detail header
            date_control = ws.cell(column=col_num, row=row_num)
            date_control.value = "Date"

            closing_control = ws.cell(column=col_num + 1, row=row_num)
            closing_control.value = "Closing"

            shares_control = ws.cell(column=col_num + 2, row=row_num)
            shares_control.value = "Shares"

            for cell in (date_control, closing_control, shares_control):
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=12, bold=True)
                cell.border = thin_border

            row_num += 1

            for day_closing in fixing["days_closing"]:
                date, closing = day_closing

                date_control = ws.cell(column=col_num, row=row_num)
                date_control.value = date
                date_control.number_format = "dd-mmm-yyyy"

                closing_control = ws.cell(column=col_num+1, row=row_num)
                closing_control.value = closing
                closing_control.number_format = "#,##0.00"

                column_letter = get_column_letter(col_num + 1)

                shares_control = ws.cell(column=col_num+2, row=row_num)
                shares_control.number_format = "#,##0"
                if fixing["transaction_type"] == "ACCU":
                    shares_control.value = f'=IF({column_letter}7<={column_letter}{row_num},0,IF({column_letter}6>={column_letter}{row_num},{column_letter}9,{column_letter}8))'
                else:
                    shares_control.value = f'=IF({column_letter}7>={column_letter}{row_num},0,IF({column_letter}6<={column_letter}{row_num},{column_letter}9,{column_letter}8))'

                for cell in (date_control, closing_control, shares_control):
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=12)
                    cell.border = thin_border

                row_num += 1

            row_num += 2
            end_row = row_num - 1

            total_control = ws.cell(column=col_num, row=row_num)
            total_control.value = "Total"

            sum_control = ws.cell(column=col_num + 2, row=row_num)
            column_letter = get_column_letter(col_num + 2)
            sum_control.value = f'=SUM({column_letter}{start_row}:{column_letter}{end_row})'
            sum_control.number_format = "#,##0"
            sum_control.alignment = Alignment(horizontal="center")
            sum_control.font = Font(size=12)

            col_num += 4

        self.top_sheet_col_num = col_num


def get_balance(db, account_code, code, trade_date):
    sql = "SELECT SUM(tbl_transaction.quantity) FROM tbl_transaction " \
          "INNER JOIN tbl_bank_account ON tbl_bank_account.ref_num = tbl_transaction.bank_ref " \
          "INNER JOIN tbl_code ON tbl_code.ref_num = tbl_transaction.code_ref " \
          "WHERE tbl_bank_account.bank_id = ? AND tbl_code.code = ? AND tbl_transaction.trade_date < ?"

    result = db.execute(sql, (account_code, code, trade_date)).fetchone()

    if result:
        return result[0]
    else:
        return 0


class WriteKnockouts:
    def __init__(self, db, wb, knockouts):
        if knockouts:
            wb.create_sheet("Knockouts")
            self.db = db
            self.ws = wb["Knockouts"]
            self.knockouts = knockouts

            self.row_num = 0
            self.write_headers()
            self.write_details()

    def write_headers(self):
        self.row_num += 1
        cell = self.ws[f"A{self.row_num}"]
        cell.value = "Knock Out Summary"

        self.row_num += 1
        cell = self.ws[f"A{self.row_num}"]
        cell.value = f"Date: {datetime.datetime.now().strftime('%B %d, %Y')}"

        self.row_num += 2
        columns = {
            "A": "No.",
            "B": "TYPE",
            "C": "BANK ACCOUNT",
            "D": "STOCK",
            "E": "CODE",
            "F": "SHARES",
            "G": "SPOT",
            "H": "STRIKE",
            "I": "KO",
            "J": "Closing",
            "K": "SHARES",
        }

        for column_letter, value in columns.items():
            cell = self.ws[f"{column_letter}{self.row_num}"]
            cell.value = value

        self.row_num += 1
        self.ws.row_dimensions[self.row_num].height = 0.01

    def write_details(self):
        for fixing in self.knockouts:
            account = self.db.execute("SELECT bank_name FROM tbl_bank_account WHERE bank_id=?;",
                                      (fixing["account_code"], )).fetchone()[0]
            self.row_num += 1
            columns = {
                "A": f'=INDIRECT("A"&ROW()-1)+1',
                "B": fixing["transaction_type"],
                "C": account,
                "D": fixing["stock_name"],
                "E": fixing["code"],
                "F": fixing["shares"],
                "G": fixing["spot"],
                "H": fixing["strike"],
                "I": fixing["ko"],
                "J": "",
                "K": fixing["shares_fixing"],
            }
            for column_letter, value in columns.items():
                cell = self.ws[f"{column_letter}{self.row_num}"]
                cell.value = value
