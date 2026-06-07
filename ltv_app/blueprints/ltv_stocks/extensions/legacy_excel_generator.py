"""
Legacy-style Excel generation for LTV Stocks Report.

This module ports the Excel generation logic from localhost/modules/ltv_stocks2.py
to work with the modern ltv_app infrastructure while maintaining the exact same
Excel output format that users are accustomed to.

Key features replicated from legacy:
- Active contract count formulas
- Price logic explanations (Strike < Close = x1, Strike > Close = x2)
- Bank Reference column
- Indicator columns (D/./KO/Done) based on price vs strike/KO
- Stock-specific colors (Alibaba=teal, Tencent=yellow, etc.)
- HKD cross-account aggregation formulas
- Grey fill for non-trading days/holidays
- Recent transactions in position section
- Complex date range (previous full week + current week to date)
"""

import io
import datetime
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class LegacyLTVStocksExcel:
    """
    Generates Excel workbook in legacy format matching localhost/modules/ltv_stocks2.py output.
    """

    def __init__(self, data: dict, report_date: date, db):
        self.data = data
        self.report_date = report_date
        self.db = db
        self.wb = None
        self.ws = None

        # Track HKD count cells for cross-sheet aggregation
        self.HKD_count = {'total_row': {}}
        self.PRIMARY_BANK_ACCOUNT = 'DBPe'

        # Calculate date range (previous full week Mon-Fri + current week to report_date)
        self.date_range = self._calculate_date_range()
        self.str_dates = [str(d)[:10] for d in self.date_range]

        # Load stock prices for all dates
        self.stock_price_cache = self._load_stock_prices()

        # Load holidays
        self.holidays = self._load_holidays()

        # Bank name mapping
        self.bank_name_map = {
            "CB1": "CITIBANK",
            "CB2": "CITIBANK",
            "CB3": "CITIBANK",
            "CBBH": "CITIBANK",
            "CBBH2": "CITIBANK",
            "CBSG": "CITIBANK",
            "BOS": "Bank of Singapore",
            "DBPe": "DEUTSCHE BANK",
            "DBPL": "DEUTSCHE BANK",
            "SC": "STANDARD CHARTERED",
            "SHK": "SUN HUNG KAI Account No. 1",
            "SHK2": "SUN HUNG KAI Account No. 2",
            "MST1": "MORGAN STANLEY",
            "MST2": "MORGAN STANLEY",
            "MSPL": "MORGAN STANLEY",
            "NSG": "NOMURA SINGAPORE"
        }

        # Sub-title configuration
        self.sub_title = {
            'CB2': {'title': 'ACCOUNT # 2 (REALGOLD)', 'color': 'FF7030A0'},
            'CB3': {'title': 'ACCOUNT # 3', 'color': 'FF0070C0'},
            'CBBH': {'title': 'BERRY HILL Account', 'color': '00FF6600'},
            'CBBH2': {'title': 'BERRY HILL Account 2', 'color': '00FF6600'},
            'CBSG': {'title': 'Singapore Account No. 1', 'color': 'FF7030A0'},
            'DBPL': {'title': 'PERFECT LEGEND HOLDINGS w/ Lucio Yan', 'color': '00000000'},
            'MST1': {'title': 'Titan Account No. 1', 'color': '00000000'},
            'MST2': {'title': 'ACCOUNT NO. 2 (Titan) - PERSONAL', 'color': '00000000'},
            'MSPL': {'title': '(Perfect Legend)', 'color': '00000000'},
        }

        # Account name mapping for position section
        self.account_names = {
            "CB1": {"HKD": "Citibank Account No. 1 Stocks"},
            "CB2": "Citibank Account No. 2 Stocks",
            "CB3": "Citibank Account No. 3 Stocks",
            "CBBH": "Citibank Berry Hill Stocks",
            "CBBH2": "Citibank Berry Hill No. 2 Stocks",
            "CBSG": "Citibank Singapore Account No. 1 Stocks",
            "BOS": "Bank of Singapore Stocks",
            "DBPe": "DEUTSCHE PERSONAL Stocks",
            "DBPL": "DEUTSCHE PERFECT LEGEND Stocks",
            "SC": "Standard Chartered Stocks",
            "SHK": "Sun Hung Kai Account No. 1 Stocks",
            "SHK2": "Sun Hung Kai Account No. 2 Stocks",
            "MST1": "MORGAN TITAN No. 1 Stocks",
            "MST2": "MORGAN TITAN No. 2 Stocks",
            "MSPL": "MORGAN PERFECT LEGEND Stocks",
            "NSG": "NOMURA SINGAPORE"
        }

    def _calculate_date_range(self):
        """
        Calculate 10-day date range: previous full week (Mon-Fri) + current week to report_date.

        Example: report_date = Tuesday, June 2, 2026
        Returns: [May 25 (Mon), May 26 (Tue), May 27 (Wed), May 28 (Thu), May 29 (Fri),
                  June 1 (Mon), June 2 (Tue)]
        """
        # Go back to previous Monday
        start = self.report_date - timedelta(days=6 + self.report_date.weekday())

        return [
            start,
            start + timedelta(days=1),
            start + timedelta(days=2),
            start + timedelta(days=3),
            start + timedelta(days=4),
            start + timedelta(days=7),
            start + timedelta(days=8),
            start + timedelta(days=9),
            start + timedelta(days=10),
            start + timedelta(days=11),
        ]

    def _load_stock_prices(self):
        """Load all stock prices for the date range into cache."""
        cache = {}

        for date_str in self.str_dates:
            rows = self.db.execute("""
                SELECT code_ref, closing_price
                FROM tbl_stock_price
                WHERE trade_date = ?
            """, (date_str,)).fetchall()

            for row in rows:
                key = (row['code_ref'], date_str)
                cache[key] = row['closing_price']

        return cache

    def _load_holidays(self):
        """Load all holidays for quick lookup."""
        rows = self.db.execute("""
            SELECT holi_date, tbl_currency.ccy_id
            FROM tbl_holiday
            INNER JOIN tbl_currency ON tbl_currency.ref_num = tbl_holiday.ccy_ref
        """).fetchall()

        return {(str(r['holi_date'])[:10], r['ccy_id']) for r in rows}

    def _is_holiday(self, date_obj, ccy='HKD'):
        """Check if date is a holiday."""
        return (str(date_obj)[:10], ccy) in self.holidays

    def _is_working_day(self, date_obj, ccy='HKD'):
        """Check if date is a working day."""
        if isinstance(date_obj, str):
            date_obj = datetime.datetime.strptime(date_obj[:10], '%Y-%m-%d')
        return not self._is_holiday(date_obj, ccy) and date_obj.isoweekday() not in (6, 7)

    def get_stock_price(self, code_ref, date_str):
        """Get stock price from cache."""
        return self.stock_price_cache.get((code_ref, date_str))

    def get_stock_price_fallback(self, code_ref, date_str):
        """Get stock price with fallback to previous days."""
        price = self.get_stock_price(code_ref, date_str)

        if price is not None:
            return price

        # Fallback to previous days
        d = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        for _ in range(10):  # Try up to 10 days back
            d = d - timedelta(days=1)
            if self._is_working_day(d):
                price = self.get_stock_price(code_ref, str(d)[:10])
                if price is not None:
                    return price

        return None

    def generate(self) -> io.BytesIO:
        """Generate Excel workbook and return as BytesIO."""
        self.wb = Workbook()
        del self.wb['Sheet']

        # Create helper sheets
        self.wb.create_sheet("closing_price")
        self.wb.create_sheet("record")

        # Process each currency and bank account
        for ccy_id in ['HKD', 'SGD']:  # Process HKD first for aggregation
            if ccy_id not in self.data:
                continue

            for bank_id in sorted(self.data[ccy_id].keys()):
                bank_data = self.data[ccy_id][bank_id]

                # Skip if no data
                if not bank_data['accu'] and not bank_data['decu'] and not bank_data['positions']:
                    continue

                sheet_name = f'{bank_id}-{ccy_id}'
                self.wb.create_sheet(sheet_name)
                self.ws = self.wb[sheet_name]

                # Configure print settings
                self.ws.page_setup.paperSize = 4  # A4
                self.ws.page_setup.orientation = 'landscape'
                self.ws.page_margins.left = 0.1
                self.ws.page_margins.right = 0.1
                self.ws.page_margins.top = 0.15
                self.ws.page_margins.bottom = 0.1
                self.ws.page_margins.header = 0.0
                self.ws.page_margins.footer = 0.0

                # Set column widths
                self._set_column_widths()

                # Build sheet content
                row_num = 1
                row_num = self._report_header(row_num, bank_id, ccy_id, bank_data)
                row_num = self._contract_section(row_num, 'ACCU', bank_data['accu'], ccy_id, bank_id)
                row_num = self._contract_section(row_num, 'DECU', bank_data['decu'], ccy_id, bank_id)
                row_num = self._position_section(row_num, bank_data['positions'], ccy_id, bank_id, bank_data)

                self.ws.print_area = f"A1:X{row_num}"

        # Add HKD aggregation formulas to primary account
        self._add_hkd_aggregation()

        # Save to BytesIO
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    def _set_column_widths(self):
        """Set column widths matching legacy format."""
        widths = {
            'A': 28.28, 'B': 5.57, 'D': 12.42, 'E': 8.28, 'F': 8.28,
            'G': 8.28, 'H': 9.85, 'I': 9.85, 'J': 5.71, 'K': 5.71,
            'L': 5.71, 'N': 9.85, 'O': 6.42, 'P': 6.42, 'Q': 6.42,
            'R': 6.42, 'S': 6.42, 'T': 6.42, 'U': 6.42, 'V': 6.42,
            'W': 6.42, 'X': 6.42, 'Z': 6.42, 'AA': 6.42, 'AB': 6.42,
            'AC': 6.42, 'AD': 6.42, 'AE': 6.42, 'AF': 6.42, 'AG': 6.42,
            'AH': 6.42, 'AI': 6.42, 'AJ': 6.42,
        }

        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width

        # Hide columns C and M
        self.ws.column_dimensions['C'].hidden = True
        self.ws.column_dimensions['M'].hidden = True

    def _report_header(self, row_num, bank_id, ccy_id, bank_data):
        """Generate report header."""
        # Main header
        bank_name = bank_data['bank_name'] or self.bank_name_map.get(bank_id, bank_id)
        date_str = self.report_date.strftime("%B %d, %Y")

        cell = self.ws[f'A{row_num}']
        cell.value = f'{bank_name} as of {date_str}'
        cell.font = Font(name='Arial', size=16, bold=True)

        cell = self.ws[f'H{row_num}']
        cell.value = 'updated'
        cell.font = Font(name='Arial', size=16, bold=True)
        cell.alignment = Alignment(horizontal='right')

        cell = self.ws[f'I{row_num}']
        cell.value = self.date_range[5]  # Monday of current week
        cell.font = Font(name='Arial', size=16, bold=True)
        cell.number_format = 'd-mmm-yyyy'
        self.ws.merge_cells(f'I{row_num}:K{row_num}')

        row_num += 1

        # Sub-header
        if bank_id in self.sub_title:
            cell = self.ws[f'A{row_num}']
            cell.value = self.sub_title[bank_id]['title']
            color = self.sub_title[bank_id]['color']
            cell.font = Font(name='Arial', size=14, bold=True, color=color)

            # Special fills for specific banks
            if bank_id == 'DBPL':
                for col in ('A', 'B', 'C', 'D', 'E'):
                    c = self.ws[f'{col}{row_num}']
                    c.fill = PatternFill('solid', fgColor='00FFFF00')
                    c.font = Font(name='Arial', size=12, bold=True, color=color)
            elif bank_id == 'MST1':
                self.ws[f'A{row_num}'].fill = PatternFill('solid', fgColor='FFF6DDC2')
            elif bank_id == 'MST2':
                for col in ('A', 'B', 'C', 'D', 'E'):
                    self.ws[f'{col}{row_num}'].fill = PatternFill('solid', fgColor='00666699')
            elif bank_id == 'MSPL':
                self.ws[f'A{row_num}'].fill = PatternFill('solid', fgColor='00FFFF00')

            row_num += 1

        # Currency sub-header for non-HKD
        if ccy_id in ('USD', 'SGD'):
            cell = self.ws[f'A{row_num}']
            cell.value = f"{ccy_id} STOCKS"
            cell.font = Font(name='Arial', size=18, bold=True, color='00000000')
            row_num += 1

        row_num += 2
        return row_num

    # This file continues but is getting large. Let me break it into parts.
    # For now, I'll save this and create the remaining methods in the next part.
