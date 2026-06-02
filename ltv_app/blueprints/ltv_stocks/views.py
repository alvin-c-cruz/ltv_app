import io
from datetime import date

from flask import Blueprint, render_template, request, send_file

from .. auth import login_required
from .. database import get_db
from ...tz import ph_today
from .create_ltv_stocks import get_ltv_stocks, get_ltv_stocks_full

bp = Blueprint('ltv_stocks', __name__, template_folder='pages', url_prefix='/ltv-stocks')


def _parse_date(val: str, default: date) -> date:
    try:
        return date.fromisoformat(val)
    except (TypeError, ValueError):
        return default


@bp.route('/', methods=['GET', 'POST'])
@login_required
def home():
    today = ph_today()
    if request.method == 'POST':
        report_date = _parse_date(request.form.get('report_date'), today)
    else:
        report_date = _parse_date(request.args.get('report_date'), today)

    db   = get_db()
    data = get_ltv_stocks(db, report_date)

    return render_template('ltv_stocks/home.html',
                           data=data,
                           report_date=str(report_date))


@bp.route('/download', methods=['POST'])
@login_required
def download():
    today = ph_today()
    report_date = _parse_date(request.form.get('report_date'), today)

    db   = get_db()
    data, price_map_multi, trading_dates = get_ltv_stocks_full(db, report_date)

    output = _generate_excel(data, report_date, price_map_multi, trading_dates)
    filename = f"{report_date} LTV Stocks.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ---------------------------------------------------------------------------
# Excel generation (self-contained, no external templates needed)
# ---------------------------------------------------------------------------

def _generate_excel(data: dict, report_date: date,
                    price_map_multi: dict, trading_dates: list) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    BOLD       = Font(name='Arial', bold=True, size=10)
    BOLD_SM    = Font(name='Arial', bold=True, size=8)
    NORM       = Font(name='Arial', size=10)
    NORM_SM    = Font(name='Arial', size=9)
    HEAD       = Font(name='Arial', bold=True, size=16)
    FONT_WHITE = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    LFT_NOWRAP = Alignment(horizontal='left', vertical='center', wrap_text=False)
    FILL_ACCU  = PatternFill('solid', fgColor='1F3864')   # dark navy for ACCU
    FILL_DECU  = PatternFill('solid', fgColor='375623')   # dark green for DECU
    FILL_POS   = PatternFill('solid', fgColor='833C00')   # dark brown for positions
    FILL_HDR   = PatternFill('solid', fgColor='BDD7EE')   # light blue column headers
    FILL_CLOSE = PatternFill('solid', fgColor='DDEBF7')   # lighter blue for close price headers
    thin       = Side(style='thin', color='AAAAAA')
    BOX        = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR        = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT        = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    wb = Workbook()
    wb.remove(wb.active)

    for ccy_id, banks in data.items():
        for bank_id, bank_data in banks.items():
            if not bank_data['accu'] and not bank_data['decu'] and not bank_data['positions']:
                continue

            sheet_name = f"{bank_id}-{ccy_id}"[:31]
            ws = wb.create_sheet(sheet_name)
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize   = 4  # A4

            row = 1
            date_str = f"{report_date.strftime('%B')} {report_date.day}, {report_date.year}"
            header_name = bank_data['report_label'] or bank_data['bank_name']
            c = ws.cell(row, 1, f"{header_name} as of {date_str}")
            c.font = HEAD; c.alignment = LFT_NOWRAP
            row += 2

            row = _xl_contracts(ws, bank_data['accu'], 'ACCUMULATOR', row,
                                 price_map_multi, trading_dates,
                                 BOLD, BOLD_SM, NORM, NORM_SM, FILL_ACCU, FILL_HDR, FILL_CLOSE,
                                 FONT_WHITE, BOX, CTR, LFT)
            row += 1

            row = _xl_contracts(ws, bank_data['decu'], 'DECUMULATOR', row,
                                 price_map_multi, trading_dates,
                                 BOLD, BOLD_SM, NORM, NORM_SM, FILL_DECU, FILL_HDR, FILL_CLOSE,
                                 FONT_WHITE, BOX, CTR, LFT)
            row += 1

            row = _xl_positions(ws, bank_data['positions'], report_date, row,
                                 BOLD, NORM, FILL_POS, FILL_HDR, FONT_WHITE, BOX, CTR, LFT)

            # Fixed columns A-M; date columns N onwards at 8 wide each
            from openpyxl.utils import get_column_letter
            base_widths = {'A': 26, 'B': 7, 'C': 12, 'D': 10, 'E': 10,
                           'F': 10, 'G': 10, 'H': 11, 'I': 11, 'J': 8, 'K': 8, 'L': 8, 'M': 12}
            for col, w in base_widths.items():
                ws.column_dimensions[col].width = w
            for i in range(len(trading_dates)):
                ws.column_dimensions[get_column_letter(14 + i)].width = 8

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _xl_contracts(ws, contracts, title, row,
                  price_map_multi, trading_dates,
                  BOLD, BOLD_SM, NORM, NORM_SM, FILL_TITLE, FILL_HDR, FILL_CLOSE,
                  FONT_WHITE, BOX, CTR, LFT):
    from openpyxl.styles import Font, PatternFill
    from datetime import date as _date

    N_dates = len(trading_dates)
    # Columns: A=Name B=Code C=Bank Ref D=Shares/Day E=Spot F=Strike G=K/O H=Start I=End
    #          J=RCVD K=REM L=Total M=Next  N+=closing dates
    TOTAL_COLS = 13 + N_dates

    # Stock-specific colors (legacy feature)
    STOCK_COLORS = {
        '2333': 'FFE5E39F',  # Light tan
        '0700': '00FFFFCC',  # Light yellow (Tencent)
        '1024': '009999FF',  # Light blue
        '0388': '00CC99FF',  # Light purple
        '3993': '00FFCC99',  # Light orange
        '0175': '00CCFFFF',  # Light cyan
        '9988': '00008080',  # Teal (Alibaba)
    }

    # ── Row 1: Active contract count formula (legacy feature) ────────
    if contracts:
        count_row = row
        next_col_start = count_row + 4
        next_col_end = next_col_start + len(contracts) - 1
        formula = f'={len(contracts)}-COUNTIF(M{next_col_start}:M{next_col_end},"*DONE*")'
        c = ws.cell(row, 1, formula)
        c.font = Font(name='Arial', size=7, bold=False)
        c.number_format = '0'
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 10.5
        row += 1

    # ── Row 2: section title ──────────────────────────────────────────
    c = ws.cell(row, 1, title)
    c.font = FONT_WHITE; c.fill = FILL_TITLE; c.alignment = LFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    row += 1

    # ── Row 3: main column headers ────────────────────────────────────
    single_hdrs = {1: 'Stock Name', 2: 'Code', 3: 'Bank Reference',
                   4: 'Shares / Day', 5: 'Spot Price', 6: 'Strike Price',
                   7: 'K/O Price', 8: 'Start Date', 9: 'End Date', 13: 'Date of Next Mo.'}
    for col, h in single_hdrs.items():
        c = ws.cell(row, col, h)
        c.font = BOLD_SM; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)

    c = ws.cell(row, 10, 'Life Term of Contract')
    c.font = BOLD_SM; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=12)

    if N_dates > 0:
        c = ws.cell(row, 14, 'Closing Price')
        c.font = BOLD_SM; c.fill = FILL_CLOSE; c.alignment = CTR; c.border = BOX
        ws.merge_cells(start_row=row, start_column=14, end_row=row, end_column=13 + N_dates)
    row += 1

    # ── Row 4: sub-headers ────────────────────────────────────────────
    for col, h in [(10, 'Rcvd mos.'), (11, 'Rem mos.'), (12, 'Total mos.')]:
        c = ws.cell(row, col, h)
        c.font = BOLD_SM; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX

    for i, d in enumerate(trading_dates):
        c = ws.cell(row, 14 + i, d)
        c.font = BOLD_SM; c.fill = FILL_CLOSE; c.alignment = CTR; c.border = BOX
        c.number_format = 'd-mmm'
    row += 1

    # ── Data rows ─────────────────────────────────────────────────────
    if not contracts:
        c = ws.cell(row, 1, 'No active contracts.')
        c.font = NORM; c.alignment = LFT
        row += 1
        return row

    STRIKE_FONT = Font(name='Arial', bold=True, size=10)

    for ct in contracts:
        if ct['is_done']:
            next_val, next_fmt = 'DONE', None
        else:
            try:
                next_val = _date.fromisoformat(ct['next_date'])
                next_fmt = 'd-mmm-yy'
            except (ValueError, AttributeError):
                next_val, next_fmt = ct['next_date'], None

        entries = [
            (1,  ct['stock_name'],       NORM,        LFT, None),
            (2,  ct['code'],             NORM,        CTR, '@'),
            (3,  ct['bank_doc'],         NORM,        CTR, None),
            (4,  ct['shares_day'],       NORM,        CTR, None),
            (5,  ct['spot_raw'],         NORM,        CTR, '#,##0.0000'),
            (6,  ct['strike_raw'],       STRIKE_FONT, CTR, '#,##0.0000'),
            (7,  ct['ko_raw'],           NORM,        CTR, '#,##0.0000'),
            (8,  ct['start_date_raw'],   NORM,        CTR, 'd-mmm-yy'),
            (9,  ct['last_end_date_raw'],NORM,        CTR, 'd-mmm-yy'),
            (10, ct['received_months'],  NORM,        CTR, '0.0'),
            (11, ct['remaining_months'], NORM,        CTR, '0.0'),
            (12, ct['total_months'],     NORM,        CTR, '0.0'),
            (13, next_val,               NORM,        CTR, next_fmt),
        ]
        for col, val, fnt, aln, fmt in entries:
            c = ws.cell(row, col, val)
            c.font = fnt; c.alignment = aln; c.border = BOX
            if fmt:
                c.number_format = fmt

            # Apply stock-specific colors to column A (Stock Name)
            if col == 1 and ct['code'] in STOCK_COLORS:
                c.fill = PatternFill('solid', fgColor=STOCK_COLORS[ct['code']])

        code_prices = price_map_multi.get(ct['code_ref'], {})
        for i, d in enumerate(trading_dates):
            price = code_prices.get(str(d))
            c = ws.cell(row, 14 + i, price)
            c.font = NORM; c.alignment = CTR; c.border = BOX
            if price is not None:
                c.number_format = '#,##0.00'

        row += 1

    return row


def _xl_positions(ws, positions, report_date, row,
                  BOLD, NORM, FILL_TITLE, FILL_HDR, FONT_WHITE, BOX, CTR, LFT):
    from openpyxl.styles import Font
    BOLD_SM = Font(name='Arial', bold=True, size=8)

    # ── Row 1: section title ──────────────────────────────────────────
    c = ws.cell(row, 1, f'STOCK POSITIONS  (as of {report_date})')
    c.font = FONT_WHITE; c.fill = FILL_TITLE; c.alignment = LFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    # ── Row 2: column headers ─────────────────────────────────────────
    for col, h in [(1, 'Stock Name'), (2, 'Code'), (3, 'Unblocked'),
                   (4, 'Blocked'), (5, 'Total Shares'), (6, 'Ave. Price'),
                   (7, '% Inc./Dec.'), (8, 'Closing Price')]:
        c = ws.cell(row, col, h)
        c.font = BOLD_SM; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX
    row += 1

    if not positions:
        c = ws.cell(row, 1, 'No positions.')
        c.font = NORM; c.alignment = LFT
        row += 1
        return row

    for code, pos in sorted(positions.items()):
        entries = [
            (1, pos['stock_name'],                                         LFT, None),
            (2, code,                                                      CTR, '@'),
            (3, pos['unblocked'],                                          CTR, '#,##0'),
            (4, pos['blocked'] or None,                                    CTR, '#,##0'),
            (5, pos['balance'],                                            CTR, '#,##0'),
            (6, round(pos['average'], 4) if pos['average'] else None,      CTR, '#,##0.0000'),
            (7, pos['pct_chg'] if pos['pct_chg'] is not None else None,    CTR, '0.00%'),
            (8, pos['closing'] if pos['closing'] is not None else None,    CTR, '#,##0.00'),
        ]
        for col, val, aln, fmt in entries:
            c = ws.cell(row, col, val)
            c.font = NORM; c.alignment = aln; c.border = BOX
            if fmt and val is not None:
                c.number_format = fmt
        ws.row_dimensions[row].height = 20
        row += 1

    return row
