"""Excel writer — contract + positions tables, and full workbook assembly.

Ports `localhost/modules/ltv_stocks2.py`'s `contract()` method (lines 410-709,
`_write_contracts` here), `position()` (712-914, `_write_positions`), `create()`
(211-295, `build_workbook`), `report_header()` (342-407, `report_header`) and
`column_width()` (298-339, `_set_column_widths`) cell-for-cell, plus a
`_write_status_grid` knock-out/strike tracking grid (Z:AJ, see docs/
superpowers/specs/2026-07-17-status-ko-tracking-grid-design.md) added on top
of the port. The off-print-area helper columns `Y`-`AC` (the reconciliation
helper columns, a separate concept from the status grid) remain dropped —
see docs/superpowers/specs/2026-07-06-ltv-stocks-legacy-exact-replica-design.md,
"The Excel writer".

`build_workbook(db, report_date, bank_ids)` is the public entry point.
"""

import io
import re
import zipfile
from datetime import date, timedelta
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import column_index_from_string

from .term_sheet_calc import _FREQ_DIV, contract_records
from .positions_calc import position_records, _average, _transactions_narrative
from .working_day import WorkingDay, position_start_date
from .stock_price import get_stock_price


# --- Style constants (port of ltv_stocks2's xl_font/xl_fill/xl_box/xl_align lambdas) ---

def xl_font(size=10, bold=False, name='Arial'):
    return Font(name=name, size=size, bold=bold)


def xl_font_color(size=10, color='00000000', bold=True, name='Arial'):
    return Font(name=name, size=size, bold=bold, color=color)


def xl_align(wrap_text=False, horizontal='center', vertical='center'):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


def xl_align_shrink(shrink_to_fit=True, vertical='center'):
    return Alignment(vertical=vertical, shrink_to_fit=shrink_to_fit)


def xl_box():
    side = Side(style='thin')
    return Border(left=side, right=side, top=side, bottom=side)


def xl_fill(fill_color):
    return PatternFill(patternType='solid', fgColor=fill_color)


# Per-code background fills for column A (ltv_stocks2.py ~688-704).
_CODE_FILLS = {
    '2333': 'FFE5E39F',
    '0700': '00FFFFCC',
    '1024': '009999FF',
    '0388': '00CC99FF',
    '3993': '00FFCC99',
    '0175': '00CCFFFF',
    '9988': '00008080',
}
_SMALL_FONT_CODES = ('0981', '2196')

_GREY_FILL = '00C0C0C0'

# O-X column -> offset into the 10-date range.
_COL_DATE_OFFSET = {
    'O': 0, 'P': 1, 'Q': 2, 'R': 3, 'S': 4,
    'T': 5, 'U': 6, 'V': 7, 'W': 8, 'X': 9,
}
_OX_COLS = ('O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X')
# AA-AJ per-day status columns, paired 1:1 with _OX_COLS (AA reads the same
# day as O, AB as P, ... AJ as X).
_STATUS_COLS = ('AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ')
_ALL_DATA_COLS = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N') + _OX_COLS
_ZERO_ROW_COLS = ('A', 'B', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X')


def week_dates(report_date: date) -> list:
    """The 10-date range: previous Mon-Fri + current Mon-Fri, built with isoweekday()."""
    start = report_date - timedelta(days=6 + report_date.isoweekday())
    offsets = (0, 1, 2, 3, 4, 7, 8, 9, 10, 11)
    return [start + timedelta(days=o) for o in offsets]


def _write_contracts(ws, records, product, row, report_date, date_range, wd, price_lookup,
                      bank_id=None):
    """Write one ACCU/DECU contract table starting at `row`. Returns the next free row.

    `wd` supplies `.is_holiday(date) -> bool` for the O-X grid (a `WorkingDay`
    instance scoped to the table's currency). `price_lookup(code_ref, date) ->
    float|None` supplies the O-X historical closing prices. `bank_id` drives the
    " bi-weekly" stock-name suffix (ltv_stocks2.py:620-621): banks other than
    DBPe/DBPL get it appended whenever the contract's J/K/L number format is the
    non-monthly '0.0' (i.e. any frequency other than 'monthly').
    """
    ws.column_dimensions['C'].hidden = True
    ws.column_dimensions['M'].hidden = True

    n = len(records)
    count_row = row
    first_data_row = count_row + 4
    last_data_row = first_data_row + n - 1

    # Head Count
    cell = ws[f'A{count_row}']
    cell.value = f'={n}-COUNTIF(N{first_data_row}:N{last_data_row},"*DONE*")'
    cell.font = xl_font(7)
    cell.number_format = '0'
    cell.alignment = xl_align(False, 'left')
    ws.row_dimensions[count_row].height = 10.5

    row = count_row + 1

    # Head Line 1
    labels = {
        'A': f'{product}MULATOR',
        'B': 'CODE',
        'C': 'Bank Reference',
        'D': 'SHARES / DAY',
        'E': 'SPOT PRICE',
        'F': 'STRIKE PRICE',
        'G': 'K/O PRICE',
        'H': 'START DATE',
        'I': 'END DATE',
        'J': 'LIFE TERM OF CONTRACT',
        'N': f'DATE OF {product}MULATOR',
        'O': 'CLOSING PRICE',
        'T': 'CLOSING PRICE',
    }
    for col, val in labels.items():
        cell = ws[f'{col}{row}']
        cell.value = val
        cell.font = xl_font(6, True) if col in ('J', 'N') else xl_font(8, True)
        cell.alignment = xl_align(True) if col in ('E', 'F', 'G', 'J', 'N') else xl_align()
        cell.border = xl_box()

    for col in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'):
        ws.merge_cells(f'{col}{row}:{col}{row + 2}')
    ws.merge_cells(f'J{row}:L{row}')
    ws.merge_cells(f'N{row}:N{row + 1}')
    ws.merge_cells(f'O{row}:S{row}')
    ws.merge_cells(f'T{row}:X{row}')

    row += 1

    # Head Line 2
    labels = {
        'A': ('Strike Price < Closing Price = QTYx1' if product == 'ACCU'
              else 'Strike Price > Closing Price = QTYx1'),
        'J': 'RCVD mos.',
        'K': 'REM mos.',
        'L': 'Total mos.',
    }
    for col, offset in _COL_DATE_OFFSET.items():
        labels[col] = date_range[offset]

    for col, val in labels.items():
        cell = ws[f'{col}{row}']
        cell.value = val
        cell.font = xl_font(7, True)
        cell.alignment = xl_align(True)
        cell.border = xl_box()
        if col in _OX_COLS:
            cell.font = xl_font(9, True)
            cell.number_format = 'm/d'

    for col in ('J', 'K', 'L') + _OX_COLS:
        ws.merge_cells(f'{col}{row}:{col}{row + 1}')

    row += 1

    # Head Line 3
    labels = {
        'A': ('Strike Price > Closing Price = QTYx2' if product == 'ACCU'
              else 'Strike Price < Closing Price = QTYx2'),
        'N': 'NEXT MO.',
    }
    for col, val in labels.items():
        cell = ws[f'{col}{row}']
        cell.value = val
        cell.font = xl_font(7, True)
        cell.alignment = xl_align(True)
        cell.border = xl_box()

    row += 1  # == first_data_row

    if n == 0:
        for col in _ZERO_ROW_COLS:
            ws[f'{col}{row}'].border = xl_box()
        ws.row_dimensions[row].height = 17
    else:
        for rec in records:
            r = row
            done = False
            divisor = _FREQ_DIV.get(rec.get('frequency'), 2)

            cols = {
                'A': rec['stock_name'],
                'B': rec['code'],
                'C': rec['bank_doc'],
                'D': rec['shares'],
                'E': rec['spot'],
                'F': rec['strike'],
                'G': rec['ko'],
                'H': rec['start_date'],
                'I': rec['end_date'],
                'J': rec['received'],
                'K': f'=L{r}-J{r}',
                'L': rec['total'],
                'M': rec['yahoo_ticker'],
                'N': rec['next_date'],
            }
            for col, offset in _COL_DATE_OFFSET.items():
                cols[col] = price_lookup(rec['code_ref'], date_range[offset])

            ws.row_dimensions[r].height = 17 if rec.get('ccy_id') == 'HKD' else 25.5

            for col in _ALL_DATA_COLS:
                cell = ws[f'{col}{r}']
                cell.font = xl_font(10) if col != 'F' else xl_font(10, True)
                cell.alignment = xl_align(False) if col != 'A' else xl_align(False, 'left')

                if col in ('E', 'F', 'G'):
                    cell.number_format = '#,##0.0000'
                elif col in ('J', 'K', 'L'):
                    if divisor == 1:
                        cell.number_format = '0'
                    else:
                        cell.number_format = '0.0'
                        if col == 'J' and bank_id not in ('DBPe', 'DBPL'):
                            a_cell = ws[f'A{r}']
                            a_cell.value = a_cell.value + ' bi-weekly'
                elif col == 'B':
                    cell.number_format = '@'
                elif col in ('H', 'I', 'N'):
                    cell.font = xl_font(9)
                    cell.number_format = 'd-mmm-yy'
                elif col in _OX_COLS:
                    cell.number_format = '#,##0.00'
                cell.border = xl_box()

                if col == 'N':
                    if rec['received'] != rec['total']:
                        cell.value = cols[col]
                    else:
                        cell.value = 'DONE'
                        cell.font = xl_font(10, True)
                elif col in _OX_COLS:
                    d = date_range[_COL_DATE_OFFSET[col]]
                    if d < rec['start_date']:
                        cell.fill = xl_fill(_GREY_FILL)
                    elif d > rec['end_date']:
                        k = _COL_DATE_OFFSET[col]
                        if k == 4:
                            if d == rec['end_date'] + timedelta(days=3):
                                cell.value = 'Done'
                                done = True
                                cell.font = xl_font(10, True)
                            elif wd.is_holiday(d):
                                if done:
                                    cell.fill = xl_fill(_GREY_FILL)
                                else:
                                    cell.value = 'Done'
                                    done = True
                                    cell.font = xl_font(10, True)
                            else:
                                cell.fill = xl_fill(_GREY_FILL)
                        elif k > 4:
                            if d == rec['end_date'] + timedelta(days=1):
                                cell.value = 'Done'
                                done = True
                                cell.font = xl_font(10, True)
                            elif wd.is_holiday(d):
                                if done:
                                    cell.fill = xl_fill(_GREY_FILL)
                                else:
                                    cell.value = 'Done'
                                    done = True
                                    cell.font = xl_font(10, True)
                            else:
                                cell.fill = xl_fill(_GREY_FILL)
                        else:
                            cell.fill = xl_fill(_GREY_FILL)
                    elif wd.is_holiday(d):
                        cell.fill = xl_fill(_GREY_FILL)
                    else:
                        if d == report_date:
                            if rec['ccy_id'] != 'USD':
                                cell.value = (
                                    f'=INDEX(closing_price!A:C,MATCH(M{r},'
                                    f'closing_price!A:A,),3)'
                                )
                            else:
                                cell.value = None
                        elif cols[col]:
                            cell.value = cols[col]
                        else:
                            cell.value = None
                else:
                    cell.value = cols[col]

            cell_a = ws[f'A{r}']
            fill_color = _CODE_FILLS.get(rec['code'])
            if fill_color:
                cell_a.fill = xl_fill(fill_color)
            elif rec['code'] in _SMALL_FONT_CODES:
                cell_a.font = xl_font(9)

            row += 1

    return row + 2


def _write_status_grid(ws, count_row, n, date_range):
    """Writes the Z:AJ knock-out/strike tracking grid for one contract block.

    `count_row`/`n` match the same block's `_write_contracts` call (the row
    passed in, and `len(records)`) -- this reproduces `_write_contracts`'
    own row placement: the date-header row sits on `count_row+2` (the same
    row `_write_contracts` uses for its own O:X date header), and the body
    spans `count_row+4 .. count_row+4+n-1` (the same rows `_write_contracts`
    used for its data rows). If n == 0 there is nothing to track -- no
    header, no body.

    Z is a literal per-row direction-flag input (always 2, not a formula).
    AA:AJ are formulas, paired 1:1 with the existing O:X closing-price
    columns via _STATUS_COLS/_OX_COLS, each referencing that row's own
    spot/strike/KO price (E/F/G, already written by _write_contracts) and
    the paired day's closing price -- so the grid recalculates live from
    O:X. Ports the "AA-AJ per-day status columns" this module's docstring
    previously said were dropped.
    """
    if n == 0:
        return

    header_row = count_row + 2
    first_data_row = count_row + 4
    last_data_row = first_data_row + n - 1

    for col, offset in zip(_STATUS_COLS, range(10)):
        cell = ws[f'{col}{header_row}']
        cell.value = date_range[offset]
        cell.font = xl_font(9, True)
        cell.alignment = xl_align(True)
        cell.number_format = 'm/d'
        cell.border = xl_box()

    for r in range(first_data_row, last_data_row + 1):
        z_cell = ws[f'Z{r}']
        z_cell.value = 2
        z_cell.font = xl_font(10)
        z_cell.alignment = xl_align(True)
        z_cell.border = xl_box()

        for col, price_col in zip(_STATUS_COLS, _OX_COLS):
            formula = (
                f'=IF($E{r}>$F{r},'
                f'IF({price_col}{r}=0,"xxx",IF({price_col}{r}="","",'
                f'IF($Z{r}=2,IF($G{r}<={price_col}{r},"KO",IF($F{r}>={price_col}{r},"D",".")),'
                f'IF($G{r}<={price_col}{r},"KO",".")))),'
                f'IF({price_col}{r}=0,"xxx",IF({price_col}{r}="","",'
                f'IF($Z{r}=2,IF($G{r}>={price_col}{r},"KO",IF($F{r}<={price_col}{r},"D",".")),'
                f'IF($G{r}>={price_col}{r},"KO",".")))))'
            )
            cell = ws[f'{col}{r}']
            cell.value = formula
            cell.font = xl_font(10)
            cell.alignment = xl_align(True)
            cell.border = xl_box()


def _compute_circle_cells(records, first_row, date_range, report_date, wd, price_lookup):
    """Returns [(col_letter, row), ...] for every O:X cell that will show a "D"
    status (strike breached without knocking out) once opened in Excel.

    openpyxl never evaluates the AA:AJ status formulas _write_status_grid writes,
    so this recomputes the same test in Python from the same source data, in
    order to know where to draw circles (openpyxl can't draw shapes at all --
    see _inject_circles). Mirrors _write_contracts' own O:X gating (start/end
    window, holiday, report_date) so a cell is only proposed for circling if it
    will actually display a numeric price there:
    - before start_date / after end_date / a holiday: the cell is blank or a
      "Done" placeholder, never a price -- skipped.
    - report_date itself: for non-USD codes _write_contracts writes a live
      formula against the never-populated `closing_price` sheet (see that
      function's `d == report_date` branch); there is no literal price to
      classify at export time, so this date is always skipped, same as the
      spec's own "px is None" skip rule.
    - otherwise: px = price_lookup(code_ref, d); skipped unless it's a real,
      nonzero number (rules out None, 0, and the "Done" *string* placeholder,
      which the bare `px in (0, None, "")` check wouldn't catch since it's
      neither of those three values).
    """
    cells = []
    for i, rec in enumerate(records):
        r = first_row + i
        e, f, g = rec['spot'], rec['strike'], rec['ko']
        above = e > f
        for col, d in zip(_OX_COLS, date_range):
            if d < rec['start_date'] or d > rec['end_date'] or wd.is_holiday(d) or d == report_date:
                continue
            px = price_lookup(rec['code_ref'], d)
            if not isinstance(px, (int, float)) or px == 0:
                continue
            if above:
                status = 'KO' if g <= px else ('D' if f >= px else '.')
            else:
                status = 'KO' if g >= px else ('D' if f <= px else '.')
            if status == 'D':
                cells.append((col, r))
    return cells


# --- Reference data ported verbatim from ltv_stocks2.py's LTV_Stocks.__init__ / position() ---

PRIMARY_BANK_ACCOUNT = 'DBPe'
PRIMARY_SHEET_NAME = f'{PRIMARY_BANK_ACCOUNT}-HKD'

_CCYS = ('HKD', 'SGD')

_BANK_NAME = {
    "CB1": "CITIBANK",
    "CB2": "CITIBANK",
    "CB3": "CITIBANK",
    "CBBH": "CITIBANK",
    "CBBH2": "CITIBANK",
    "CBSG": "CITIBANK",
    "BOS": "Bank of Singapore",
    "DBPe": "DEUTSCHE BANK",
    "DBPL": "DEUTSCHE BANK",
    "SC": "STANDARD CHARTERED",
    "SHK": "SUN HUNG KAI Account No. 1",
    "SHK2": "SUN HUNG KAI Account No. 2",
    "MST1": "MORGAN STANLEY",
    "MST2": "MORGAN STANLEY",
    "MSPL": "MORGAN STANLEY",
    "NSG": "NOMURA SINGAPORE",
}

_SUB_TITLE = {
    'CB2': {'title': 'ACCOUNT # 2 (REALGOLD)', 'color': 'FF7030A0'},
    'CB3': {'title': 'ACCOUNT # 3', 'color': 'FF0070C0'},
    'CBBH': {'title': 'BERRY HILL Account', 'color': '00FF6600'},
    'CBBH2': {'title': 'BERRY HILL Account 2', 'color': '00FF6600'},
    'CBSG': {'title': 'Singapore Account No. 1', 'color': 'FF7030A0'},
    'DBPL': {'title': 'PERFECT LEGEND HOLDINGS w/ Lucio Yan', 'color': '00000000'},
    'MST1': {'title': 'Titan Account No. 1', 'color': '00000000'},
    'MST2': {'title': 'ACCOUNT NO. 2 (Titan) - PERSONAL', 'color': '00000000'},
    'MSPL': {'title': '(Perfect Legend)', 'color': '00000000'},
}

# Per-bank font colors used on the positions-table account label (position() ~716-733).
_POSITION_COLOR = {
    'CB1': 'FF7030A0', 'CB2': 'FF7030A0', 'CB3': 'FF0070C0',
    'CBBH': '00FF6600', 'CBBH2': '00FF6600', 'CBSG': 'FF7030A0',
    'BOS': '00000000', 'DBPe': '00000000', 'DBPL': '00000000',
    'SC': '00000000', 'SHK': '00000000', 'SHK2': '00000000',
    'MST1': '00000000', 'MST2': '00000000', 'MSPL': '00000000', 'NSG': '00000000',
}

# Positions-table account label (position() ~735-758); values are either a
# plain string or a dict keyed by ccy.
_ACCOUNT_LABEL = {
    "CB1": {
        "HKD": "Citibank Account No. 1 Stocks",
        "JPY": "Citibank Account No. 1 Stocks",
        "AUD": "Citibank Account No. 1 Stocks",
        "USD": "Citibank Account No. 1 Stocks",
        "SGD": "Citibank Account No. 1 Stocks",
    },
    "CB2": "Citibank Account No. 2 Stocks",
    "CB3": "Citibank Account No. 3 Stocks",
    "CBBH": "Citibank Berry Hill  Stocks",
    "CBBH2": "Citibank Berry Hill No. 2 Stocks",
    "CBSG": "Citibank Singapore Account No. 1 Stocks",
    "BOS": "Bank of Singapore Stocks",
    "DBPe": "DEUTSCHE PERSONAL Stocks",
    "DBPL": "DEUTSCHE PERFECT LEGEND Stocks",
    "SC": "Standard Chartered Stocks",
    "SHK": "Sun Hung Kai Account No. 1 Stocks",
    "SHK2": "Sun Hung Kai Account No. 2 Stocks",
    "MST1": "MORGAN TITAN No. 1 Stocks",
    "MST2": "MORGAN TITAN No. 2 Stocks",
    "MSPL": "MORGAN PERFECT LEGEND Stocks",
    "NSG": "NOMURA SINGAPORE",
}

# column_width() (298-339). Z-AJ intentionally omitted — those are widths for
# the out-of-scope reconciliation helper columns.
_COLUMN_WIDTHS = {
    'A': 27.57 + 0.71, 'B': 4.86 + 0.71, 'D': 11.71 + 0.71, 'E': 7.57 + 0.71,
    'F': 7.57 + 0.71, 'G': 7.57 + 0.71, 'H': 9.14 + 0.71, 'I': 9.14 + 0.71,
    'J': 5 + 0.71, 'K': 5 + 0.71, 'L': 5 + 0.71, 'N': 9.14 + 0.71,
    'O': 5.71 + 0.71, 'P': 5.71 + 0.71, 'Q': 5.71 + 0.71, 'R': 5.71 + 0.71,
    'S': 5.71 + 0.71, 'T': 5.71 + 0.71, 'U': 5.71 + 0.71, 'V': 5.71 + 0.71,
    'W': 5.71 + 0.71, 'X': 5.71 + 0.71,
}


def _set_column_widths(ws):
    """Port of column_width() (298-339), minus the Z-AJ helper-column widths."""
    for col, width in _COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    for col in ('C', 'M'):
        ws.column_dimensions[col].hidden = True


def report_header(ws, bank_id, ccy, report_date, row, date_range):
    """Port of report_header() (342-407)."""
    cell = ws[f'A{row}']
    cell.value = f'{_BANK_NAME.get(bank_id, bank_id)} as of {report_date.strftime("%B %d, %Y")}'
    cell.font = xl_font(16, True)

    cell = ws[f'H{row}']
    cell.value = 'updated'
    cell.font = xl_font(16, True)
    cell.alignment = xl_align(False, 'right')

    cell = ws[f'I{row}']
    cell.value = date_range[5]
    cell.font = xl_font(16, True)
    cell.number_format = 'd-mmm-yyyy'
    ws.merge_cells(f'I{row}:K{row}')

    if bank_id in _SUB_TITLE:
        row += 1
        cell = ws[f'A{row}']
        cell.value = _SUB_TITLE[bank_id]['title']
        cell.font = xl_font_color(14, _SUB_TITLE[bank_id]['color'])

        if bank_id == 'DBPL':
            for col in ('A', 'B', 'C', 'D', 'E'):
                cell = ws[f'{col}{row}']
                cell.fill = xl_fill('00FFFF00')
                cell.font = xl_font_color(12, _SUB_TITLE[bank_id]['color'])
        elif bank_id == 'MST1':
            ws[f'A{row}'].fill = xl_fill('FFF6DDC2')
        elif bank_id == 'MST2':
            for col in ('A', 'B', 'C', 'D', 'E'):
                ws[f'{col}{row}'].fill = xl_fill('00666699')
        elif bank_id == 'MSPL':
            ws[f'A{row}'].fill = xl_fill('00FFFF00')

    if ccy == 'JPY':
        row += 1
        cell = ws[f'A{row}']
        cell.value = 'JAPAN STOCKS'
        cell.font = xl_font_color(14, 'FF008000')
    elif ccy == 'AUD':
        row += 1
        cell = ws[f'A{row}']
        cell.value = 'AUSTRALIAN STOCKS'
        cell.font = xl_font_color(14, 'FF008000')
    elif ccy in ('USD', 'SGD'):
        row += 1
        cell = ws[f'A{row}']
        cell.value = f'{ccy} STOCKS'
        cell.font = xl_font_color(18, '00000000')

    row += 2
    return row


def _write_positions(ws, positions, report_date, row, wd, price_lookup, bank_id=None, ccy='HKD',
                      hkd_wd=None):
    """Write the positions table starting at `row`. Returns the next free row.

    Port of position() (712-914), excluding the Y-AC reconciliation helper
    columns (out of scope for this port). `bank_id`/`ccy` drive the account-name
    header label and per-bank font color (position() ~716-758); both default so
    the function still degrades gracefully if the caller omits them.

    `hkd_wd` is a `WorkingDay(db, 'HKD')` instance: the legacy start_date
    walk-back and header column F (ltv_stocks2.py:121-123, 765) are always
    computed against HKD, regardless of the sheet's own `ccy`. Falls back to
    `wd` (the sheet's own-ccy WorkingDay) if the caller omits it.
    """
    ws.column_dimensions['C'].hidden = True

    account_label = _ACCOUNT_LABEL.get(bank_id)
    if isinstance(account_label, dict):
        account_label = account_label.get(ccy)
    if account_label is None:
        account_label = bank_id or ''

    color = _POSITION_COLOR.get(bank_id, '00000000')

    if hkd_wd is None:
        hkd_wd = wd
    start_date = position_start_date(report_date, hkd_wd)
    end_date = report_date

    # Head Line 1
    row1 = row
    labels = {
        'A': account_label,
        'B': 'CODE',
        'D': 'AS OF',
        'F': hkd_wd.previous_day(start_date),
        'I': end_date,
        'J': '% Inc. / Dec.',
        'L': report_date if ccy != 'USD' else wd.previous_day(report_date),
    }

    for col, value in labels.items():
        cell = ws[f'{col}{row1}']
        cell.value = value

        if col == 'A':
            if bank_id in ('CB2', 'CB3', 'CBBH', 'CBBH2'):
                cell.font = xl_font_color(9, color)
            elif ccy == 'HKD':
                cell.font = xl_font(8, True) if bank_id == 'DBPL' else xl_font(9, True)
            else:
                cell.font = xl_font_color(9, color)
        else:
            cell.font = xl_font(9, True)

        if col == 'D':
            cell.alignment = xl_align(True, 'right')
        elif col == 'F':
            cell.alignment = xl_align(True, 'left')
        else:
            cell.alignment = xl_align()
        cell.border = xl_box()

        if col == 'F':
            cell.number_format = 'd-mmm-yyyy'
        if col in ('I', 'L'):
            cell.number_format = 'd-mmm'
        if col == 'I':
            cell.fill = xl_fill('00FFFF00')

    for col in ('A', 'B'):
        ws.merge_cells(f'{col}{row1}:{col}{row1 + 1}')
    ws.merge_cells(f'D{row1}:E{row1}')
    ws.merge_cells(f'F{row1}:H{row1}')
    ws.merge_cells(f'J{row1}:K{row1 + 1}')
    ws.merge_cells(f'L{row1}:N{row1 + 1}')

    ws.row_dimensions[row1].height = 16

    # Head Line 2
    row2 = row1 + 1
    labels2 = {
        'D': 'UNBLOCKED',
        'E': 'BLOCKED',
        'G': 'TOTAL SHARES',
        'I': 'Ave. Price',
    }
    for col, value in labels2.items():
        cell = ws[f'{col}{row2}']
        cell.value = value
        cell.font = xl_font(9, True)
        cell.alignment = xl_align(True)
        cell.border = xl_box()

    ws.merge_cells(f'E{row2}:F{row2}')
    ws.merge_cells(f'G{row2}:H{row2}')
    ws.row_dimensions[row2].height = 16

    # Stock Positions
    r = row2 + 1
    for code, pos in positions.items():
        if ccy == 'USD':
            trade_date = wd.previous_day(report_date)
            closing_price = price_lookup(pos['code_ref'], trade_date)
        else:
            closing_price = f'=INDEX(closing_price!A:C,MATCH(C{r},closing_price!A:A,),3)'

        cols = {
            'A': pos['stock_name'],
            'B': code,
            'C': pos['yahoo_ticker'],
            'D': pos['unblocked'],
            'E': pos['blocked'],
            'G': f'=D{r}+E{r}',
            'I': pos['average'],
            'J': f'=(L{r}/I{r})-1',
            'L': closing_price,
            'O': pos.get('transactions'),
        }

        for col, value in cols.items():
            cell = ws[f'{col}{r}']
            if col in ('D', 'E', 'G'):
                if value:
                    cell.value = value
            else:
                cell.value = value

            cell.font = xl_font(10)

            if col == 'A':
                cell.alignment = xl_align(True, 'left')
            elif col == 'O':
                cell.alignment = xl_align(True, 'left', 'top')
            else:
                cell.alignment = xl_align(True)

            if col != 'O':
                cell.border = xl_box()

            if col in ('D', 'E', 'G'):
                cell.number_format = '#,##0'
            elif col == 'I':
                cell.number_format = '#,##0.0000'
            elif col == 'L':
                cell.number_format = '#,##0.00'
            elif col == 'J':
                cell.number_format = '0.00%'

        ws.merge_cells(f'E{r}:F{r}')
        ws.merge_cells(f'G{r}:H{r}')
        ws.merge_cells(f'J{r}:K{r}')
        ws.merge_cells(f'L{r}:N{r}')
        ws.merge_cells(f'O{r}:X{r}')

        ws.row_dimensions[r].height = 50

        r += 1

    r += 2
    return r


def _inject_accu_only_positions(positions, accu, db, bank_ref, bank_id, report_date):
    """Port of `contract()`'s side-effect on `self.stock_position` (ltv_stocks2.py
    ~561-563): any ACCU contract whose code has no balance-derived position (zero
    net shares as of the report's AS-OF snapshot date) gets a placeholder row
    (`unblocked=0, blocked=0`). DECU contracts get no such placeholder (legacy sets
    them to `None`, which `position()` skips entirely).

    BUG FIX (diverges from legacy): the placeholder's `average` used to always be
    the contract's strike price. But `position_records()` gates solely on the
    AS-OF-snapshot-date balance (deliberately -- that's what Unblocked/Blocked/Total
    Shares show); a code that accumulated real shares THIS WEEK (trades landed after
    the snapshot but on/before report_date) has a real, charge-inclusive cost basis
    via `_average()` that was being discarded in favor of the strike. Every other
    position row already prices Ave. Price fresh as of report_date rather than the
    walked-back AS-OF date (see position_records()'s own comment on this) -- apply
    that same principle here: use the real average (and attach the week's
    transactions narrative) whenever one exists; fall back to the strike only when
    the code truly has no computable average as of report_date either (no trades
    ever, or still net zero). Returns a fresh dict re-sorted by code.
    """
    positions = dict(positions)
    for rec in accu:
        code = rec['code']
        if code not in positions:
            real_average = _average(db, bank_id, code, report_date)
            if real_average:
                positions[code] = {
                    'stock_name': rec['stock_name_plain'],
                    'code': code,
                    'code_ref': rec['code_ref'],
                    'yahoo_ticker': rec['yahoo_ticker'],
                    'balance': 0,
                    'blocked': 0,
                    'unblocked': 0,
                    'average': real_average,
                    'transactions': _transactions_narrative(db, bank_ref, rec['code_ref'], report_date),
                }
            else:
                positions[code] = {
                    'stock_name': rec['stock_name_plain'],
                    'code': code,
                    'code_ref': rec['code_ref'],
                    'yahoo_ticker': rec['yahoo_ticker'],
                    'balance': 0,
                    'blocked': 0,
                    'unblocked': 0,
                    'average': rec['strike'],
                    'transactions': None,
                }
    return dict(sorted(positions.items()))


_NS_CT = 'http://schemas.openxmlformats.org/package/2006/content-types'
_NS_PKGREL = 'http://schemas.openxmlformats.org/package/2006/relationships'
_NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_DRAWING_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.drawing+xml'
_DRAWING_REL_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing'
_RELS_REL_TYPE = 'http://schemas.openxmlformats.org/package/2006/relationships'

ET.register_namespace('', _NS_CT)


def _drawing_xml(cells):
    """One <xdr:twoCellAnchor> ellipse (transparent fill, 2pt solid black outline)
    per (col_letter, row) cell, anchored just inside that cell."""
    anchors = []
    for i, (col, row) in enumerate(cells, start=1):
        col_idx = column_index_from_string(col) - 1
        row_idx = row - 1
        anchors.append(
            f'<xdr:twoCellAnchor editAs="oneCell">'
            f'<xdr:from><xdr:col>{col_idx}</xdr:col><xdr:colOff>30000</xdr:colOff>'
            f'<xdr:row>{row_idx}</xdr:row><xdr:rowOff>15000</xdr:rowOff></xdr:from>'
            f'<xdr:to><xdr:col>{col_idx + 1}</xdr:col><xdr:colOff>-30000</xdr:colOff>'
            f'<xdr:row>{row_idx + 1}</xdr:row><xdr:rowOff>-15000</xdr:rowOff></xdr:to>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{i}" name="D-Circle {i}"/><xdr:cNvSpPr/></xdr:nvSpPr>'
            f'<xdr:spPr>'
            f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
            f'<a:prstGeom prst="ellipse"><a:avLst/></a:prstGeom>'
            f'<a:noFill/>'
            f'<a:ln w="25400"><a:solidFill><a:srgbClr val="000000"/></a:solidFill></a:ln>'
            f'</xdr:spPr>'
            f'</xdr:sp>'
            f'<xdr:clientData/>'
            f'</xdr:twoCellAnchor>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr '
        'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        + ''.join(anchors) +
        '</xdr:wsDr>'
    ).encode('utf-8')


def _inject_circles(buf, circle_cells):
    """Post-processes a saved workbook's zip to draw a circle over every O:X
    cell in `circle_cells` (``{sheet_name: [(col_letter, row), ...]}``).

    openpyxl can only create images/charts, never freeform drawing shapes, so
    the ellipses are injected as raw DrawingML: one `xl/drawings/drawingN.xml`
    part per affected sheet, wired to that sheet via a `<drawing r:id="..."/>`
    element and a worksheet-level `_rels` relationship, with the corresponding
    `[Content_Types].xml` override added. Every other part in the original zip
    is copied through unchanged -- existing values, formulas, formatting, and
    merged cells (and any drawing a sheet already had) are untouched.
    """
    circle_cells = {name: cells for name, cells in circle_cells.items() if cells}
    if not circle_cells:
        return buf

    buf.seek(0)
    zin = zipfile.ZipFile(buf, 'r')
    originals = {info.filename: zin.read(info.filename) for info in zin.infolist()}

    # sheet name -> worksheet part path, via workbook.xml + workbook.xml.rels
    # (never assume sheetN.xml ordering matches sheet creation order).
    wb_root = ET.fromstring(originals['xl/workbook.xml'])
    ns_main = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    rels_root = ET.fromstring(originals['xl/_rels/workbook.xml.rels'])
    rid_to_target = {
        rel.get('Id'): rel.get('Target')
        for rel in rels_root
    }
    sheet_name_to_part = {}
    for sheet_el in wb_root.find('m:sheets', ns_main):
        rid = sheet_el.get(f'{{{_NS_R}}}id')
        target = rid_to_target[rid]
        if target.startswith('/'):
            part = target[1:]  # package-root-absolute, e.g. "/xl/worksheets/sheet3.xml"
        elif target.startswith('xl/'):
            part = target
        else:
            part = f'xl/{target}'  # relative to xl/_rels/, e.g. "worksheets/sheet3.xml"
        sheet_name_to_part[sheet_el.get('name')] = part

    existing_drawing_nums = [
        int(m.group(1)) for name in originals
        for m in [re.match(r'xl/drawings/drawing(\d+)\.xml$', name)]
        if m
    ]
    next_drawing_num = max(existing_drawing_nums, default=0) + 1

    new_or_changed = {}
    ct_xml = originals['[Content_Types].xml'].decode('utf-8')
    ct_overrides = []

    for sheet_name, cells in circle_cells.items():
        ws_part = sheet_name_to_part[sheet_name]
        drawing_num = next_drawing_num
        next_drawing_num += 1
        drawing_part = f'xl/drawings/drawing{drawing_num}.xml'
        drawing_rels_part = f'xl/drawings/_rels/drawing{drawing_num}.xml.rels'

        new_or_changed[drawing_part] = _drawing_xml(cells)
        new_or_changed[drawing_rels_part] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{_RELS_REL_TYPE}"></Relationships>'
        ).encode('utf-8')
        ct_overrides.append(
            f'<Override PartName="/{drawing_part}" ContentType="{_DRAWING_CONTENT_TYPE}"/>'
        )

        ws_dir, ws_file = ws_part.rsplit('/', 1)
        ws_rels_part = f'{ws_dir}/_rels/{ws_file}.rels'
        ws_rels_xml = originals.get(ws_rels_part) or new_or_changed.get(ws_rels_part)
        if ws_rels_xml:
            rels_root = ET.fromstring(ws_rels_xml)
            existing_rids = [rel.get('Id') for rel in rels_root]
            next_num = 1
            while f'rId{next_num}' in existing_rids:
                next_num += 1
            new_rid = f'rId{next_num}'
            rel_xml = (
                f'<Relationship Id="{new_rid}" Type="{_DRAWING_REL_TYPE}" '
                f'Target="../drawings/drawing{drawing_num}.xml"/>'
            )
            ws_rels_xml = ws_rels_xml.decode('utf-8').replace(
                '</Relationships>', rel_xml + '</Relationships>'
            ).encode('utf-8')
        else:
            new_rid = 'rId1'
            ws_rels_xml = (
                f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                f'<Relationships xmlns="{_RELS_REL_TYPE}">'
                f'<Relationship Id="{new_rid}" Type="{_DRAWING_REL_TYPE}" '
                f'Target="../drawings/drawing{drawing_num}.xml"/>'
                f'</Relationships>'
            ).encode('utf-8')
        new_or_changed[ws_rels_part] = ws_rels_xml

        ws_xml = (originals.get(ws_part) or new_or_changed.get(ws_part)).decode('utf-8')
        # The worksheet root openpyxl writes declares no "r" namespace prefix
        # (only xl/workbook.xml does) -- declare it locally on this element
        # rather than assuming an ancestor already bound it.
        drawing_el = f'<drawing xmlns:r="{_NS_R}" r:id="{new_rid}"/>'
        if '</worksheet>' in ws_xml:
            ws_xml = ws_xml.replace('</worksheet>', drawing_el + '</worksheet>')
        new_or_changed[ws_part] = ws_xml.encode('utf-8')

    ct_xml = ct_xml.replace('</Types>', ''.join(ct_overrides) + '</Types>')
    new_or_changed['[Content_Types].xml'] = ct_xml.encode('utf-8')

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for info in zin.infolist():
            content = new_or_changed.get(info.filename, originals[info.filename])
            zout.writestr(info, content)
            written.add(info.filename)
        for name, content in new_or_changed.items():
            if name not in written:
                zout.writestr(name, content)

    zin.close()
    out.seek(0)
    return out


def build_workbook(db, report_date, bank_ids):
    """Port of create() (211-295). Returns the workbook as an `io.BytesIO`.

    For each ccy in ('HKD', 'SGD') x bank in `bank_ids`: resolves the bank_ref,
    gathers ACCU/DECU contracts (filtered to `ccy`) and positions; skips the
    sheet unless it has an ACCU, a DECU, or a position. Writes report_header,
    ACCU then DECU contract tables, then the positions table. After all sheets,
    assembles the cross-sheet ACCU/DECU total formulas into the primary
    `DBPe-HKD` sheet (skipped if that sheet wasn't built — see Escalate note).
    """
    wb = openpyxl.Workbook()
    del wb['Sheet']
    wb.create_sheet('closing_price')
    wb.create_sheet('record')

    date_range = week_dates(report_date)
    hkd_count = {'total_row': {}}
    hkd_wd = WorkingDay(db, 'HKD')
    circle_cells = {}

    for ccy in _CCYS:
        for bank_id in bank_ids:
            bank_row = db.execute(
                "SELECT ref_num FROM tbl_bank_account WHERE bank_id = ?", (bank_id,)
            ).fetchone()
            if bank_row is None:
                continue
            bank_ref = bank_row[0]

            accu = [r for r in contract_records(db, bank_ref, 'ACCU') if r['ccy_id'] == ccy]
            decu = [r for r in contract_records(db, bank_ref, 'DECU') if r['ccy_id'] == ccy]
            positions = position_records(db, bank_ref, bank_id, ccy, report_date, hkd_wd=hkd_wd)
            positions = _inject_accu_only_positions(positions, accu, db, bank_ref, bank_id, report_date)

            if not (accu or decu or positions):
                continue

            sheet_name = f'{bank_id}-{ccy}'
            ws = wb.create_sheet(sheet_name)

            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(
                ws, paper_size=4, orientation='landscape'
            )
            ws.page_margins.left = 0.1
            ws.page_margins.right = 0.1
            ws.page_margins.top = 0.15
            ws.page_margins.bottom = 0.1
            ws.page_margins.header = 0.0
            ws.page_margins.footer = 0.0

            _set_column_widths(ws)

            wd = WorkingDay(db, ccy)
            price_lookup = lambda code_ref, d: get_stock_price(db, code_ref, d)

            r = 1
            r = report_header(ws, bank_id, ccy, report_date, r, date_range)

            accu_count_row = r
            r = _write_contracts(ws, accu, 'ACCU', r, report_date, date_range, wd,
                                  price_lookup=price_lookup, bank_id=bank_id)
            _write_status_grid(ws, accu_count_row, len(accu), date_range)

            decu_count_row = r
            r = _write_contracts(ws, decu, 'DECU', r, report_date, date_range, wd,
                                  price_lookup=price_lookup, bank_id=bank_id)
            _write_status_grid(ws, decu_count_row, len(decu), date_range)

            sheet_circles = (
                _compute_circle_cells(accu, accu_count_row + 4, date_range, report_date, wd, price_lookup)
                + _compute_circle_cells(decu, decu_count_row + 4, date_range, report_date, wd, price_lookup)
            )
            if sheet_circles:
                circle_cells[sheet_name] = sheet_circles

            r = _write_positions(ws, positions, report_date, r, wd, price_lookup,
                                  bank_id=bank_id, ccy=ccy, hkd_wd=hkd_wd)

            ws.print_area = f'A1:X{r}'

            if ccy == 'HKD':
                hkd_count.setdefault(bank_id, {})
                hkd_count[bank_id]['ACCU'] = f'A{accu_count_row}'
                hkd_count[bank_id]['DECU'] = f'A{decu_count_row}'
                if bank_id == PRIMARY_BANK_ACCOUNT:
                    hkd_count['total_row']['ACCU'] = accu_count_row
                    hkd_count['total_row']['DECU'] = decu_count_row

    # Cross-sheet ACCU/DECU totals (create() 251-289).
    accu_formula = ''
    decu_formula = ''
    first = True
    for bank_account in hkd_count:
        if bank_account == 'total_row':
            accu_formula = '=SUM('
            decu_formula = '=SUM('
        else:
            accu_cell = hkd_count[bank_account]['ACCU']
            decu_cell = hkd_count[bank_account]['DECU']
            sep = '' if first else ', '
            accu_formula += f"{sep}'{bank_account}-HKD'!{accu_cell}"
            decu_formula += f"{sep}'{bank_account}-HKD'!{decu_cell}"
            first = False
    accu_formula += ')'
    decu_formula += ')'

    if PRIMARY_SHEET_NAME in wb.sheetnames and 'ACCU' in hkd_count.get('total_row', {}):
        ws_primary = wb[PRIMARY_SHEET_NAME]
        accu_row = hkd_count['total_row']['ACCU']
        decu_row = hkd_count['total_row']['DECU']

        cell_accu_total = ws_primary[f'B{accu_row}']
        cell_accu_total.value = accu_formula

        cell_decu_total = ws_primary[f'B{decu_row}']
        cell_decu_total.value = decu_formula

        for cell in (cell_accu_total, cell_decu_total):
            cell.font = xl_font(9, True)
            cell.alignment = xl_align()
    # else: no DBPe-HKD sheet was built (e.g. tests seeding only other banks) —
    # cross-sheet totals cannot be recorded; see Escalate note in the task report.

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf = _inject_circles(buf, circle_cells)
    return buf
