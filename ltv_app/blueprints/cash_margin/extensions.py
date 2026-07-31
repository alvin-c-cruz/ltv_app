import os
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill

from .. term_sheet import StockContract
from ...tz import ph_today


# Per-bank fill color used to mark the "live" (E or G) shares-per-day column --
# ported verbatim from localhost/modules/cash_margin.py::update_file's
# ACCOUNT_COLOR dict.
ACCOUNT_COLOR = {
    "CB1": "0099CC00", "CB2": "0099CC00", "CB3": "0099CC00",
    "CBBH": "0099CC00", "CBBH2": "0099CC00", "CBSG": "0099CC00",
    "BOS": "00FFFF00",
    "DBPe": "00CCFFFF", "DBPL": "00CCFFFF",
    "SC": "00FFCC00",
    "SHK": "00FF00FF", "SHK2": "00FF00FF",
    "MST1": "00FF8080", "MST2": "00FF8080", "MSPL": "00FF8080", "NSG": "00FF8080",
}

# Bank groups whose "received" / "total" progress is tracked in fixing PERIODS
# (ts_dict["received"] / ts_dict["total"]) rather than DAYS
# (ts_dict["days_received"] / ts_dict["days_max"]) -- matches legacy
# update_file()'s `if bank_account in ("CB1", "CB2", "CB3", "CBBH", "CBBH2")`
# branch exactly, including CBSG's deliberate *exclusion* from this list even
# though it shares CB1's fill color in ACCOUNT_COLOR above.
_PERIOD_TRACKED_BANKS = ("CB1", "CB2", "CB3", "CBBH", "CBBH2")

_ALL_COLS = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T")


def _fill(ws, row_num, hex_color, cols=_ALL_COLS):
    for col in cols:
        ws[f"{col}{row_num}"].fill = PatternFill(patternType='solid', fgColor=hex_color)


# How many working days sd_3_days() will look back for a closing price before
# giving up on a code with no (or very stale) tbl_stock_price history. Without
# this cap, a code with zero price rows (e.g. a stock traded for the first
# time today) sends previous_day() walking backwards forever -- confirmed hang
# on newly-added ticker 3308 (2026-07-31), zero rows in tbl_stock_price at the
# time. Ported from localhost/modules/cash_margin.py's already-fixed constant
# of the same name.
_MAX_PRICE_LOOKBACK_DAYS = 60


def _short_date(date_str):
    """Match localhost/modules/dates.py::short_date's exact output format --
    confirmed via a throwaway script (deleted) against sample dates:
    short_date('2026-08-15') -> '15-Aug-2026', short_date('2026-08-05') ->
    '5-Aug-2026' (day is NOT zero-padded), short_date('2026-01-31') ->
    '31-Jan-2026'. Reproduced by hand rather than via strftime, since
    strftime's %d would wrongly zero-pad single-digit days.
    """
    months = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
    }
    year = date_str[:4]
    month = int(date_str[5:7])
    day = int(date_str[-2:])
    return f"{day}-{months[month]}-{year}"


def get_next_month(observation_month):
    year = int(observation_month[:4])
    month = int(observation_month[-2:])
    if month != 12:
        month += 1
    else:
        year += 1
        month = 1
    return f"{year}-{str(month).zfill(2)}"


def _stock_name_with_gtd(ts):
    """Matches legacy localhost/modules/term_sheet.py::summary_ts's stock_name
    construction exactly (gtd == 'Yes' -> 'GTD 1m', gtd == 'No' -> 'NO GTD',
    else -> 'GTD {gtd.upper()}' -- e.g. real DB values '1m'/'2m'/'3m'/'4m'
    become 'GTD 1M'/'GTD 2M'/etc.)."""
    if ts.gtd == "Yes":
        return f"{ts.stock_name} GTD 1m"
    elif ts.gtd == "No":
        return f"{ts.stock_name} NO GTD"
    else:
        return f"{ts.stock_name} GTD {ts.gtd.upper()}"


def _contract_to_dict(ts):
    single = ts.daily_shares
    if ts.leveraged == "Yes":
        double_str = "{0:,.0f}".format(single * 2)
        single_str = "{0:,.0f}".format(single)
        shares = f"{single_str} / {double_str}"
    else:
        shares = "{0:,.0f}".format(single)

    last_period = len(ts.schedules)
    # ts.end_date is only ever set by __post_init__'s loop over periods, so it
    # doesn't exist at all when a contract has zero schedule periods -- fall
    # back to start_date (always a populated dataclass field) rather than an
    # attribute that may not exist.
    end_date = ts.schedules[-1].end_date if ts.schedules else ts.start_date

    if ts.frequency == "monthly":
        total = last_period
    elif ts.frequency == "weekly":
        total = last_period / 4
    else:
        total = last_period / 2

    received = 0
    for period in ts.schedules:
        if period.received in (None, ""):
            break
        if ts.frequency == "monthly":
            received += 1
        elif ts.frequency == "weekly":
            received += 0.25
        else:
            received += 0.5

    return {
        "contract_ref": ts.ref_num,
        "reference": ts.reference,
        "code": ts.code,
        "stock_name": _stock_name_with_gtd(ts),
        "shares": shares,
        "spot": "{0:,.4f}".format(ts.spot),
        "strike": "{0:,.4f}".format(ts.strike_value),
        "ko": "{0:,.4f}".format(ts.ko_value),
        "strike_value": ts.strike_value,
        "ko_value": ts.ko_value,
        "start_date": _short_date(ts.start_date),
        "end_date": _short_date(end_date),
        "total": total,
        "received": received,
        "remaining": total - received,
        "days_received": ts.received_days,
        "days_max": ts.total_days,
        "this_month": 0,
        "next_month": 0,
        "_ts": ts,
    }


def sd_3_days(db, code, strike_value, product):
    """Port of legacy localhost/modules/cash_margin.py::sd_3_days (already
    fixed there this session with the bounded lookback via
    _MAX_PRICE_LOOKBACK_DAYS above). Returns 1 (single) or 2 (double) signal
    based on whether strike_value has been breached on each of the 3 most
    recent working days going back from today.

    Takes strike_value as a raw float (Task B2's ts_dict["strike_value"])
    instead of re-parsing a comma-formatted display string with float() --
    that exact bug crashed production earlier this session in a different
    file, generate_fixings.py.

    Uses the same in-memory holiday-set + previous_day() closure pattern as
    ltv_app/blueprints/fixings/extensions/generate_fixings.py, rather than
    legacy's working_day() class, which re-queries tbl_holiday per call.
    """
    ccy = db.execute(
        "SELECT tbl_currency.ccy_id FROM tbl_code "
        "INNER JOIN tbl_currency ON tbl_code.ccy_ref = tbl_currency.ref_num "
        "WHERE tbl_code.code=?", (code,)
    ).fetchone()['ccy_id']

    holidays = {
        (str(r['holi_date'])[:10], r['ccy_id'])
        for r in db.execute(
            "SELECT holi_date, tbl_currency.ccy_id FROM tbl_holiday "
            "INNER JOIN tbl_currency ON tbl_currency.ref_num = tbl_holiday.ccy_ref"
        ).fetchall()
    }

    def is_holiday(date_str):
        return (date_str[:10], ccy) in holidays

    def previous_day(date_str):
        d = datetime.strptime(date_str[:10], '%Y-%m-%d') - timedelta(days=1)
        while is_holiday(str(d)[:10]) or d.isoweekday() in (6, 7):
            d -= timedelta(days=1)
        return str(d)[:10]

    def get_price(date_str):
        row = db.execute(
            "SELECT closing_price FROM tbl_stock_price "
            "INNER JOIN tbl_code ON tbl_code.ref_num = tbl_stock_price.code_ref "
            "WHERE tbl_code.code=? AND tbl_stock_price.trade_date=?",
            (code, date_str)
        ).fetchone()
        return row['closing_price'] if row else None

    def sdk(closing):
        if closing is None:
            # No price found within the lookback window -- can't tell whether
            # strike has been breached, so don't count this day either way.
            return 1
        if product == "ACCU":
            return 2 if strike_value >= closing else 1
        else:
            return 2 if strike_value <= closing else 1

    def find_closing(start_date):
        trade_date = start_date
        closing = get_price(trade_date)
        lookback = 0
        # Falsy check (not closing), not `is None` -- matches legacy's
        # `while not closing` exactly: a literal 0 closing price is treated
        # as "not found" and the walk continues further back, same as legacy.
        while not closing and lookback < _MAX_PRICE_LOOKBACK_DAYS:
            trade_date = previous_day(trade_date)
            closing = get_price(trade_date)
            lookback += 1
        return trade_date, closing

    trade_date_1, closing_1 = find_closing(str(ph_today()))
    trade_date_2, closing_2 = find_closing(previous_day(trade_date_1))
    trade_date_3, closing_3 = find_closing(previous_day(trade_date_2))

    sd_total = sdk(closing_1) + sdk(closing_2) + sdk(closing_3)
    return 2 if sd_total >= 5 else 1


def gather_margin_data(db, ccy, observation_month):
    """Port of legacy cash_margin() -> get_term_sheets() -> group_accounts()
    -> get_two_fixings() (localhost/modules/cash_margin.py), rebuilt on top of
    StockContract instead of the legacy summary_ts/term_sheet classes.

    Drops the legacy bank_group layer ("All") -- see Task 6 brief / plan doc:
    that layer was already a dead no-op in the legacy code (every lookup
    always indexed dict_margin["All"][bank_ref]), so this is a simplification,
    not a fidelity break. Returns {bank_id: {"ACCU": {reference: ts_dict},
    "DECU": {...}}}.
    """
    next_observation_month = get_next_month(observation_month)

    sql = """
        SELECT tbl_stock_contract.ref_num
        FROM tbl_stock_contract
        INNER JOIN tbl_bank_account ON tbl_bank_account.ref_num = tbl_stock_contract.bank_ref
        INNER JOIN tbl_code ON tbl_code.ref_num = tbl_stock_contract.code_ref
        INNER JOIN tbl_currency ON tbl_currency.ref_num = tbl_code.ccy_ref
        WHERE tbl_stock_contract.status = "active"
            AND tbl_currency.ccy_id = ?
        ORDER BY tbl_bank_account.priority
    """
    contract_refs = [row['ref_num'] for row in db.execute(sql, (ccy,)).fetchall()]

    dict_margin = {}
    for contract_ref in contract_refs:
        ts = StockContract(db=db)
        ts.get(ref_num=contract_ref)
        ts.get_schedules()  # get_schedules() already calls __post_init__() internally

        bank_id = ts.bank_id
        product = ts.transaction_type  # "ACCU" or "DECU"

        if bank_id not in dict_margin:
            dict_margin[bank_id] = {"ACCU": {}, "DECU": {}}

        ts_dict = _contract_to_dict(ts)

        this_month = 0
        next_month = 0
        for period in ts.schedules:
            if period.received in (None, ""):
                fixing_date = period.end_date
                if fixing_date[:7] == observation_month:
                    this_month += period.days
                if fixing_date[:7] == next_observation_month:
                    next_month += period.days
        ts_dict["this_month"] = this_month
        ts_dict["next_month"] = next_month

        dict_margin[bank_id][product][ts_dict["reference"]] = ts_dict

    return dict_margin


def build_cash_margin_file(db, ccy, observation_month, instance_path):
    """Port of legacy localhost/modules/cash_margin.py::update_file's
    template-walking loop (lines ~107-225), minus the NAS-share load/save
    branches (this is a download-on-demand report, not legacy's
    scheduled-file-refresh flow -- always load the packaged template and
    save to the instance temp/ dir, matching every other ltv_app Excel
    export).

    Three real discrepancies were found and fixed versus the plan's draft
    pseudocode, by reading legacy's actual update_file() source rather than
    trusting the draft:

    1. known_bank_ids must be the FULL set of bank_id values from
       tbl_bank_account (legacy's `accounts = list_bank_accounts()`), not
       just dict_margin.keys(). dict_margin only contains bank_ids that have
       at least one active contract in this ccy -- a bank with zero such
       contracts would have its D-column header row misclassified as an
       unmatched "reference" and wrongly red-filled.

    2. Column B must be the PLAIN stock name (contract.stock_name), not the
       GTD-suffixed ts_dict["stock_name"]. Legacy's update_file() explicitly
       recomputes `stock_name = getstock_name(code)` -- a fresh, un-suffixed
       lookup -- rather than reusing the GTD-suffixed name summary_ts()
       already built into the ts dict for other purposes. Confirmed by
       cross-referencing /hkd-margin's StockContract.as_dict(), which also
       exposes the plain `self.stock_name`, never a GTD-suffixed variant.

    3. Columns E/G (single/double shares-per-day) must be the RAW
       contract.daily_shares int (and daily_shares*2 when leveraged),
       matching legacy's `ts['single'] = dict_ts.header['daily_shares']` and
       `double = ts.get('double') if ts.get('double') else ts['single']`
       (double is only ever set when leveraged == "Yes"; otherwise it falls
       back to the same raw single value) -- not the comma-formatted display
       string embedded in ts_dict["shares"] (e.g. "12,000"). Legacy writes
       pre-formatted STRINGS for H/I/J (spot/strike/ko) but a raw NUMBER for
       E/G; the two columns are not interchangeable in kind.

    4. K/M (received / total) must branch on bank_id: legacy uses
       ts['received']/ts['total'] (fixing-PERIOD counts) only for
       ("CB1", "CB2", "CB3", "CBBH", "CBBH2"); every other bank -- including
       CBSG despite sharing CB1's fill color -- uses
       ts['days_received']/ts['days_max'] (day counts) instead. The draft
       used the period-count pair unconditionally.
    """
    dict_margin = gather_margin_data(db, ccy, observation_month)

    known_bank_ids = {
        row['bank_id'] for row in db.execute("SELECT bank_id FROM tbl_bank_account").fetchall()
    }

    template_path = os.path.join(
        os.path.dirname(__file__), '..', '..', 'excel_templates', 'cash_margin.xlsx'
    )
    wb = load_workbook(template_path)

    for product in ("ACCU", "DECU"):
        ws = wb[product]

        counter = 0
        row_num = 5
        bank_id = ""
        while counter < 100:
            cell_value = ws[f'D{row_num}'].value
            if cell_value is None:
                counter += 1
            else:
                counter = 0
                if cell_value in known_bank_ids:
                    bank_id = cell_value
                else:
                    reference = cell_value
                    bank_data = dict_margin.get(bank_id, {}).get(product, {})

                    if reference in bank_data:
                        _fill(ws, row_num, "00FFFFFF")

                        ts = bank_data[reference]
                        contract = ts["_ts"]

                        stock_name = contract.stock_name

                        single = contract.daily_shares
                        double = (
                            contract.daily_shares * 2
                            if contract.leveraged == "Yes"
                            else contract.daily_shares
                        )

                        if bank_id in _PERIOD_TRACKED_BANKS:
                            rvd = ts["received"]
                            total_mos = ts["total"]
                        else:
                            rvd = ts["days_received"]
                            total_mos = ts["days_max"]

                        sd = sd_3_days(db, ts["code"], ts["strike_value"], product)

                        cols = {
                            "B": stock_name,
                            "C": ts["code"],
                            "E": single,
                            "F": "/",
                            "G": double,
                            "H": ts["spot"],
                            "I": ts["strike"],
                            "J": ts["ko"],
                            "K": rvd,
                            "L": f"=M{row_num}-K{row_num}",
                            "M": total_mos,
                            "N": ts["start_date"],
                            "O": ts["end_date"],
                            "V": ts["this_month"],
                            "W": ts["this_month"] + ts["next_month"],
                            "X": sd,
                        }
                        for col, value in cols.items():
                            ws[f"{col}{row_num}"].value = value

                        if sd == 2:
                            ws[f"G{row_num}"].fill = PatternFill(patternType='solid', fgColor=ACCOUNT_COLOR[bank_id])
                        else:
                            ws[f"E{row_num}"].fill = PatternFill(patternType='solid', fgColor=ACCOUNT_COLOR[bank_id])

                        if single == double:
                            ws[f"E{row_num}"].fill = PatternFill(patternType='solid', fgColor=ACCOUNT_COLOR[bank_id])

                        if rvd == total_mos:
                            _fill(ws, row_num, "00FFFF00")
                    else:
                        _fill(ws, row_num, "00FF0000")

            row_num += 1

    temp_path = os.path.join(instance_path, 'temp', f'cash_margin_{ccy}_{observation_month}.xlsx')
    wb.save(temp_path)
    wb.close()
    return temp_path
