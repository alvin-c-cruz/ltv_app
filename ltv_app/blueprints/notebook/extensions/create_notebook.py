from flask import current_app
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from datetime import datetime

from ltv_app.blueprints.transactions import TradesDoneAverage

CARD_ROW_HEIGHT = 12
ROW_HEIGHT = 17.25
ROW_NUM_START = 24
STOCK_TRANSACTIONS = [
    "Buy (Accu)",
    "Buy (Accu-KO)",
    "Buy (Spot)",
    "Stock Dividends",
    "Sell (Decu)",
    "Sell (Decu-KO)",
    "Sell (Spot)"
]

DERIVATIVES = [
    "ACCU",
    "DECU"
]

BORROW_RETURN = ["Return Shares", "Borrow Shares"]

RED_COLOR = "00FF0000"


class CreateNotebook:
    def __init__(self, db, trade_date, transactions, transfers):
        self.db = db
        self.trade_date = trade_date
        self.transactions = transactions
        self.transfers = transfers

        self.template_file = os.path.join(current_app.instance_path, "excel_templates", "notebook.xlsx")
        self.filename = os.path.join(current_app.instance_path, "temp", f"{trade_date} notebook.xlsx")

        self.create_file()

    def create_file(self):
        wb = load_workbook(self.template_file)
        ws = wb["notebook"]

        self.write_date(ws)
        self.write_stock_card(ws)
        row_num = self.write_transactions(ws=ws, row_num=ROW_NUM_START)
        row_num = self.write_transfers(ws, row_num)

        ws.print_area = f"A1:Q{row_num}"

        wb.save(self.filename)
        wb.close()

    def write_date(self, ws):
        # Page date
        cell = ws["A1"]
        cell.value = datetime.strptime(self.trade_date, "%Y-%m-%d").strftime("%B %d, %Y %A")
        cell.font = Font(size=13, bold=True)
        cell.alignment = Alignment(horizontal="center")

        ws.merge_cells("A1:Q1")
        ws.row_dimensions[1].height = ROW_HEIGHT

        # Card date
        ws["K6"].value = datetime.strptime(self.trade_date, "%Y-%m-%d")

    def write_stock_card(self, ws):
        for row_num in range(5, ROW_NUM_START):
            ws.row_dimensions[row_num].height = CARD_ROW_HEIGHT

    def write_transactions(self, ws, row_num):
        for ccy, accounts in self.transactions.items():
            for bank_name, codes in accounts.items():
                bank_id = self.db.execute("SELECT bank_id FROM tbl_bank_account WHERE bank_name = ?",
                                          (bank_name, )).fetchone()[0]
                counter = 1

                border_line(ws, row_num)
                cell = ws[f"A{row_num}"]
                cell.value = bank_name if ccy == "HKD" else f"{bank_name} ({ccy})"
                cell.font = Font(size=13, bold=True)
                cell.alignment = Alignment(horizontal="center")
                ws.merge_cells(f"A{row_num}:Q{row_num}")
                ws.row_dimensions[row_num].height = ROW_HEIGHT

                row_num += 1
                ws.row_dimensions[row_num].height = ROW_HEIGHT
                border_line(ws, row_num)

                for code in codes:
                    trades = codes[code]

                for code, trades in codes.items():
                    stock_trade_count = sum([len(trades) for _type, trades in trades.items() if _type in STOCK_TRANSACTIONS])

                    # Running long/short balances for this (bank, stock), started
                    # from the day-opening balance and carried forward across every
                    # transaction below so multiple same-stock trades follow through.
                    running_long = self._opening_balance(bank_name, code)
                    running_short = self._opening_short_balance(bank_name, code)

                    transaction_reference = []
                    for transaction_type in STOCK_TRANSACTIONS:
                        if transaction_type in trades:
                            for row in trades[transaction_type]:
                                row_num, counter, cell_reference = self.write_line(ws, row_num, counter, row)
                                transaction_reference.append(cell_reference)
                                border_line(ws, row_num)

                                row_num += 1
                                ws.row_dimensions[row_num].height = ROW_HEIGHT
                                border_line(ws, row_num)

                    # Balance lines
                    if stock_trade_count:
                        shares_remaining = 0
                        row_num -= 1
                        # Shares beginning
                        balance_formula = f"=G{row_num}"
                        cell = ws[f"G{row_num}"]
                        shares_balance = self.db.execute(
                            "SELECT SUM(tbl_transaction.quantity) as balance "
                            "FROM tbl_transaction "
                            "INNER JOIN tbl_bank_account ON tbl_bank_account.ref_num = tbl_transaction.bank_ref "
                            "INNER JOIN tbl_code ON tbl_code.ref_num = tbl_transaction.code_ref "
                            "WHERE tbl_bank_account.bank_name = ? "
                            "   AND tbl_code.code = ? "
                            "   AND tbl_transaction.trade_date < ?",
                            (bank_name, code, self.trade_date)
                        ).fetchone()
                        cell.value = shares_balance[0] if shares_balance else 0
                        if shares_balance:
                            try:
                                shares_remaining += shares_balance[0]
                            except TypeError:
                                pass

                        cell.font = Font(size=13)
                        cell.number_format = "#,##0_ ;[Red](#,##0)"
                        ws.merge_cells(f"G{row_num}:J{row_num}")
                        row_num += 1
                        ws.row_dimensions[row_num].height = ROW_HEIGHT
                        border_line(ws, row_num)

                        # Quantities
                        i = 0
                        for _type in STOCK_TRANSACTIONS:
                            if _type in trades:
                                for trade in trades[_type]:
                                    shares_remaining += trade["quantity"]
                                    quantity = abs(trade["quantity"])
                                    sign = "+" if trade["quantity"] >= 0 else "-"

                                    cell = ws[f"F{row_num}"]
                                    cell.value = sign
                                    cell.alignment = Alignment(horizontal="right")
                                    cell.font = Font(size=13, name="AR JULIAN")

                                    cell = ws[f"G{row_num}"]
                                    cell.value = f"={transaction_reference[i]}"
                                    i += 1
                                    cell.font = Font(size=13)
                                    cell.number_format = "#,##0_ ;[Red](#,##0)"
                                    ws.merge_cells(f"G{row_num}:J{row_num}")

                                    balance_formula += f"{sign}G{row_num}"

                                    row_num += 1
                                    ws.row_dimensions[row_num].height = ROW_HEIGHT
                                    border_line(ws, row_num)

                        # Remaining shares
                        cell = ws[f"F{row_num}"]
                        cell.value = f'=IF(G{row_num}>=0,"shares","short")'
                        cell.font = Font(size=13)
                        cell.alignment = Alignment(horizontal="right")

                        cell = ws[f"G{row_num}"]
                        cell.value = balance_formula
                        cell.font = Font(size=13)
                        cell.border = Border(top=Side(style='thin'))
                        cell.number_format = "#,##0_ ;[Red](#,##0)"
                        ws.merge_cells(f"G{row_num}:J{row_num}")
                        row_num += 1
                        ws.row_dimensions[row_num].height = ROW_HEIGHT
                        border_line(ws, row_num)

                        # Closing Rate
                        for transaction_type, _ in trades.items():
                            if "KO" in transaction_type:
                                cell = ws[f"O{row_num-3}"]
                                cell.value = "Closing"
                                cell.alignment = Alignment(horizontal="center")
                                cell.font = Font(size=9, color=RED_COLOR)
                                ws.merge_cells(f"O{row_num-3}:Q{row_num-3}")

                                cell = ws[f"O{row_num-2}"]
                                cell.alignment = Alignment(horizontal="center")
                                cell.number_format = "#,##0.00"
                                cell.font = Font(size=9, color=RED_COLOR)
                                ws.merge_cells(f"O{row_num-2}:Q{row_num-2}")

                                cell = ws[f"O{row_num-1}"]
                                cell.alignment = Alignment(horizontal="center")
                                cell.number_format = "d-mmm-yyyy"
                                cell.font = Font(size=9, color=RED_COLOR)
                                ws.merge_cells(f"O{row_num-1}:Q{row_num-1}")

                        # Average
                        if "+" in balance_formula:
                            if shares_balance[0] == 0 and shares_remaining == 0:
                                pass
                            else:
                                cell = ws[f"L{row_num - 2}"]
                                cell.value = "Average"
                                cell.alignment = Alignment(horizontal="center")
                                cell.font = Font(size=13)
                                ws.merge_cells(f"L{row_num - 2}:N{row_num - 2}")

                                cell = ws[f"L{row_num - 1}"]
                                cell.alignment = Alignment(horizontal="center")
                                average = TradesDoneAverage(db=self.db, trade_date=self.trade_date,
                                                            code=code, bank_id=bank_id).average
                                cell.value = average
                                cell.font = Font(size=13)
                                cell.number_format = "#,##0.0000"
                                ws.merge_cells(f"L{row_num - 1}:N{row_num - 1}")

                        # Extra line after remaining shares
                        row_num += 1
                        ws.row_dimensions[row_num].height = ROW_HEIGHT
                        border_line(ws, row_num)

                        # Carry the post-stock-trade long balance into the running
                        # total so a following Borrow/Return begins from it.
                        running_long = shares_remaining

                    for transaction_type in BORROW_RETURN:
                        if transaction_type in trades:
                            for row in trades[transaction_type]:
                                row_num, counter, running_long, running_short = self.write_borrow_return(
                                    ws, row_num, counter, row, bank_name, bank_id,
                                    running_long, running_short)

                    for transaction_type in DERIVATIVES:
                        if transaction_type in trades:
                            for row in trades[transaction_type]:
                                row_num, counter, _ = self.write_line(ws, row_num, counter, row)
                                row_num += 1
                                ws.row_dimensions[row_num].height = ROW_HEIGHT
                                border_line(ws, row_num)

                # Additional empty lines after each bank name
                row_num += 1
                ws.row_dimensions[row_num].height = ROW_HEIGHT
                border_line(ws, row_num)

        return row_num

    def _opening_balance(self, bank_name, code):
        row = self.db.execute(
            "SELECT SUM(tbl_transaction.quantity) AS balance "
            "FROM tbl_transaction "
            "INNER JOIN tbl_bank_account ON tbl_bank_account.ref_num = tbl_transaction.bank_ref "
            "INNER JOIN tbl_code ON tbl_code.ref_num = tbl_transaction.code_ref "
            "WHERE tbl_bank_account.bank_name = ? "
            "  AND tbl_code.code = ? "
            "  AND tbl_transaction.trade_date < ?",
            (bank_name, code, self.trade_date)
        ).fetchone()
        try:
            return int(row[0]) if row and row[0] is not None else 0
        except (TypeError, IndexError):
            return 0

    def _opening_short_balance(self, bank_name, code):
        row = self.db.execute(
            "SELECT SUM(tbl_transaction_short.quantity) AS balance "
            "FROM tbl_transaction_short "
            "INNER JOIN tbl_bank_account ON tbl_bank_account.ref_num = tbl_transaction_short.bank_ref "
            "INNER JOIN tbl_code ON tbl_code.ref_num = tbl_transaction_short.code_ref "
            "WHERE tbl_bank_account.bank_name = ? "
            "  AND tbl_code.code = ? "
            "  AND tbl_transaction_short.trade_date < ?",
            (bank_name, code, self.trade_date)
        ).fetchone()
        try:
            return int(row[0]) if row and row[0] is not None else 0
        except (TypeError, IndexError):
            return 0

    def write_borrow_return(self, ws, row_num, counter, row, bank_name, bank_id,
                            prior_long, prior_short):
        # prior_long / prior_short are the RUNNING long/short balances threaded in
        # by write_transactions -- the balance after every earlier same-day
        # transaction of this (bank, stock), not a fresh start-of-day re-query.
        # This makes multiple same-stock transactions follow through (e.g. a Sell
        # then a Return: Long beginning = the post-Sell balance, not the day open).
        ccy = row["ccy_id"]
        transaction_type = row["transaction_type"]
        stock_name = row["stock_name"]
        code = row["code"]
        quantity = abs(int(row["quantity"]))
        price = row["price"]

        is_return = transaction_type == "Return Shares"
        long_sign = "-" if is_return else "+"
        short_sign = "+" if is_return else "-"

        price_str = str(price)
        dec_len = len(price_str) - price_str.find(".") - 1
        number_format = "#,##0.00" if dec_len <= 2 else "#,##0.0000"

        r1 = row_num      # title
        r2 = row_num + 1  # hidden data (R, S) + formula summary
        r3 = row_num + 2  # Long / Short headers
        r4 = row_num + 3  # prior positions
        r5 = row_num + 4  # change (+/- quantity)
        r6 = row_num + 5  # result (top rule)
        r7 = row_num + 6  # blank spacer

        # --- Row 1: counter + title ---
        border_line(ws, r1)
        ws[f"B{r1}"].value = counter
        ws[f"B{r1}"].font = Font(size=13)
        ws[f"B{r1}"].number_format = "0)"
        ws[f"B{r1}"].alignment = Alignment(horizontal="right")
        ws[f"C{r1}"].value = f"{transaction_type} - {stock_name} ({code})"
        ws[f"C{r1}"].font = Font(size=13)
        ws.row_dimensions[r1].height = ROW_HEIGHT

        # --- Row 2: hidden data + formula summary ---
        border_line(ws, r2)
        ws[f"R{r2}"].value = quantity
        ws[f"S{r2}"].value = price
        ws[f"C{r2}"].value = (
            f'=TEXT(R{r2},"#,###,###,##0")&" shares @ "'
            f'&TEXT(S{r2},"{number_format}")&'
            f'" Average = {ccy} "&TEXT(R{r2}*S{r2},"#,###,###,###,##0.00")'
        )
        ws[f"C{r2}"].font = Font(size=13)
        ws.row_dimensions[r2].height = ROW_HEIGHT

        # --- Row 3: Long / Short headers (font 11, merged D:G and L:O) ---
        border_line(ws, r3)
        ws[f"D{r3}"].value = "Long"
        ws[f"D{r3}"].font = Font(size=11)
        ws[f"D{r3}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"D{r3}:G{r3}")
        ws[f"L{r3}"].value = "Short"
        ws[f"L{r3}"].font = Font(size=11)
        ws[f"L{r3}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"L{r3}:O{r3}")
        ws.row_dimensions[r3].height = ROW_HEIGHT

        # --- Row 4: prior positions (merged D:G and L:O) ---
        border_line(ws, r4)
        ws[f"D{r4}"].value = int(prior_long)
        ws[f"D{r4}"].font = Font(size=13)
        ws[f"D{r4}"].number_format = "#,##0_ ;[Red](#,##0)"
        ws.merge_cells(f"D{r4}:G{r4}")
        ws[f"L{r4}"].value = int(prior_short)
        ws[f"L{r4}"].font = Font(size=13)
        ws[f"L{r4}"].number_format = "#,##0_ ;[Red](#,##0)"
        ws.merge_cells(f"L{r4}:O{r4}")
        ws.row_dimensions[r4].height = ROW_HEIGHT

        # --- Row 5: change — live formula =R{r2}, merged D:G and L:O ---
        # Short sign moves to K (one left) to free P for Average in Borrow
        border_line(ws, r5)
        ws[f"C{r5}"].value = long_sign
        ws[f"C{r5}"].font = Font(size=13)
        ws[f"C{r5}"].alignment = Alignment(horizontal="right")
        ws[f"D{r5}"].value = f"=R{r2}"
        ws[f"D{r5}"].font = Font(size=13)
        ws[f"D{r5}"].number_format = "#,##0"
        ws.merge_cells(f"D{r5}:G{r5}")
        ws[f"K{r5}"].value = short_sign
        ws[f"K{r5}"].font = Font(size=13)
        ws[f"K{r5}"].alignment = Alignment(horizontal="right")
        ws[f"L{r5}"].value = f"=R{r2}"
        ws[f"L{r5}"].font = Font(size=13)
        ws[f"L{r5}"].number_format = "#,##0"
        ws.merge_cells(f"L{r5}:O{r5}")
        if not is_return:
            ws[f"P{r5}"].value = "Average"
            ws[f"P{r5}"].font = Font(size=11)
            ws[f"P{r5}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[r5].height = ROW_HEIGHT

        # --- Row 6: result — live formulas, top rule ---
        border_line(ws, r6)
        result_border = Border(top=Side(style="thin"))
        long_formula = f"=D{r4}-D{r5}" if is_return else f"=D{r4}+D{r5}"
        short_formula = f"=+L{r4}+L{r5}" if is_return else f"=L{r4}-L{r5}"
        ws[f"C{r6}"].value = f'=IF(D{r6}>=0,"shares","short")'
        ws[f"C{r6}"].font = Font(size=13)
        ws[f"C{r6}"].alignment = Alignment(horizontal="right")
        ws[f"D{r6}"].value = long_formula
        ws[f"D{r6}"].font = Font(size=13)
        ws[f"D{r6}"].border = result_border
        ws[f"D{r6}"].number_format = "#,##0_ ;[Red](#,##0)"
        ws.merge_cells(f"D{r6}:G{r6}")
        ws[f"K{r6}"].value = f'=IF(L{r6}>=0,"shares","short")'
        ws[f"K{r6}"].font = Font(size=13)
        ws[f"K{r6}"].alignment = Alignment(horizontal="right")
        ws[f"L{r6}"].value = short_formula
        ws[f"L{r6}"].font = Font(size=13)
        ws[f"L{r6}"].border = result_border
        ws[f"L{r6}"].number_format = "#,##0_ ;[Red](#,##0)"
        ws.merge_cells(f"L{r6}:O{r6}")
        if not is_return:
            average = TradesDoneAverage(
                db=self.db, trade_date=self.trade_date,
                code=code, bank_id=bank_id
            ).average
            ws[f"P{r6}"].value = average
            ws[f"P{r6}"].font = Font(size=13)
            ws[f"P{r6}"].number_format = "#,##0.0000"
            ws[f"P{r6}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[r6].height = ROW_HEIGHT

        # --- Row 7: blank spacer ---
        border_line(ws, r7)
        ws.row_dimensions[r7].height = ROW_HEIGHT

        counter += 1
        # Carry the running balances forward for the next same-stock transaction.
        new_long = prior_long - quantity if is_return else prior_long + quantity
        new_short = prior_short + quantity if is_return else prior_short - quantity
        return r7 + 1, counter, new_long, new_short

    def write_transfers(self, ws, row_num):
        if not self.transfers:
            return row_num

        # Section header: C:P merged
        border_line(ws, row_num)
        cell = ws[f"C{row_num}"]
        cell.value     = "Transfer of Stocks"
        cell.font      = Font(size=13, bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"C{row_num}:P{row_num}")
        ws.row_dimensions[row_num].height = ROW_HEIGHT
        row_num += 1

        counter = 1
        # Running balances per (account, code): each transfer's opening carries
        # from the prior same-day transfer of that stock in that account, rather
        # than re-reading the day-opening balance every time.
        running = {}
        for pair in self.transfers:
            out_bank   = pair['out_bank']
            in_bank    = pair['in_bank']
            in_bank_id = pair['in_bank_id']
            stock_name = pair['stock_name']
            code       = pair['code']
            qty        = int(pair['quantity'])

            if (out_bank, code) not in running:
                running[(out_bank, code)] = self._opening_balance(out_bank, code)
            if (in_bank, code) not in running:
                running[(in_bank, code)] = self._opening_balance(in_bank, code)
            opening_out = running[(out_bank, code)]
            opening_in  = running[(in_bank, code)]

            # Row 1: counter + title
            border_line(ws, row_num)
            ws[f"B{row_num}"].value         = counter
            ws[f"B{row_num}"].font          = Font(size=13)
            ws[f"B{row_num}"].number_format = "0\\)"
            ws[f"C{row_num}"].value         = f"Transfer {stock_name} ({code})"
            ws[f"C{row_num}"].font          = Font(size=13)
            ws[f"C{row_num}"].alignment     = Alignment(horizontal="left")
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Row 2: from
            border_line(ws, row_num)
            ws[f"C{row_num}"].value     = f"from {out_bank}"
            ws[f"C{row_num}"].font      = Font(size=13)
            ws[f"C{row_num}"].alignment = Alignment(horizontal="left")
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Row 3: to
            border_line(ws, row_num)
            ws[f"C{row_num}"].value     = f"to {in_bank}"
            ws[f"C{row_num}"].font      = Font(size=13)
            ws[f"C{row_num}"].alignment = Alignment(horizontal="left")
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Row 4: quantity in shares
            border_line(ws, row_num)
            ws[f"C{row_num}"].value     = f"{qty:,} shares"
            ws[f"C{row_num}"].font      = Font(size=13)
            ws[f"C{row_num}"].alignment = Alignment(horizontal="left")
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Row 5: bank name headers
            border_line(ws, row_num)
            ws[f"C{row_num}"].value = out_bank
            ws[f"C{row_num}"].font  = Font(size=12)
            ws[f"K{row_num}"].value = in_bank
            ws[f"K{row_num}"].font  = Font(size=12)
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Row 6: opening balances
            border_line(ws, row_num)
            cell = ws[f"E{row_num}"]
            cell.value         = opening_out
            cell.font          = Font(size=13)
            cell.alignment     = Alignment(horizontal="right")
            cell.number_format = "#,##0"
            ws.merge_cells(f"E{row_num}:H{row_num}")
            cell = ws[f"M{row_num}"]
            cell.value         = opening_in
            cell.font          = Font(size=13)
            cell.alignment     = Alignment(horizontal="right")
            cell.number_format = "#,##0"
            ws.merge_cells(f"M{row_num}:P{row_num}")
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Row 7: movement (- out, + in)
            border_line(ws, row_num)
            ws[f"D{row_num}"].value     = "-"
            ws[f"D{row_num}"].font      = Font(size=13)
            ws[f"D{row_num}"].alignment = Alignment(horizontal="right")
            cell = ws[f"E{row_num}"]
            cell.value         = qty
            cell.font          = Font(size=13)
            cell.alignment     = Alignment(horizontal="right")
            cell.number_format = "#,##0"
            ws.merge_cells(f"E{row_num}:H{row_num}")
            ws[f"L{row_num}"].value     = "+"
            ws[f"L{row_num}"].font      = Font(size=13)
            ws[f"L{row_num}"].alignment = Alignment(horizontal="right")
            cell = ws[f"M{row_num}"]
            cell.value         = qty
            cell.font          = Font(size=13)
            cell.alignment     = Alignment(horizontal="right")
            cell.number_format = "#,##0"
            ws.merge_cells(f"M{row_num}:P{row_num}")
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Row 8: closing balances — formula (opening +/- movement) with a
            # top "total" rule above each balance, like the remaining-shares line.
            border_line(ws, row_num)
            balance_rule = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin', color='00C0C0C0'),
            )
            ws[f"D{row_num}"].value     = "shares"
            ws[f"D{row_num}"].font      = Font(size=13)
            ws[f"D{row_num}"].alignment = Alignment(horizontal="right")
            cell = ws[f"E{row_num}"]
            cell.value         = f"=E{row_num-2}-E{row_num-1}"
            cell.font          = Font(size=13)
            cell.alignment     = Alignment(horizontal="right")
            cell.number_format = "#,##0"
            ws.merge_cells(f"E{row_num}:H{row_num}")
            for col in ("E", "F", "G", "H"):
                ws[f"{col}{row_num}"].border = balance_rule
            ws[f"L{row_num}"].value     = "shares"
            ws[f"L{row_num}"].font      = Font(size=13)
            ws[f"L{row_num}"].alignment = Alignment(horizontal="right")
            cell = ws[f"M{row_num}"]
            cell.value         = f"=M{row_num-2}+M{row_num-1}"
            cell.font          = Font(size=13)
            cell.alignment     = Alignment(horizontal="right")
            cell.number_format = "#,##0"
            ws.merge_cells(f"M{row_num}:P{row_num}")
            for col in ("M", "N", "O", "P"):
                ws[f"{col}{row_num}"].border = balance_rule
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Row 9: average at destination bank
            border_line(ws, row_num)
            ws[f"K{row_num}"].value = "Average"
            ws[f"K{row_num}"].font  = Font(size=13)
            average = TradesDoneAverage(
                db=self.db, trade_date=self.trade_date,
                code=code, bank_id=in_bank_id,
                include_transfers=True
            ).average
            cell = ws[f"N{row_num}"]
            cell.value         = average
            cell.font          = Font(size=13)
            cell.alignment     = Alignment(horizontal="center")
            cell.number_format = "#,##0.0000"
            ws.merge_cells(f"N{row_num}:P{row_num}")
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Blank spacing row
            border_line(ws, row_num)
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            row_num += 1

            # Carry balances forward for the next same-day transfer of this stock.
            running[(out_bank, code)] = opening_out - qty
            running[(in_bank, code)]  = opening_in + qty

            counter += 1

        return row_num

    def write_line(self, ws, row_num, counter, row):
        ccy = row["ccy_id"]
        transaction_type = row["transaction_type"]
        stock_name = row["stock_name"]
        code = row["code"]
        cell_reference = None

        if transaction_type in STOCK_TRANSACTIONS:
            # Stock line
            cell = ws[f"B{row_num}"]
            cell.value = counter
            cell.font = Font(size=13)
            cell.number_format = "0)"
            cell.alignment = Alignment(horizontal="right")

            cell = ws[f"C{row_num}"]
            cell.value = f"{transaction_type} {stock_name} ({code})"
            cell.font = Font(size=13)

            row_num += 1
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            border_line(ws, row_num)

            # Transaction Line
            cell_reference = f"R{row_num}"
            ws[f"R{row_num}"].value = abs(row["quantity"])
            ws[f"S{row_num}"].value = row["price"]

            strike_spot = "Spot" if "Spot" in transaction_type else "Strike"
            number_format = "#,##0.00" if len(str(row["price"])[-(len(str(row["price"])) - (str(row["price"]).find(".")) - 1):]) <= 2 else "#,##0.0000"

            cell = ws[f"C{row_num}"]
            cell.value = f'=TEXT(R{row_num},"#,###,###,##0")&" shares @ "&TEXT(S{row_num},"{number_format}")&" ' \
                         f'{strike_spot} = {ccy} "&TEXT(R{row_num}*S{row_num},"#,###,###,###,##0.00")'
            cell.font = Font(size=13)

            row_num += 1
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            border_line(ws, row_num)

            # Spot and KO Line
            if "(Accu" in transaction_type or "(Decu" in transaction_type:
                cell = ws[f"C{row_num}"]
                cell.value = "Spot"
                cell.font = Font(size=13)

                cell = ws[f"E{row_num}"]
                cell.value = row["spot"]
                cell.font = Font(size=13)
                cell.alignment = Alignment(horizontal="left")
                cell.number_format = "#,##0.00" if len(str(row["spot"])[-(len(str(row["spot"])) - (str(row["spot"]).find(".")) - 1):]) <= 2 else "#,##0.0000"
                ws.merge_cells(f"E{row_num}:G{row_num}")

                cell = ws[f"I{row_num}"]
                cell.value = "KO"
                cell.font = Font(size=13)

                cell = ws[f"J{row_num}"]
                cell.value = row["ko"]
                cell.font = Font(size=13)
                cell.alignment = Alignment(horizontal="left")
                cell.number_format = "#,##0.00" if len(str(row["ko"])[-(len(str(row["ko"])) - (str(row["ko"]).find(".")) - 1):]) <= 2 else "#,##0.0000"
                ws.merge_cells(f"J{row_num}:L{row_num}")

                row_num += 1
                ws.row_dimensions[row_num].height = ROW_HEIGHT
                border_line(ws, row_num)

        elif transaction_type in DERIVATIVES:
            # Stock line
            cell = ws[f"B{row_num}"]
            cell.value = counter
            cell.font = Font(size=13)
            cell.number_format = "0)"
            cell.alignment = Alignment(horizontal="right")

            cell = ws[f"C{row_num}"]
            cell.value = f"New {transaction_type.title()}mulator - {stock_name} ({code})"
            cell.font = Font(size=13)

            row_num += 1
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            border_line(ws, row_num)

            # Terms line
            tenor = row["tenor"].replace("m", "")
            if row["gtd"] in ("1m", "Yes"):
                gtd = "1 month GTD"
            elif row["gtd"] == "No":
                gtd = "No GTD"
            else:
                gtd = f"{row['gtd'].replace('m', '')} months GTD"

            if row["leveraged"] == "Yes":
                shares = '{:,.0f}'.format(row["daily_shares"]) + " / " + '{:,.0f}'.format(row["daily_shares"] * 2)
            else:
                shares = '{:,.0f}'.format(row["daily_shares"])

            cell = ws[f"C{row_num}"]
            cell.value = f"{tenor} months, {gtd}, shares {shares} "
            cell.font = Font(size=13)

            row_num += 1
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            border_line(ws, row_num)

            # Spot and Strike Line
            cell = ws[f"C{row_num}"]
            cell.value = "Spot"
            cell.font = Font(size=13)

            cell = ws[f"E{row_num}"]
            cell.value = row["spot"]
            cell.font = Font(size=13)
            cell.alignment = Alignment(horizontal="left")
            cell.number_format = "#,##0.0000"
            ws.merge_cells(f"E{row_num}:G{row_num}")

            cell = ws[f"H{row_num}"]
            cell.value = "Strike"
            cell.font = Font(size=13)

            cell = ws[f"J{row_num}"]
            cell.value = row["strike_rate"] / 100
            cell.font = Font(size=13)
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0.00%"
            ws.merge_cells(f"J{row_num}:L{row_num}")

            cell = ws[f"M{row_num}"]
            cell.value = f"=E{row_num}*J{row_num}"
            cell.font = Font(size=13)
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "#,##0.0000"
            ws.merge_cells(f"M{row_num}:O{row_num}")

            row_num += 1
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            border_line(ws, row_num)

            # KO line
            cell = ws[f"H{row_num}"]
            cell.value = "KO"
            cell.font = Font(size=13)

            cell = ws[f"J{row_num}"]
            cell.value = row["ko_rate"] / 100
            cell.font = Font(size=13)
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0.00%"
            ws.merge_cells(f"J{row_num}:L{row_num}")

            cell = ws[f"M{row_num}"]
            cell.value = f"=E{row_num-1}*J{row_num}"
            cell.font = Font(size=13)
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "#,##0.0000"
            ws.merge_cells(f"M{row_num}:O{row_num}")

            row_num += 1
            ws.row_dimensions[row_num].height = ROW_HEIGHT
            border_line(ws, row_num)

        counter += 1
        return row_num, counter, cell_reference


def border_line(ws, row_num):
    for column in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"):
        cell = ws[f'{column}{row_num}']
        cell.border = Border(
            top=Side(style='thin', color='00C0C0C0'),
            bottom=Side(style='thin', color='00C0C0C0')
        )