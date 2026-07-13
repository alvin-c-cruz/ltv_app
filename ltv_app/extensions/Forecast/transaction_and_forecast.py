import sqlite3
from datetime import datetime
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
import os

from .forecast import forecast as futures
from .stock_price import Stock_Price

class transaction_and_forecast:
  def __init__(self, db = None, code = None, template_path = None, save_path = None, trade_date = None):
    ## This class will need database connection and stock code
    ## It is to be understood that the database is updated and correct
    ##  else the generated forecast will be incorrect
    
    db.row_factory = sqlite3.Row  # accessing the columns by name

    self.db = db                  # An sqlite connection
    self.code = code              # ei. '2333'
    self.save_path = save_path    # Folder where to save the generated excel file

    # trade_date is now() if not explicitly given
    if trade_date is None: trade_date = str(datetime.now().date()) 
    print(f'Reporting date is: {trade_date}')
    self.beginning_date = trade_date[:7] + '-01'

    # Used for forecast iteration and excel creation
    self.banks = {bank[0]: bank[1] for bank in self.get_banks()}

    beginning_balance = self.get_beginning_balance()
    transactions = self.get_transactions()
    forecast_data = futures(db, code)

    # The main variable for this class
    data = {}
    for bank_id in self.banks:
      data[bank_id] = {
        'beginning_balance': beginning_balance[bank_id] if bank_id in beginning_balance else 0,
        'average': 0,
        'transactions': transactions[bank_id] if bank_id in transactions else [],
        'forecast': forecast_data[bank_id]
      }

    stock_price = Stock_Price(db, code)
    closing_price = stock_price.get_price(trade_date)

    self.filename = CreateExcel(
            db=db,
            data=data, 
            trade_date=trade_date, 
            code=code, 
            template_path=template_path, 
            save_path=save_path,
            closing_price=closing_price).filename

    # Convert dictionary to json for better presentation
    # and also for possible microservice function
    self.json_data = json.dumps(data, indent=4)
    


  def get_banks(self):
    sql = "SELECT bank_id, bank_name FROM tbl_bank_account ORDER BY priority;"
    banks = self.db.execute(sql).fetchall()
    return banks


  def get_beginning_balance(self):
    sql = f"""
    SELECT tbl_bank_account.bank_id, SUM(tbl_transaction.quantity) 
    FROM tbl_transaction
    INNER JOIN tbl_bank_account ON tbl_transaction.bank_ref = tbl_bank_account.ref_num
    INNER JOIN tbl_code ON tbl_transaction.code_ref = tbl_code.ref_num
    WHERE tbl_transaction.trade_date < "{self.beginning_date}"
      AND tbl_code.code = "{self.code}"
    GROUP BY tbl_bank_account.bank_id
    ORDER BY tbl_bank_account.priority
    """
    return {row[0]:row[1] for row in self.db.execute(sql).fetchall()}


  def get_transactions(self):
    sql = f"""
    SELECT  tbl_bank_account.bank_id, 
            tbl_transaction.trade_date, 
            tbl_transaction.transaction_type, 
            tbl_transaction.price, 
            tbl_transaction.quantity
    FROM tbl_transaction
    INNER JOIN tbl_bank_account ON tbl_transaction.bank_ref = tbl_bank_account.ref_num
    INNER JOIN tbl_code ON tbl_transaction.code_ref = tbl_code.ref_num
    WHERE tbl_transaction.trade_date >= "{self.beginning_date}"
      AND tbl_code.code = "{self.code}"
    ORDER BY tbl_bank_account.priority
    """
    transactions = {}
    for row in self.db.execute(sql).fetchall():
      bank_id, trade_date, transaction_type, price, quantity = row    
      if bank_id not in transactions: transactions[bank_id] = []
      transactions[bank_id].append((trade_date, transaction_type, price, quantity))

    return transactions
  

class CreateExcel:
  def __init__(self, 
            db=None, 
            data=None, 
            trade_date=None, 
            code=None, 
            template_path=None, 
            save_path=None,
            closing_price=None):

    self.db = db

    template_file = os.path.join(template_path, "transaction_forecast.xlsx")
    self.filename = os.path.join(save_path, f"{str(trade_date)} transaction_and_forecast {code}.xlsx")

    stock_name = self.get_stock_name(code)

    wb = load_workbook(template_file)

    for bank_id in data:
        self.write_sheet(wb, bank_id, data[bank_id], closing_price, stock_name)

    wb.save(self.filename)
    # wb.save(rf"excel_files\Forecasts\Cover Page\{str(trade_date)} transaction_and_forecast {code}.xlsx")
    wb.close()

  def get_bank_name(self, bank_id):
    return self.db.execute(
        """
        SELECT bank_name
        FROM tbl_bank_account
        WHERE bank_id = ?;
        """, (bank_id, )
        ).fetchone()[0]

  def get_stock_name(self, code):
    stock_name, ccy_id = self.db.execute(
        """
        SELECT c.stock_name, cy.ccy_id
        FROM tbl_code as c
        INNER JOIN tbl_currency as cy ON cy.ref_num = c.ccy_ref
        WHERE c.code = ?
        ;
        """, (code, )
        ).fetchone()

    return f"Stock: {stock_name} ({code} {ccy_id[:2]})"


  def write_sheet(self, wb, bank_id, bank_data, closing_price, stock_name):
    ws = wb[bank_id]
    bank_name = self.get_bank_name(bank_id)
    if all((not bank_data["beginning_balance"], not bank_data["transactions"], not bank_data["forecast"])):
        ws.sheet_state = 'hidden'
    else:
        self.write_heading(ws, bank_name, closing_price, bank_data["beginning_balance"], stock_name, bank_data["average"])
        row_num = 12  # Starting line for transactions
        row_num = self.write_transactions(ws, bank_data["transactions"], row_num)
        row_num = self.write_forecast(ws, bank_data["forecast"], row_num)

        ws.print_area = f"A1:Z{row_num}"


  def write_heading(self, ws, bank_name, closing_price, beginning_balance, stock_name, average):
    references = {
        'B4': bank_name,
        'G4': stock_name,
        "R2": closing_price,
        "T11": beginning_balance,
        "W11": average
    }

    for ref in references:
        cell = ws[ref]
        cell.value = references[ref]

  def write_transactions(self, ws, transactions, row_num):
    for transaction in transactions:
        trade_date, transaction_type, price, quantity = transaction

        year, month, day = trade_date.split("-")
        cols = {
            "A": datetime(int(year), int(month), int(day)),
            "B": transaction_type if transaction_type != "From Account (Pay Short)" else "Pay Short",
            "C": price,
            "Q": quantity,
            "T": f'=INDIRECT("T"&ROW()-1)+Q{row_num}',
            "W": f'=IF(T{row_num}<0,"",IF(INDIRECT("T"&ROW()-1)<0,C{row_num},IF(Q{row_num}>0,(INDIRECT("T"&ROW()-1)*INDIRECT("W"&ROW()-1)+C{row_num}*Q{row_num})/T{row_num},INDIRECT("W"&ROW()-1))))', 
            "AC":f'=MONTH(A{row_num})&B{row_num}',
            "AD":f'=Q{row_num}',
        }

        if "-KO" in transaction_type:
            cols['Y'] = 'KO'

        for col in cols:
            cell = ws[f"{col}{row_num}"]
            cell.value = cols[col]


        self.format_transaction(ws, row_num, transaction_type, cols)

        ws.merge_cells(f"Q{row_num}:R{row_num}")
        ws.merge_cells(f"T{row_num}:U{row_num}")

        row_num += 1

    for col in ('A', 'B', 'C', 'D', 'E', 'F', 
            'G', 'H', 'I', 'J', 'K', 'L', 'M', 
            'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U'):
        cell = ws[f"{col}{row_num}"]
        cell.border = Border(bottom=Side(style='double'))

    row_num += 1

    return row_num

  def format_transaction(self, ws, row_num, transaction_type, cols):
    font_size = lambda col: 9 if col in ('B', ) else 12
    font_family = 'Times New Roman'
    font_color = lambda : '00000000' if transaction_type in ("Buy (Accu-KO)", "Sell (Decu-KO)") else '00000000'
    bold_face = lambda col : True if col == "A" else False
    text_align = lambda col: 'center' if col in ('B', 'C') else 'right'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    for col in cols:
        cell = ws[f"{col}{row_num}"]

        if col != 'Y':
            cell.font = Font(name=font_family, size=font_size(col), color=font_color(), bold=bold_face(col))
            cell.alignment = Alignment(
                        horizontal=text_align(col),
                        vertical='center')
        else:
            cell.font = Font(name=font_family, size=8, color="00FF0000")
            cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center')

    for col in ('A', 'B', 'C', 'D', 'E', 'G', 'H', 'I', 'L', 'M', 'N', 'P', 'Q', 'R', 'T', 'U'):
        cell = ws[f"{col}{row_num}"]
        cell.border = thin_border

    #  Fill color for Buy or Sell
    fill_color = lambda : '00CCFFFF' if 'Sell' in transaction_type or 'Out' in transaction_type  or 'Pay Short' in transaction_type else '00FFFFFF'
    for col in ('A', 'B', 'C', 'Q'):
        cell = ws[f"{col}{row_num}"]
        cell.fill = PatternFill(
            patternType='solid',
            fgColor=fill_color()
            )

    #  Fill color for empty cells
    for col in ('D', 'E', 'G', 'H', 'I', 'L', 'M', 'N', 'P'):
        cell = ws[f"{col}{row_num}"]
        cell.fill = PatternFill(
            patternType='solid',
            fgColor='00C0C0C0'
            )

    ws[f'A{row_num}'].number_format = "dd-mmm-yy"
    ws[f'C{row_num}'].number_format = "#,###,##0.0000"
    ws[f'W{row_num}'].number_format = "#,###,##0.0000"
    ws[f'Q{row_num}'].number_format = "#,##0_);(#,##0)"
    ws[f'T{row_num}'].number_format = "#,##0_);[Red](#,##0)"

  def write_forecast(self, ws, forecasts, row_num):
    row_num += 2

    # Write Forecast Heading
    cell = ws[f'A{row_num}']
    cell.value = "Show calculated estimate"
    cell.font = Font(name='Times New Roman', size=14, bold=True)

    cell = ws[f'T{row_num}']
    cell.value = '=INDIRECT("T"&ROW()-4)'
    cell.font = Font(name='Times New Roman', size=14, color='00FFFFFF')

    cell = ws[f'W{row_num}']
    cell.value = '=INDIRECT("W"&ROW()-4)'
    cell.font = Font(name='Times New Roman', size=14, color='00FFFFFF')

    row_num += 1

    fixing_num = 1
    for forecast in forecasts:
        if forecast['fixing_num'] != fixing_num:
            fixing_num = forecast['fixing_num']
            cell = ws[f'T{row_num}']
            cell.value = f'=INDIRECT("T"&ROW()-1)'
            cell.font = Font(color='00FFFFFF')

            cell = ws[f'W{row_num}']
            cell.value = f'=INDIRECT("W"&ROW()-1)'
            cell.font = Font(color='00FFFFFF')

            row_num += 1

        year, month, day = forecast['fixing_date'].split("-")
        cols = {
            "A": datetime(int(year), int(month), int(day)),
            "B": forecast['transaction_type'],
            "C": forecast['strike'],
            "D": forecast['ko'],
            "E": forecast['single'],
            "G": f'=(H{row_num}*E{row_num}+I{row_num}*E{row_num})*IF(B{row_num}="DECU",-1,1)',
            "H": forecast['days_done'],
            "I": forecast['days_double'],
            "L": forecast['days_indicative'],
            "M": f'=E{row_num}*L{row_num}*IF(LEFT(B{row_num},4)="ACCU",1,-1)*IF(Y{row_num}="D",2,1)',
            "P": f'=H{row_num}+L{row_num}',
            "Q": f'=G{row_num}+M{row_num}',
            "T": f'=INDIRECT("T"&ROW()-1)+Q{row_num}',
            "W": f'=IF(T{row_num}<0,"",IF(INDIRECT("T"&ROW()-1)<0,C{row_num},IF(Q{row_num}>0,(INDIRECT("T"&ROW()-1)*INDIRECT("W"&ROW()-1)+C{row_num}*Q{row_num})/T{row_num},INDIRECT("W"&ROW()-1))))', 
            "Y": f'=IF(C{row_num}<=R$2,"D",IF(D{row_num}>=R$2,"KO","S"))' if forecast['transaction_type'] == 'DECU' else f'=IF(C{row_num}>=R$2,"D",IF(D{row_num}<=R$2,"KO","S"))',
            "Z": '' if forecast['gtd'] == 'No' else "GTD",
            "AC":f'=MONTH(A{row_num})&B{row_num}',
            "AD":f'=Q{row_num}',
            "AF": forecast['contract_ref'],
            "AG": forecast['frequency'],
            "AH": forecast['fixing_num'],
        }

        for col in cols:
            cell = ws[f"{col}{row_num}"]
            cell.value = cols[col]

        self.format_forecast(ws, row_num, forecast['transaction_type'], cols)

        ws.merge_cells(f"M{row_num}:N{row_num}")
        ws.merge_cells(f"Q{row_num}:R{row_num}")
        ws.merge_cells(f"T{row_num}:U{row_num}")


        row_num += 1

    return row_num


  def format_forecast(self, ws, row_num, transaction_type, cols):
    font_size = lambda col: 8 if col in ('Y', 'Z') else 12
    font_family = 'Times New Roman'
    font_color = lambda : '00FF0000' if transaction_type in ("Buy (Accu-KO)", "Sell (Decu-KO)") else '00000000'
    bold_face = lambda col : True if col == "A" else False
    text_align = lambda col: 'center' if col in ('B', 'C', 'Y', 'Z') else 'right'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    for col in cols:
        cell = ws[f"{col}{row_num}"]
        # cell.style.font.name = font_family
        cell.font = Font(name=font_family, size=font_size(col), color=font_color(), bold=bold_face(col))
        cell.alignment = Alignment(
                    horizontal=text_align(col),
                    vertical='center')

    for col in ('A', 'B', 'C', 'D', 'E', 'G', 'H', 'I', 'L', 'M', 'N', 'P', 'Q', 'R', 'T', 'U'):
        cell = ws[f"{col}{row_num}"]
        cell.border = thin_border

    #  Fill color for Buy or Sell
    fill_color = lambda : '0099CC00' if 'DECU' in transaction_type else '00FFFFFF'
    for col in ('A', 'B', 'C', 'D', 'E', 'Q'):
        cell = ws[f"{col}{row_num}"]
        cell.fill = PatternFill(
            patternType='solid',
            fgColor=fill_color()
            )

    ws[f'A{row_num}'].number_format = "dd-mmm-yy"
    ws[f'C{row_num}'].number_format = "#,###,##0.0000"
    ws[f'D{row_num}'].number_format = "#,###,##0.0000"
    ws[f'E{row_num}'].number_format = "#,###,##0"
    ws[f'G{row_num}'].number_format = "#,##0_);(#,##0)"
    ws[f'H{row_num}'].number_format = "#,###,##0"
    ws[f'I{row_num}'].number_format = "#,###,##0"
    ws[f'L{row_num}'].number_format = "#,###,##0"
    ws[f'M{row_num}'].number_format = "#,##0_);(#,##0)"
    ws[f'P{row_num}'].number_format = "#,###,##0"
    ws[f'Q{row_num}'].number_format = "#,##0_);(#,##0)"
    ws[f'T{row_num}'].number_format = "#,##0_);[Red](#,##0)"
    ws[f'W{row_num}'].number_format = "#,###,##0.0000"
