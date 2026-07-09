import datetime
from flask import Blueprint, current_app, g, request
import sqlite3
import os
from openpyxl import load_workbook

from .. auth import login_required
from ..bank import BankAccount
from ..currency import Currency
from ..stocks import Stocks

bp = Blueprint('database', __name__, url_prefix="/database")


@bp.route('/')
@login_required
def home():
    return "Database home"


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(current_app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row

    # log_request()

    return g.db


def log_request():
    request_url = str(request.url)
    request_data = ""
    if "user_url" not in g:
        if 'static' not in request_url:
            g.user_url = request_url
            save_log(request_url, request_data)
    else:
        if g.user_url != request_url:
            if 'static' not in request_url:
                g.user_url = request_url
                save_log(request_url, request_data)


def save_log(request_url, request_data):
    filename = os.path.join(current_app.instance_path, "data_logs.xlsx")
    wb = load_workbook(filename)
    ws = wb["LOGS"]

    #  Go to next empty row
    row_num = 1
    date_time = ws[f"A{row_num}"].value
    while date_time:
        row_num += 1
        date_time = ws[f"A{row_num}"].value

    date_time = datetime.datetime.now()

    ws[f"A{row_num}"].value = date_time
    ws[f"B{row_num}"].value = request_url
    ws[f"C{row_num}"].value = request_data

    wb.save(filename)
    wb.close()


@bp.before_app_request
def base_variables():
    db = get_db()
    if "bank_accounts" not in g:
        g.bank_accounts = db.execute(
            "SELECT ref_num, bank_name, bank_id FROM tbl_bank_account "
            "WHERE is_active = 1 ORDER BY priority"
        ).fetchall()

    if "stocks" not in g:
        stocks = Stocks(db=db)
        g.stocks = stocks.all(fields=["ref_num", "code", "stock_name", "ccy_ref"], order_by=["code"])

    if "currencies" not in g:
        currencies = Currency(db=db)
        g.currencies = currencies.all(fields=["ref_num", "ccy_id"], order_by=["priority"])
