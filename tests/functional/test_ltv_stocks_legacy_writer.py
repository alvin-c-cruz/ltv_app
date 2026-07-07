import sqlite3
from datetime import date

import openpyxl

from ltv_app.blueprints.ltv_stocks.legacy_port.excel_writer import (
    _write_contracts,
    build_workbook,
    week_dates,
)


class _FakeWorkingDay:
    """Weekend-only holiday check; no DB access needed for this unit test."""

    def is_holiday(self, d):
        return d.weekday() >= 5


def _record(ref_num, code, code_ref, *, received, total, next_date,
            start_date, end_date, ccy_id='HKD'):
    return {
        'ref_num': ref_num,
        'reference': f'REF{ref_num}',
        'bank_doc': f'DOC{ref_num}',
        'frequency': 'monthly',
        'stock_name': f'Stock {code} GTD 1m',
        'code': code,
        'code_ref': code_ref,
        'yahoo_ticker': f'{code}.HK',
        'shares': '1,000',
        'spot': 100.0,
        'strike': 95.0,
        'ko': 110.0,
        'start_date': start_date,
        'end_date': end_date,
        'total': total,
        'received': received,
        'next_date': next_date,
        'remaining': total - received,
        'ccy_id': ccy_id,
        'daily_shares': 1000,
        'leveraged': 'No',
        'indicative': 'YES',
        'status': 'active',
    }


def test_write_contracts_headers_and_data_rows():
    report_date = date(2026, 7, 6)
    date_range = week_dates(report_date)

    active = _record(
        1, '0700', 700, received=1.0, total=3.0,
        next_date=date(2026, 8, 1),
        start_date=date(2026, 1, 1), end_date=date(2026, 12, 31),
    )
    done = _record(
        2, '9988', 9988, received=3.0, total=3.0,
        next_date=None,
        start_date=date(2026, 1, 1), end_date=date(2026, 3, 31),
    )
    records = [active, done]

    wb = openpyxl.Workbook()
    ws = wb.active

    next_row = _write_contracts(
        ws, records, 'ACCU', 1, report_date, date_range,
        _FakeWorkingDay(), price_lookup=lambda code_ref, d: None,
    )

    values = {cell.value for row in ws.iter_rows() for cell in row}
    assert 'ACCUMULATOR' in values
    assert 'RCVD mos.' in values
    assert 'NEXT MO.' in values

    first_data_row = 1 + 4
    active_row = first_data_row
    done_row = first_data_row + 1

    assert ws[f'K{active_row}'].value == f'=L{active_row}-J{active_row}'
    assert ws[f'K{done_row}'].value == f'=L{done_row}-J{done_row}'

    assert ws[f'N{done_row}'].value == 'DONE'
    assert isinstance(ws[f'N{active_row}'].value, date)

    assert ws['A1'].value == f'=2-COUNTIF(N{first_data_row}:N{done_row},"*DONE*")'

    assert ws.column_dimensions['C'].hidden is True
    assert ws.column_dimensions['M'].hidden is True

    assert next_row == done_row + 1 + 2


def test_week_dates_uses_isoweekday_and_has_ten_dates():
    report_date = date(2026, 7, 6)  # Monday
    dates = week_dates(report_date)
    assert len(dates) == 10
    # Monday isoweekday() == 1 -> start = report_date - timedelta(days=7)
    assert dates[0] == date(2026, 6, 29)
    assert dates == [
        date(2026, 6, 29), date(2026, 6, 30), date(2026, 7, 1),
        date(2026, 7, 2), date(2026, 7, 3),
        date(2026, 7, 6), date(2026, 7, 7), date(2026, 7, 8),
        date(2026, 7, 9), date(2026, 7, 10),
    ]


def test_write_contracts_zero_records_excludes_hidden_columns_from_border():
    """Verify that zero-contracts blank row excludes C and M from border (legacy port)."""
    report_date = date(2026, 7, 6)
    date_range = week_dates(report_date)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Call with empty records list
    next_row = _write_contracts(
        ws, [], 'ACCU', 1, report_date, date_range,
        _FakeWorkingDay(), price_lookup=lambda code_ref, d: None,
    )

    # Blank row is at first_data_row (1 + 4 = 5)
    blank_row = 1 + 4

    # Columns C (3) and M (13) should have NO border
    cell_c = ws.cell(blank_row, 3)  # Column C
    cell_m = ws.cell(blank_row, 13)  # Column M
    assert cell_c.border.left.style is None, "Column C should have no border in zero-contracts row"
    assert cell_m.border.left.style is None, "Column M should have no border in zero-contracts row"

    # Visible columns like B (2) and N (14) SHOULD have borders
    cell_b = ws.cell(blank_row, 2)  # Column B
    cell_n = ws.cell(blank_row, 14)  # Column N
    assert cell_b.border.left.style is not None, "Column B should have border in zero-contracts row"
    assert cell_n.border.left.style is not None, "Column N should have border in zero-contracts row"


# --- build_workbook: full workbook assembly (positions + contracts + cross-sheet totals) ---

BANK_REF = 1       # seeded 'CB1', indicative=NULL, transaction_basis='value_date'
BANK_ID = 'CB1'
CODE_REF = 1       # seeded '700'


def _txn(conn, ref, bank_ref, code_ref, trade_date, quantity, price=100.0, ttype='Buy (Spot)'):
    conn.execute(
        "INSERT INTO tbl_transaction "
        "(ref_num, trade_date, value_date, bank_ref, code_ref, transaction_type, quantity, price, "
        " brokerage, commission, foreign_charge, stamp_duty, misc) "
        "VALUES (?,?,?,?,?,?,?,?,0,0,0,0,0)",
        (ref, trade_date, trade_date, bank_ref, code_ref, ttype, quantity, price)
    )


def _contract(conn, ref, ttype, *, bank_ref=BANK_REF, code_ref=CODE_REF, leveraged='No',
              daily=1000, spot=100.0, strike_rate=95.0, ko_rate=110.0, status='active',
              frequency='monthly', start='2026-01-01'):
    conn.execute(
        "INSERT INTO tbl_stock_contract (ref_num, reference, bank_ref, code_ref, trade_date, "
        " start_date, transaction_type, daily_shares, leveraged, spot, strike_rate, ko_rate, "
        " tenor, frequency, gtd, bank_doc, status) VALUES (?,?,?,?,?,?,?,?,?,?,?,?, '12m',?,'1m', 'DOC',?)",
        (ref, f'REF{ref}', bank_ref, code_ref, start, start, ttype, daily, leveraged, spot,
         strike_rate, ko_rate, frequency, status))


def _period(conn, ref, end_date, received='', days='20'):
    conn.execute("INSERT INTO tbl_stock_contract_period "
                 "(contract_ref, start_date, end_date, days, received, gtd) "
                 "VALUES (?, '2026-01-01', ?, ?, ?, '1m')", (ref, end_date, days, received))


def test_build_workbook_assembles_sheets_and_positions_block(app):
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row

    # A position: a spot balance in code '700'.
    _txn(conn, 1, BANK_REF, CODE_REF, '2026-01-01', 50000)

    # An active DECU contract (with a period) so the ACCU/DECU tables aren't empty.
    _contract(conn, 200, 'DECU')
    _period(conn, 200, '2026-07-31')
    conn.commit()

    report_date = date(2026, 7, 6)
    buf = build_workbook(conn, report_date, [BANK_ID])
    conn.close()

    wb = openpyxl.load_workbook(buf)

    assert 'closing_price' in wb.sheetnames
    assert 'record' in wb.sheetnames
    assert 'CB1-HKD' in wb.sheetnames

    # closing_price/record are manual-paste targets and stay empty.
    assert wb['closing_price'].max_row == 1
    assert wb['closing_price']['A1'].value is None
    assert wb['record'].max_row == 1
    assert wb['record']['A1'].value is None

    ws = wb['CB1-HKD']
    assert ws.column_dimensions['C'].hidden is True

    # Locate the positions row: the row where column G holds the
    # `=D{r}+E{r}` total-shares formula (distinct from the contracts table,
    # where G holds the literal K/O price).
    position_row = None
    for row_cells in ws.iter_rows(min_col=7, max_col=7):
        cell = row_cells[0]
        if isinstance(cell.value, str) and cell.value.startswith('=D') and '+E' in cell.value:
            position_row = cell.row
            break

    assert position_row is not None, "expected a positions row with the G=D+E formula"
    r = position_row
    assert ws[f'B{r}'].value == '700'

    assert ws[f'G{r}'].value == f'=D{r}+E{r}'
    assert ws[f'J{r}'].value == f'=(L{r}/I{r})-1'
    l_value = ws[f'L{r}'].value
    assert isinstance(l_value, str)
    assert l_value.startswith('=INDEX(closing_price!A:C,MATCH(')
    assert 'MATCH(' in l_value and ',3)' in l_value


def test_build_workbook_skips_bank_with_no_data(app):
    """A bank with no transactions, no ACCU, no DECU produces no sheet at all."""
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    conn.commit()

    report_date = date(2026, 7, 6)
    buf = build_workbook(conn, report_date, [BANK_ID])
    conn.close()

    wb = openpyxl.load_workbook(buf)
    assert 'CB1-HKD' not in wb.sheetnames
    assert 'closing_price' in wb.sheetnames
    assert 'record' in wb.sheetnames


def test_build_workbook_unknown_bank_id_is_skipped():
    """A bank_id not present in tbl_bank_account is silently skipped, not an error."""
    import sqlite3 as _sqlite3
    from tests.functional.conftest import _SCHEMA, _seed

    conn = _sqlite3.connect(':memory:')
    conn.row_factory = _sqlite3.Row
    conn.executescript(_SCHEMA)
    _seed(conn)

    report_date = date(2026, 7, 6)
    buf = build_workbook(conn, report_date, ['NOPE'])
    conn.close()

    wb = openpyxl.load_workbook(buf)
    assert wb.sheetnames == ['closing_price', 'record']
