"""Active-count / KO handling for the LTV Stocks report."""
import sqlite3
from datetime import date

from openpyxl import load_workbook

from ltv_app.blueprints.ltv_stocks.create_ltv_stocks import _load_contracts
from ltv_app.blueprints.ltv_stocks.views import _active_count, _generate_excel


def _add_contract(conn, ref_num, ttype, status, *, tenor='12m', trade_date='2026-01-01'):
    """Insert a contract with a single period.

    With the default tenor ('12m') the contract expects 12 periods, so its one
    period leaves it not-DONE (remaining > 0). Pass tenor='1m' to make the one
    period complete the contract (remaining == 0). `trade_date` sets the
    contract's inception date used by the current-week filter.
    """
    conn.execute(
        "INSERT INTO tbl_stock_contract "
        "(ref_num, reference, bank_ref, code_ref, trade_date, start_date, "
        " transaction_type, daily_shares, leveraged, spot, strike_rate, ko_rate, "
        " tenor, frequency, gtd, bank_doc, status) "
        "VALUES (?, ?, 1, 1, ?, '2026-01-01', ?, 1000, 'No', "
        " 100.0, 95.0, 110.0, ?, 'monthly', '1m', 'DOC', ?)",
        (ref_num, f"REF{ref_num}", trade_date, ttype, tenor, status),
    )
    conn.execute(
        "INSERT INTO tbl_stock_contract_period "
        "(contract_ref, start_date, end_date, days, received, gtd) "
        "VALUES (?, '2026-01-01', '2026-01-31', '20', '1000', '1m')",
        (ref_num,),
    )


def test_load_contracts_sets_is_ko(app):
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    _add_contract(conn, 10, 'ACCU', 'active')
    _add_contract(conn, 11, 'ACCU', 'KO')
    conn.commit()

    result = {}
    _load_contracts(conn, result, {}, date(2026, 1, 1))
    conn.close()

    by_ref = {c['ref_num']: c for c in result['HKD']['CB1']['accu']}

    assert by_ref[10]['is_ko'] is False
    assert by_ref[11]['is_ko'] is True
    # KO contract is not DONE and keeps a date in the next-month column.
    assert by_ref[11]['is_done'] is False
    assert by_ref[11]['next_date'] != 'DONE'


def test_ko_overrides_done_when_all_periods_received(app):
    """A KO contract is never DONE, even when every period is received."""
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    _add_contract(conn, 20, 'ACCU', 'active', tenor='1m')  # completes -> DONE
    _add_contract(conn, 21, 'ACCU', 'KO', tenor='1m')      # completes but is KO
    conn.commit()

    result = {}
    _load_contracts(conn, result, {}, date(2026, 1, 1))
    conn.close()

    by_ref = {c['ref_num']: c for c in result['HKD']['CB1']['accu']}

    # Completed non-KO contract is DONE.
    assert by_ref[20]['is_done'] is True
    assert by_ref[20]['is_ko'] is False
    assert by_ref[20]['next_date'] == 'DONE'

    # KO overrides DONE: not DONE, keeps a next-month date (not 'DONE').
    assert by_ref[21]['is_ko'] is True
    assert by_ref[21]['is_done'] is False
    assert by_ref[21]['next_date'] != 'DONE'


def test_active_count_excludes_done_and_ko():
    contracts = [
        {'is_done': False, 'is_ko': False},  # active   -> counted
        {'is_done': True,  'is_ko': False},  # done     -> excluded
        {'is_done': False, 'is_ko': True},   # ko       -> excluded
        {'is_done': False, 'is_ko': False},  # active   -> counted
    ]
    assert _active_count(contracts) == 2
    assert _active_count([]) == 0


def _contract(**overrides):
    base = {
        'stock_name': 'Tencent', 'code': '700', 'code_ref': 1, 'bank_doc': 'DOC',
        'shares_day': '1,000', 'spot_raw': 100.0, 'strike_raw': 95.0, 'ko_raw': 110.0,
        'start_date_raw': date(2026, 1, 1), 'last_end_date_raw': date(2026, 1, 31),
        'received_months': 1.0, 'remaining_months': 11.0, 'total_months': 12,
        'next_date': '2026-08-01', 'is_done': False, 'is_ko': False,
    }
    base.update(overrides)
    return base


def test_generate_excel_count_is_plain_integer():
    accu = [
        _contract(),                    # active
        _contract(is_ko=True),          # ko
        _contract(is_done=True, next_date='DONE'),  # done
    ]
    data = {'HKD': {'CB1': {
        'bank_name': 'Citibank No. 1', 'bank_priority': 1, 'report_label': None,
        'accu': accu, 'decu': [], 'positions': {},
    }}}

    output = _generate_excel(data, date(2026, 7, 2), {}, [])
    wb = load_workbook(output)
    ws = wb['CB1-HKD']

    # ACCU count cell is column A of the section's first row (A3).
    value = ws['A3'].value
    assert value == 1
    assert isinstance(value, int)


def test_current_week_filters_done_and_ko(app):
    """Active always shown; DONE/KO only when trade_date is in the current week."""
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    # Report date 2026-07-02 (Thu) -> week Mon 2026-06-29 .. Fri 2026-07-03.
    report_date = date(2026, 7, 2)
    in_week, out_week = '2026-06-30', '2026-05-01'

    _add_contract(conn, 30, 'ACCU', 'active', trade_date=out_week)             # active, old
    _add_contract(conn, 31, 'ACCU', 'active', tenor='1m', trade_date=in_week)  # DONE in week
    _add_contract(conn, 32, 'ACCU', 'active', tenor='1m', trade_date=out_week) # DONE out of week
    _add_contract(conn, 33, 'ACCU', 'KO', trade_date=in_week)                  # KO in week
    _add_contract(conn, 34, 'ACCU', 'KO', trade_date=out_week)                 # KO out of week
    conn.commit()

    result = {}
    _load_contracts(conn, result, {}, report_date)
    conn.close()

    refs = {c['ref_num'] for c in result['HKD']['CB1']['accu']}
    assert 30 in refs            # active always listed, even though trade_date is old
    assert 31 in refs            # DONE in current week -> listed
    assert 33 in refs            # KO in current week -> listed
    assert 32 not in refs        # DONE out of week -> omitted
    assert 34 not in refs        # KO out of week -> omitted


def test_missing_trade_date_done_ko_omitted(app):
    """A DONE/KO contract with NULL trade_date is treated as not-in-week (omitted)."""
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    _add_contract(conn, 40, 'ACCU', 'KO', trade_date=None)
    conn.commit()

    result = {}
    _load_contracts(conn, result, {}, date(2026, 7, 2))
    conn.close()

    accu = result.get('HKD', {}).get('CB1', {}).get('accu', [])
    assert all(c['ref_num'] != 40 for c in accu)
