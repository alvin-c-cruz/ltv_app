"""
Integration tests — Trades Done download routes.

Old Trades Done: GET /trades/download/<trade_date>
New Trades Done: GET /trades/download_with_gain_loss/<trade_date>

Both routes:
  - require login
  - redirect to home with a flash when there is no data for the date
  - return a valid .xlsx file when data exists

New Trades Done adds gain/loss columns for sell transactions.
"""
import io

import pytest
from openpyxl import load_workbook

TEST_DATE  = '2026-05-18'
PRIOR_DATE = '2026-05-01'   # used to establish a cost basis before the sell
BANK_REF   = 1
CODE_REF   = 1

OLD_URL = f'/trades/download/{TEST_DATE}'
NEW_URL = f'/trades/download_with_gain_loss/{TEST_DATE}'

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
XLSX_MAGIC = b'PK'   # xlsx is a zip archive


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _insert_transaction(db_conn, **overrides):
    """Insert a tbl_transaction row and return its ref_num."""
    params = {
        'trade_date':       TEST_DATE,
        'value_date':       '2026-05-20',
        'bank_ref':         BANK_REF,
        'code_ref':         CODE_REF,
        'transaction_type': 'Buy (Spot)',
        'quantity':         1000,
        'price':            320.50,
        'brokerage':        0, 'commission': 0,
        'foreign_charge':   0, 'stamp_duty': 0, 'misc': 0,
        'locked':           0,
    }
    params.update(overrides)
    cur = db_conn.execute(
        """INSERT INTO tbl_transaction
           (trade_date, value_date, bank_ref, code_ref, transaction_type,
            quantity, price, brokerage, commission, foreign_charge,
            stamp_duty, misc, locked)
           VALUES (:trade_date, :value_date, :bank_ref, :code_ref,
                   :transaction_type, :quantity, :price, :brokerage,
                   :commission, :foreign_charge, :stamp_duty, :misc, :locked)""",
        params,
    )
    db_conn.commit()
    return cur.lastrowid


def _insert_buy_and_sell(db_conn):
    """Insert a prior buy (for cost basis) and a sell on TEST_DATE."""
    _insert_transaction(db_conn,
                        trade_date=PRIOR_DATE, value_date='2026-05-03',
                        transaction_type='Buy (Spot)', quantity=1000, price=300.00)
    _insert_transaction(db_conn,
                        trade_date=TEST_DATE, value_date='2026-05-20',
                        transaction_type='Sell (Spot)', quantity=-500, price=350.00)


# ===========================================================================
# 1. Old Trades Done — /trades/download/<trade_date>
# ===========================================================================
class OldTradesDoneTests:

    # ── Access control ──────────────────────────────────────────────────────

    def test_unauthenticated_redirects_to_login(self, client):
        response = client.get(OLD_URL)
        assert response.status_code == 302
        assert '/login' in response.headers['Location']

    # ── Empty state ─────────────────────────────────────────────────────────

    def test_no_data_redirects_to_home(self, auth_client):
        response = auth_client.get(OLD_URL)
        assert response.status_code == 302
        assert '/trades' in response.headers['Location']

    def test_no_data_shows_flash_message(self, auth_client):
        response = auth_client.get(OLD_URL, follow_redirects=True)
        assert b'No data to download' in response.data

    # ── Successful download with a Buy transaction ───────────────────────────

    def test_buy_transaction_returns_200(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(OLD_URL)
        assert response.status_code == 200

    def test_buy_transaction_content_type_is_xlsx(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(OLD_URL)
        assert XLSX_MIME in response.content_type

    def test_buy_transaction_response_is_valid_xlsx(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(OLD_URL)
        assert response.data[:2] == XLSX_MAGIC

    def test_buy_transaction_filename_contains_date(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(OLD_URL)
        disposition = response.headers.get('Content-Disposition', '')
        assert TEST_DATE in disposition

    def test_buy_transaction_filename_contains_trades_done(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(OLD_URL)
        disposition = response.headers.get('Content-Disposition', '')
        assert 'Trades Done' in disposition

    # ── Successful download with a Sell transaction ──────────────────────────

    def test_sell_transaction_returns_200(self, auth_client, db_conn):
        _insert_buy_and_sell(db_conn)
        response = auth_client.get(OLD_URL)
        assert response.status_code == 200

    def test_sell_transaction_response_is_valid_xlsx(self, auth_client, db_conn):
        _insert_buy_and_sell(db_conn)
        response = auth_client.get(OLD_URL)
        assert response.data[:2] == XLSX_MAGIC

    # ── Multiple banks (triggers bank divider code path) ────────────────────

    def test_multiple_transactions_same_date_returns_200(self, auth_client, db_conn):
        _insert_transaction(db_conn, quantity=1000, price=300.00)
        _insert_transaction(db_conn, quantity=500, price=310.00)
        response = auth_client.get(OLD_URL)
        assert response.status_code == 200


# ===========================================================================
# 2. New Trades Done — /trades/download_with_gain_loss/<trade_date>
# ===========================================================================
class NewTradesDoneTests:

    # ── Access control ──────────────────────────────────────────────────────

    def test_unauthenticated_redirects_to_login(self, client):
        response = client.get(NEW_URL)
        assert response.status_code == 302
        assert '/login' in response.headers['Location']

    # ── Empty state ─────────────────────────────────────────────────────────

    def test_no_data_redirects_to_home(self, auth_client):
        response = auth_client.get(NEW_URL)
        assert response.status_code == 302
        assert '/trades' in response.headers['Location']

    def test_no_data_shows_flash_message(self, auth_client):
        response = auth_client.get(NEW_URL, follow_redirects=True)
        assert b'No data to download' in response.data

    # ── Successful download with a Buy transaction ───────────────────────────

    def test_buy_transaction_returns_200(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(NEW_URL)
        assert response.status_code == 200

    def test_buy_transaction_content_type_is_xlsx(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(NEW_URL)
        assert XLSX_MIME in response.content_type

    def test_buy_transaction_response_is_valid_xlsx(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(NEW_URL)
        assert response.data[:2] == XLSX_MAGIC

    def test_buy_transaction_filename_contains_gain_loss(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(NEW_URL)
        disposition = response.headers.get('Content-Disposition', '')
        assert 'gain loss' in disposition.lower()

    def test_buy_transaction_filename_contains_date(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        response = auth_client.get(NEW_URL)
        disposition = response.headers.get('Content-Disposition', '')
        assert TEST_DATE in disposition

    # ── Successful download with a Sell transaction (gain/loss code path) ────

    def test_sell_transaction_returns_200(self, auth_client, db_conn):
        _insert_buy_and_sell(db_conn)
        response = auth_client.get(NEW_URL)
        assert response.status_code == 200

    def test_sell_transaction_response_is_valid_xlsx(self, auth_client, db_conn):
        _insert_buy_and_sell(db_conn)
        response = auth_client.get(NEW_URL)
        assert response.data[:2] == XLSX_MAGIC

    def test_sell_transaction_content_type_is_xlsx(self, auth_client, db_conn):
        _insert_buy_and_sell(db_conn)
        response = auth_client.get(NEW_URL)
        assert XLSX_MIME in response.content_type

    # ── Filename distinguishes new from old ──────────────────────────────────

    def test_new_filename_differs_from_old_filename(self, auth_client, db_conn):
        _insert_transaction(db_conn)
        old_resp = auth_client.get(OLD_URL)
        new_resp = auth_client.get(NEW_URL)
        old_disp = old_resp.headers.get('Content-Disposition', '')
        new_disp = new_resp.headers.get('Content-Disposition', '')
        assert old_disp != new_disp


# ===========================================================================
# 3. New Trades Done — gain/loss columns belong to Sell (Spot) blocks only
# ===========================================================================
def _first_sheet(response):
    wb = load_workbook(io.BytesIO(response.data))
    return wb[wb.sheetnames[0]]


def _column_values(ws, letters):
    values = {}
    for letter in letters:
        values[letter] = [
            cell.value for cell in ws[letter] if cell.value is not None
        ]
    return values


class GainLossColumnTests:

    def test_buy_only_has_no_gain_loss_columns(self, auth_client, db_conn):
        _insert_transaction(db_conn)  # Buy (Spot)
        response = auth_client.get(NEW_URL)
        ws = _first_sheet(response)
        values = _column_values(ws, 'LMN')
        assert values == {'L': [], 'M': [], 'N': []}

    def test_sell_short_has_no_gain_loss_columns(self, auth_client, db_conn):
        _insert_transaction(db_conn, transaction_type='Sell (Short)',
                            quantity=-500, price=310.00)
        response = auth_client.get(NEW_URL)
        ws = _first_sheet(response)
        values = _column_values(ws, 'LMN')
        assert values == {'L': [], 'M': [], 'N': []}

    def test_sell_spot_has_gain_loss_columns(self, auth_client, db_conn):
        _insert_buy_and_sell(db_conn)  # prior buy + Sell (Spot) on TEST_DATE
        response = auth_client.get(NEW_URL)
        ws = _first_sheet(response)
        values = _column_values(ws, 'LMN')
        assert values['L'], "expected cost formulas in column L"
        assert values['M'], "expected gain/loss formulas in column M"
        assert values['N'], "expected % formulas in column N"
