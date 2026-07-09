"""The instance/ directory must bootstrap itself on a fresh clone.

`instance/` is gitignored wholesale (`.gitignore: /instance/*`), so a fresh
clone has no `temp/` directory and no `data_logs.xlsx`. Both were previously
assumed to exist, which made every Excel download and every request log raise
on a new machine.
"""
import os

import pytest
from flask import Flask
from openpyxl import load_workbook

from ltv_app import _ensure_instance_dirs
from ltv_app.blueprints.database.views import save_log


def test_ensure_instance_dirs_creates_temp_on_a_fresh_instance(tmp_path):
    inst = tmp_path / "instance"

    _ensure_instance_dirs(str(inst))

    assert (inst / "temp").is_dir(), "temp/ must exist; six blueprints write into it"
    assert (inst / "test_database").is_dir()


def test_ensure_instance_dirs_creates_temp_when_instance_already_exists(tmp_path):
    """Regression for the original bug: `os.makedirs(instance_path)` raised
    OSError when instance/ already existed, and the bare `except OSError: pass`
    swallowed it — so the subdirectories after it were never created."""
    inst = tmp_path / "instance"
    inst.mkdir()

    _ensure_instance_dirs(str(inst))

    assert (inst / "temp").is_dir()
    assert (inst / "test_database").is_dir()


def test_ensure_instance_dirs_is_idempotent(tmp_path):
    inst = tmp_path / "instance"

    _ensure_instance_dirs(str(inst))
    _ensure_instance_dirs(str(inst))  # must not raise

    assert (inst / "temp").is_dir()


def _app_with_instance(tmp_path):
    return Flask(__name__, instance_path=str(tmp_path))


def test_save_log_creates_the_workbook_when_it_is_missing(tmp_path):
    app = _app_with_instance(tmp_path)
    log = tmp_path / "data_logs.xlsx"
    assert not log.exists()

    with app.app_context():
        save_log("http://example/trades", "")

    assert log.exists(), "save_log must create data_logs.xlsx rather than raise"
    ws = load_workbook(log)["LOGS"]
    assert ws["B1"].value == "http://example/trades"
    assert ws["A1"].value is not None  # timestamp


def test_save_log_appends_to_an_existing_workbook(tmp_path):
    app = _app_with_instance(tmp_path)

    with app.app_context():
        save_log("http://example/one", "")
        save_log("http://example/two", "data")

    ws = load_workbook(tmp_path / "data_logs.xlsx")["LOGS"]
    assert ws["B1"].value == "http://example/one"
    assert ws["B2"].value == "http://example/two"
    assert ws["C2"].value == "data"
