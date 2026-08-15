"""Before/after behaviour-preservation check for the gain_loss query refactor.

Generates real gain-loss workbooks via create_file() against a FROZEN COPY of
the local DB, for five parameter combinations covering every branch of
_get_refs_and_name() plus the short book, and dumps every cell's
(value, number_format, bold, fill, merge, hyperlink) for every sheet -- plus
the gather_gain_loss() dict -- to a deterministic text file.

Run once BEFORE the refactor (saves snapshot), then again AFTER (compares) --
any diff means the refactor changed behaviour, which is not allowed.

The baseline DB is frozen so that a red result always means "your code
changed", never "the data moved underneath you". Regenerate it only with
--freeze, which is a deliberate act that discards the old before-state.

Run: server/.venv/Scripts/python.exe scripts/verify_gain_loss_refactor.py [--freeze|--save|--check]
"""
import argparse
import json
import os
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
SERVER = os.path.dirname(HERE)
sys.path.insert(0, SERVER)

import openpyxl

from ltv_app import create_app

LIVE_DB = os.path.join(SERVER, "instance", "LTV Stocks.db")
# instance/ is gitignored wholesale. Do NOT put this under scripts/ -- that
# directory is un-ignored recursively, so a copy of real client data there
# would be one `git add -A` away from a public repo.
BASELINE_DB = os.path.join(SERVER, "instance", "_gain_loss_baseline.db")
# Matches the existing /scripts/_*_snapshot.txt ignore rule.
SNAPSHOT_PATH = os.path.join(HERE, "_gain_loss_snapshot.txt")

# Five cases covering every _get_refs_and_name() branch and both books.
# MSPL (bank_ref=11) / 0175 (code_ref=10) is the richest short pair (17 rows).
# Case 5 uses 2018-04 because tbl_transaction_short has NO rows in Aug 2026 --
# without it the entire short-detail writing block would go unexercised.
CASES = [
    ("all-banks_all-codes_2026-08", "2026-08-01", "2026-08-31", 0, 0),
    ("all-banks_one-code_2026-08", "2026-08-01", "2026-08-31", 0, 10),
    ("one-bank_all-codes_2026-08", "2026-08-01", "2026-08-31", 11, 0),
    ("one-bank_one-code_2026-08", "2026-08-01", "2026-08-31", 11, 10),
    ("all-banks_all-codes_2018-04_shorts", "2018-04-01", "2018-04-30", 0, 0),
]


def _fill_of(cell):
    """Foreground colour if a solid fill is set, else None."""
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    rgb = getattr(fill.fgColor, "rgb", None)
    return rgb if isinstance(rgb, str) else None


def dump_workbook(buf, lines):
    wb = openpyxl.load_workbook(buf)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # sheet_state is load-bearing: hidden/visible is computed from which
        # (ccy, bank) groups have data.
        lines.append(f"=== {sheet_name} ({ws.dimensions}) state={ws.sheet_state} ===")
        lines.append("merged\t" + ",".join(sorted(str(r) for r in ws.merged_cells.ranges)))
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None and cell.hyperlink is None:
                    continue
                target = cell.hyperlink.target if cell.hyperlink is not None else None
                lines.append(
                    "\t".join([
                        cell.coordinate,
                        repr(cell.value),
                        str(cell.number_format),
                        "bold" if (cell.font is not None and cell.font.bold) else "-",
                        str(_fill_of(cell)),
                        str(target),
                    ])
                )
    wb.close()


def snapshot_text():
    if not os.path.exists(BASELINE_DB):
        sys.exit(f"No frozen baseline at {BASELINE_DB}. Run with --freeze first.")

    app = create_app()
    app.config["DATABASE"] = BASELINE_DB
    app.config["LOGIN_DISABLED"] = True
    # Pinned so url_for(..., _external=True) hyperlinks are deterministic.
    app.config["SERVER_NAME"] = "localhost"

    lines = []
    for label, date_from, date_to, bank_ref, code_ref in CASES:
        with app.test_request_context("/gain-loss/"):
            from ltv_app.blueprints.database.views import get_db
            actual = get_db().execute("PRAGMA database_list").fetchall()[0][2]
            if os.path.abspath(actual) != os.path.abspath(BASELINE_DB):
                sys.exit(f"REFUSING: connected to {actual}, not the frozen baseline")

            from ltv_app.blueprints.gain_loss.create_gain_loss import (
                create_file, gather_gain_loss, _get_refs_and_name,
            )

            lines.append(f"########## CASE {label} "
                         f"({date_from}..{date_to} bank={bank_ref} code={code_ref}) ##########")

            # Data layer, pinned separately so a red result says *which* layer broke.
            _, bank_refs, code_refs = _get_refs_and_name(date_from, date_to, bank_ref, code_ref)
            data = gather_gain_loss(date_from, date_to, bank_refs, code_refs)
            lines.append("--- gather_gain_loss ---")
            lines.append(json.dumps(data, sort_keys=True, indent=1, default=str))

            # Writing layer.
            lines.append("--- workbook ---")
            output, download_name = create_file(date_from=date_from, date_to=date_to,
                                                bank_ref=bank_ref, code_ref=code_ref)
            lines.append(f"download_name\t{download_name}")
            dump_workbook(output, lines)

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--freeze", action="store_true",
                       help="copy the live DB to the frozen baseline (discards old before-state)")
    group.add_argument("--save", action="store_true", help="save the before-snapshot")
    group.add_argument("--check", action="store_true", help="compare against the saved snapshot")
    args = parser.parse_args()

    if args.freeze:
        shutil.copy2(LIVE_DB, BASELINE_DB)
        print(f"Froze baseline: {BASELINE_DB} ({os.path.getsize(BASELINE_DB) / 1e6:.2f} MB)")
        return

    text = snapshot_text()

    if args.save:
        with open(SNAPSHOT_PATH, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"Saved snapshot: {SNAPSHOT_PATH} ({len(text.splitlines())} lines)")
        return

    with open(SNAPSHOT_PATH, encoding="utf-8") as f:
        before = f.read()

    if text == before:
        print(f"RESULT: IDENTICAL -- refactor is behaviour-preserving "
              f"({len(text.splitlines())} lines, {len(CASES)} cases).")
        return

    before_lines, after_lines = before.splitlines(), text.splitlines()
    print("RESULT: DIFFERS")
    print(f"  before: {len(before_lines)} lines, after: {len(after_lines)} lines")
    case = "(before first case marker)"
    shown = 0
    for i in range(max(len(before_lines), len(after_lines))):
        b = before_lines[i] if i < len(before_lines) else "<missing>"
        a = after_lines[i] if i < len(after_lines) else "<missing>"
        if b.startswith("########## CASE"):
            case = b.split("CASE ")[1].split(" (")[0]
        if b != a:
            print(f"  diff at line {i} [case {case}]:\n    before: {b!r}\n    after:  {a!r}")
            shown += 1
            if shown >= 10:
                print("  ... (further diffs suppressed)")
                break
    sys.exit(1)


if __name__ == "__main__":
    main()
