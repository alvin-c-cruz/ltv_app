"""Verification for the D-circle DrawingML injection (excel_writer.py).

Two parts: an isolated check of _compute_circle_cells' classification logic
(no DB, synthetic rows -- mirrors the exact algorithm given in the spec), and
a live end-to-end check that generates a real workbook against the real DB,
confirms every XML part it touched is well-formed, that openpyxl can reopen
the file without error, and that the injected circles land on exactly the
cells independently recomputed here.

Run: server/.venv/Scripts/python.exe scripts/verify_circle_injection.py
"""
import os
import sys
import zipfile
from datetime import date, timedelta
from xml.etree import ElementTree as ET

HERE = os.path.dirname(os.path.abspath(__file__))
SERVER = os.path.dirname(HERE)
sys.path.insert(0, SERVER)

from openpyxl.utils import get_column_letter

from ltv_app import create_app
from ltv_app.blueprints.database.views import get_db
from ltv_app.blueprints.ltv_stocks.legacy_port.excel_writer import (
    _compute_circle_cells, build_workbook,
)
from ltv_app.blueprints.ltv_stocks.legacy_port.working_day import WorkingDay
from ltv_app.tz import ph_today

ok = True


def check(label, actual, expected):
    global ok
    if actual == expected:
        print(f"  {label}: PASS")
    else:
        ok = False
        print(f"  {label}: FAIL  expected {expected!r}, got {actual!r}")


class _Rec(dict):
    def __getitem__(self, key):
        return dict.__getitem__(self, key)


class _NeverHoliday:
    def is_holiday(self, d):
        return False


def isolated_checks():
    print("Isolated _compute_circle_cells checks:")
    date_range = [date(2026, 7, i) for i in (6, 7, 8, 9, 10, 13, 14)]
    wd = _NeverHoliday()
    price_lookup = lambda code_ref, d: {
        date(2026, 7, 6): 100.0,
        date(2026, 7, 7): 120.0,
        date(2026, 7, 8): 0,
        date(2026, 7, 9): None,
        date(2026, 7, 10): "Done",
        date(2026, 7, 13): 150.0,
        date(2026, 7, 14): 100.0,  # report_date -- must be skipped regardless
    }[d]

    # Row: spot=200 > strike=150 (above branch), ko=180.
    rec_above = _Rec(code_ref=1, spot=200.0, strike=150.0, ko=180.0,
                      start_date=date(2026, 7, 6), end_date=date(2026, 7, 20))
    cells = _compute_circle_cells([rec_above], 10, date_range, date(2026, 7, 14), wd, price_lookup)
    # 07/06 px=100: KO? 180<=100 no. D? 150>=100 yes -> D at O10.
    # 07/07 px=120: 150>=120 yes -> D at P10.
    # 07/08 px=0 -> skip (zero).
    # 07/09 px=None -> skip.
    # 07/10 px="Done" (string) -> skip (not int/float).
    # 07/13 px=150: KO? 180<=150 no. D? 150>=150 yes -> D at T10.
    # 07/14 == report_date, has a recorded price (100.0) -> classified like any
    # other day: KO? 180<=100 no. D? 150>=100 yes -> D at U10.
    check("above-branch D cells", sorted(cells), sorted([('O', 10), ('P', 10), ('T', 10), ('U', 10)]))

    # report_date with NO recorded price, AND report_date IS the actual current
    # day -> inherits the last classified real trading day's status instead of
    # being skipped. Must use the real ph_today() as report_date here, since
    # the inheritance path is gated on `report_date == ph_today()` (there is
    # no live closing_price formula, and therefore nothing to inherit toward,
    # for any date other than today).
    today = ph_today()
    today_date_range = [today - timedelta(days=n) for n in (8, 7, 6, 5, 4, 1, 0)]
    price_lookup_today_no_price = lambda code_ref, d: {
        today_date_range[0]: 100.0,
        today_date_range[1]: 120.0,
        today_date_range[2]: 0,
        today_date_range[3]: None,
        today_date_range[4]: "Done",
        today_date_range[5]: 150.0,
        today_date_range[6]: None,  # report_date == today -- no data yet
    }[d]
    rec_above_today = _Rec(code_ref=1, spot=200.0, strike=150.0, ko=180.0,
                            start_date=today_date_range[0], end_date=today_date_range[0] + timedelta(days=30))
    cells = _compute_circle_cells(
        [rec_above_today], 10, today_date_range, today, wd, price_lookup_today_no_price
    )
    check("report_date (== today) inherits previous day's D status", sorted(cells),
          sorted([('O', 10), ('P', 10), ('T', 10), ('U', 10)]))

    # Same setup, but report_date is NOT today (a hypothetical/other-date run
    # -- report_date still matches the grid's last column, it just isn't the
    # real current day) -> no live formula exists for that date, so nothing to
    # inherit toward; a missing price there must be skipped like any other
    # gap, not circled.
    not_today = date(2020, 1, 10)  # deterministic, clearly never "today"
    not_today_date_range = [not_today - timedelta(days=n) for n in (8, 7, 6, 5, 4, 1, 0)]
    price_lookup_not_today_no_price = lambda code_ref, d: {
        not_today_date_range[0]: 100.0,
        not_today_date_range[1]: 120.0,
        not_today_date_range[2]: 0,
        not_today_date_range[3]: None,
        not_today_date_range[4]: "Done",
        not_today_date_range[5]: 150.0,
        not_today_date_range[6]: None,  # report_date == not_today_date_range[-1] -- no data
    }[d]
    rec_above_not_today = _Rec(code_ref=1, spot=200.0, strike=150.0, ko=180.0,
                                start_date=not_today_date_range[0], end_date=not_today_date_range[0] + timedelta(days=30))
    cells = _compute_circle_cells(
        [rec_above_not_today], 10, not_today_date_range, not_today, wd, price_lookup_not_today_no_price
    )
    # Earlier real-priced days still classify normally (O/P/T, same pattern as
    # "above-branch D cells") -- only the last column (report_date, no price,
    # not today) must be excluded, proving specifically that it doesn't
    # inherit when report_date isn't the real current day.
    check("report_date != today -> no inheritance, missing price skipped", sorted(cells),
          sorted([('O', 10), ('P', 10), ('T', 10)]))

    # Row: spot=100 <= strike=150 (else branch), ko=90.
    rec_below = _Rec(code_ref=1, spot=100.0, strike=150.0, ko=90.0,
                      start_date=date(2026, 7, 6), end_date=date(2026, 7, 20))
    cells = _compute_circle_cells([rec_below], 20, date_range, date(2026, 7, 14), wd, price_lookup)
    # 07/06 px=100: KO? 90>=100 no. D? 150<=100 no -> "." (no circle).
    # 07/07 px=120: KO? 90>=120 no. D? 150<=120 no -> "." (no circle).
    # 07/13 px=150: KO? 90>=150 no. D? 150<=150 yes -> D at T20.
    check("else-branch D cells", cells, [('T', 20)])

    # start_date/end_date window: date outside range never circled even with a real price.
    rec_window = _Rec(code_ref=1, spot=200.0, strike=150.0, ko=180.0,
                       start_date=date(2026, 7, 8), end_date=date(2026, 7, 9))
    cells = _compute_circle_cells([rec_window], 30, date_range, date(2026, 7, 14), wd, price_lookup)
    check("outside start/end window -> no circles", cells, [])


def live_check():
    print("Live DBPe-HKD end-to-end check:")
    app = create_app()
    app.config["DATABASE"] = os.path.join(SERVER, "instance", "LTV Stocks.db")
    ctx = app.app_context()
    ctx.push()
    try:
        db = get_db()
        from ltv_app.tz import ph_today
        report_date = ph_today()
        from ltv_app.blueprints.ltv_stocks.legacy_port.term_sheet_calc import contract_records
        from ltv_app.blueprints.ltv_stocks.legacy_port.excel_writer import week_dates
        from ltv_app.blueprints.ltv_stocks.legacy_port.stock_price import get_stock_price

        bank_ref = 5  # DBPe
        accu = [r for r in contract_records(db, bank_ref, 'ACCU') if r['ccy_id'] == 'HKD']
        decu = [r for r in contract_records(db, bank_ref, 'DECU') if r['ccy_id'] == 'HKD']
        date_range = week_dates(report_date)
        wd = WorkingDay(db, 'HKD')
        price_lookup = lambda code_ref, d: get_stock_price(db, code_ref, d)

        # DBPe-HKD's own report_header() shape (no subtitle, ccy == 'HKD') always
        # returns 3 for accu_count_row; decu_count_row is accu_count_row + 6 + n_accu,
        # matching _write_contracts' own return-value arithmetic (first_data_row + n + 2).
        accu_count_row = 3
        decu_count_row = accu_count_row + 6 + len(accu)

        expected = set(
            _compute_circle_cells(accu, accu_count_row + 4, date_range, report_date, wd, price_lookup)
            + _compute_circle_cells(decu, decu_count_row + 4, date_range, report_date, wd, price_lookup)
            # Cells (col B) above "CODE", circled on DBPe-HKD only.
            + [('B', accu_count_row), ('B', decu_count_row)]
        )

        buf = build_workbook(db, report_date, ['DBPe'])
    finally:
        ctx.pop()

    zf = zipfile.ZipFile(buf)

    bad = []
    for name in zf.namelist():
        if name.endswith('.xml') or name.endswith('.rels'):
            try:
                ET.fromstring(zf.read(name))
            except ET.ParseError as e:
                bad.append((name, str(e)))
    check("all XML parts well-formed", bad, [])

    drawing_parts = [n for n in zf.namelist() if n.startswith('xl/drawings/drawing') and n.endswith('.xml')]
    check("exactly one drawing part for one-sheet export", len(drawing_parts), 1)

    ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'}
    root = ET.fromstring(zf.read(drawing_parts[0]))
    actual = set()
    for anchor in root.findall('xdr:twoCellAnchor', ns):
        col = int(anchor.find('xdr:from/xdr:col', ns).text)
        row = int(anchor.find('xdr:from/xdr:row', ns).text)
        actual.add((get_column_letter(col + 1), row + 1))
    check("circled cells match independently recomputed set", actual, expected)
    print(f"    ({len(expected)} circles expected on DBPe-HKD as of {report_date})")

    import openpyxl
    buf.seek(0)
    try:
        openpyxl.load_workbook(buf)
        check("openpyxl reopens the file without error", True, True)
    except Exception as e:
        check("openpyxl reopens the file without error", str(e), None)


def main():
    isolated_checks()
    live_check()
    print("RESULT:", "ALL PASS" if ok else "FAIL")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
