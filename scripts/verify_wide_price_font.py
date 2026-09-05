"""Verification that four-figure prices are rendered small enough to fit.

Columns E:G (spot, strike, ko) are 8.28 characters wide and the daily closing
prices in O:X are 6.42, both formatted to 4 and 2 decimals respectively. A price
in the hundreds fits at the normal 10pt; a price in the thousands renders as
"1,226.8250" -- ten characters -- and Excel clips it or shows ####. Reported
2026-09-06 against a stock trading either side of HKD 1,000 (see server/BUGS.md).

Contract rows priced that high now drop to 8pt. The rule is judged per row and
from the contract's own spot/strike/ko, so one expensive stock does not shrink
the figures on every other row of the sheet.

Case A pins the size rule itself, including the boundary. Case B is structural
and is the one that matters: it walks every contract row of every sheet in a
real generated workbook and asserts the size is 8 exactly when a price exceeds
the threshold and 10 otherwise -- an invariant read entirely off the workbook,
so it holds for stocks nobody has thought of yet. B needs no knowledge of which
codes are expensive, and C guards it against becoming vacuous.

Contract rows are told apart from position rows by their number format:
`_write_contracts` writes E:G as '#,##0.0000', while `_write_positions` writes
share counts as '#,##0' in overlapping columns.

Read-only: builds a workbook in memory from the live database and inspects it.
No client figures are pinned here -- every expectation is derived at runtime.

Run: server/.venv/Scripts/python.exe scripts/verify_wide_price_font.py
"""
import os
import sys
from datetime import date

import openpyxl
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SERVER = os.path.dirname(HERE)
sys.path.insert(0, SERVER)

from ltv_app import create_app
from ltv_app.blueprints.database.views import get_db
from ltv_app.blueprints.ltv_stocks.legacy_port.excel_writer import (
    build_workbook, _price_font_size, _WIDE_PRICE_THRESHOLD, _WIDE_PRICE_FONT_SIZE,
    _OX_COLS,
)

DB_PATH = os.path.join(SERVER, "instance", "LTV Stocks.db")
BANK_IDS = ["DBPe", "DBPL", "SHK", "SHK2", "MST1", "MST2", "MSPL", "NSG"]
REPORT_DATE = date(2026, 9, 4)

CONTRACT_PRICE_FORMAT = '#,##0.0000'
NORMAL_SIZE = 10

_ok = True


def _case(label, actual, expected):
    global _ok
    if actual == expected:
        print(f"  {label}: PASS")
        return True
    # Never interpolates a price into the failure line -- this repository is public.
    print(f"  {label}: FAIL  expected {expected!r}, got {actual!r}")
    _ok = False
    return False


# --- Case A: the rule ---------------------------------------------------------
print("A  the size rule")
over = _WIDE_PRICE_THRESHOLD + 1
_case("A1 a four-figure spot shrinks the row",
      _price_font_size({"spot": over, "strike": 1.0, "ko": 1.0}), _WIDE_PRICE_FONT_SIZE)
_case("A2 a four-figure strike shrinks the row",
      _price_font_size({"spot": 1.0, "strike": over, "ko": 1.0}), _WIDE_PRICE_FONT_SIZE)
_case("A3 a four-figure ko shrinks the row",
      _price_font_size({"spot": 1.0, "strike": 1.0, "ko": over}), _WIDE_PRICE_FONT_SIZE)
_case("A4 ordinary prices stay at the normal size",
      _price_font_size({"spot": 1.0, "strike": 2.0, "ko": 3.0}), NORMAL_SIZE)
# "exceeds a thousand" -- exactly at the threshold still fits, so it stays.
_case("A5 exactly at the threshold is not shrunk",
      _price_font_size({"spot": _WIDE_PRICE_THRESHOLD, "strike": 1.0, "ko": 1.0}),
      NORMAL_SIZE)
_case("A6 a missing price is not treated as wide",
      _price_font_size({"spot": None, "strike": None, "ko": None}), NORMAL_SIZE)


# --- Case B: the invariant over a real workbook -------------------------------
print("B  every contract row in a generated workbook")
app = create_app()
app.config["DATABASE"] = DB_PATH
with app.app_context():
    buf = build_workbook(get_db(), REPORT_DATE, BANK_IDS)

wb = openpyxl.load_workbook(buf)

wide_rows = 0
normal_rows = 0
wrong_size = 0
wrong_daily = 0

for name in wb.sheetnames:
    ws = wb[name]
    for row in ws.iter_rows():
        # Merged cells (the position table merges several ranges) carry no
        # column_letter; address by index instead.
        cells = {get_column_letter(c.column): c for c in row}
        price_cells = [cells.get(col) for col in ("E", "F", "G")]
        if not all(c is not None and c.number_format == CONTRACT_PRICE_FORMAT
                   for c in price_cells):
            continue
        values = [c.value for c in price_cells
                  if isinstance(c.value, (int, float))]
        if not values:
            continue

        is_wide = max(values) > _WIDE_PRICE_THRESHOLD
        expected = _WIDE_PRICE_FONT_SIZE if is_wide else NORMAL_SIZE
        if is_wide:
            wide_rows += 1
        else:
            normal_rows += 1

        for c in price_cells:
            if c.font.size != expected:
                wrong_size += 1

        # The daily closing prices on the same row carry the same magnitude and
        # sit in narrower columns, so they follow the row. Text placeholders
        # ("Done") and empties are not prices and are left alone.
        for col in _OX_COLS:
            c = cells.get(col)
            if c is not None and isinstance(c.value, (int, float)):
                if c.font.size != expected:
                    wrong_daily += 1

_case("B1 contract rows whose price cells are the wrong size", wrong_size, 0)
_case("B2 daily price cells that do not follow their row", wrong_daily, 0)


# --- Case C: the invariant is not vacuous -------------------------------------
print("C  the sample actually exercises both sides")
_case("C1 at least one four-figure contract row was rendered", wide_rows > 0, True)
_case("C2 at least one ordinary-priced row was rendered (control)",
      normal_rows > 0, True)

print("RESULT:", "ALL PASS" if _ok else "FAIL")
sys.exit(0 if _ok else 1)
