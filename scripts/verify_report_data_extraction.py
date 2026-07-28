"""Before/after behavior-preservation check for the report_data.py extraction.

Generates a real workbook via build_workbook() against the live local DB,
for a report date/bank list with real ACCU+DECU+position data, and dumps
every cell's (value, number_format) for every sheet to a deterministic text
file. Run once BEFORE the extraction (saves snapshot), then again AFTER
(compares against the saved snapshot) -- any diff means the refactor changed
behavior, which is not allowed.

Run: server/.venv/Scripts/python.exe scripts/verify_report_data_extraction.py [--save|--check]
"""
import argparse
import os
import sys
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
SERVER = os.path.dirname(HERE)
sys.path.insert(0, SERVER)
SNAPSHOT_PATH = os.path.join(HERE, "_report_data_extraction_snapshot.txt")

import openpyxl

from ltv_app import create_app
from ltv_app.blueprints.database.views import get_db
from ltv_app.blueprints.ltv_stocks.legacy_port.excel_writer import build_workbook

REPORT_DATE = date(2026, 7, 27)
BANK_IDS = ['DBPe', 'DBPL', 'SHK', 'SHK2', 'MST1', 'MST2', 'MSPL', 'NSG']


def snapshot_text():
    app = create_app()
    app.config["DATABASE"] = os.path.join(SERVER, "instance", "LTV Stocks.db")
    ctx = app.app_context()
    ctx.push()
    try:
        db = get_db()
        buf = build_workbook(db, REPORT_DATE, BANK_IDS)
    finally:
        ctx.pop()

    wb = openpyxl.load_workbook(buf)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== {sheet_name} ({ws.dimensions}) ===")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    lines.append(f"{cell.coordinate}\t{cell.value!r}\t{cell.number_format}")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--save", action="store_true", help="save the before-snapshot")
    group.add_argument("--check", action="store_true", help="compare against the saved snapshot")
    args = parser.parse_args()

    text = snapshot_text()

    if args.save:
        with open(SNAPSHOT_PATH, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"Saved snapshot: {SNAPSHOT_PATH} ({len(text.splitlines())} lines)")
        sys.exit(0)

    with open(SNAPSHOT_PATH, encoding="utf-8") as f:
        before = f.read()
    if text == before:
        print("RESULT: IDENTICAL -- extraction is behavior-preserving.")
        sys.exit(0)
    else:
        before_lines = before.splitlines()
        after_lines = text.splitlines()
        print("RESULT: DIFFERS")
        print(f"  before: {len(before_lines)} lines, after: {len(after_lines)} lines")
        for i, (b, a) in enumerate(zip(before_lines, after_lines)):
            if b != a:
                print(f"  first diff at line {i}:\n    before: {b!r}\n    after:  {a!r}")
                break
        sys.exit(1)


if __name__ == "__main__":
    main()
