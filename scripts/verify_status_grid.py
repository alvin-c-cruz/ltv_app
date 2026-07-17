"""Verification for _write_status_grid (excel_writer.py).

Builds a fresh in-memory worksheet and calls _write_status_grid directly with
synthetic inputs (no DB needed -- the function only writes cells from
count_row/n/date_range, and the formulas it writes are pure cell references,
not value-dependent). Then does one live end-to-end sanity check against the
real DB, confirming the DBPe-HKD sheet's actual ACCU/DECU row placement
matches what the isolated checks predict.

Run: server/.venv/Scripts/python.exe scripts/verify_status_grid.py
"""
import os
import sys
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
SERVER = os.path.dirname(HERE)
sys.path.insert(0, SERVER)

import openpyxl

from ltv_app import create_app
from ltv_app.blueprints.database.views import get_db
from ltv_app.blueprints.ltv_stocks.legacy_port.excel_writer import (
    _write_status_grid, build_workbook,
)

DATE_RANGE = [date(2026, 6, 29), date(2026, 6, 30), date(2026, 7, 1),
              date(2026, 7, 2), date(2026, 7, 3), date(2026, 7, 6),
              date(2026, 7, 7), date(2026, 7, 8), date(2026, 7, 9),
              date(2026, 7, 10)]


def _formula(r, price_col):
    return (
        f'=IF($E{r}>$F{r},'
        f'IF({price_col}{r}=0,"xxx",IF({price_col}{r}="","",'
        f'IF($Z{r}=2,IF($G{r}<={price_col}{r},"KO",IF($F{r}>={price_col}{r},"D",".")),'
        f'IF($G{r}<={price_col}{r},"KO",".")))),'
        f'IF({price_col}{r}=0,"xxx",IF({price_col}{r}="","",'
        f'IF($Z{r}=2,IF($G{r}>={price_col}{r},"KO",IF($F{r}<={price_col}{r},"D",".")),'
        f'IF($G{r}>={price_col}{r},"KO",".")))))'
    )


def _case(label, actual, expected):
    if actual == expected:
        print(f"  {label}: PASS")
        return True
    print(f"  {label}: FAIL  expected {expected!r}, got {actual!r}")
    return False


def check_isolated():
    ok = True
    wb = openpyxl.Workbook()
    ws = wb.active

    # ACCU-shaped block: count_row=3, n=7 -> header row 5, body rows 7-13.
    _write_status_grid(ws, count_row=3, n=7, date_range=DATE_RANGE)

    ok &= _case("ACCU header AA5 value", ws['AA5'].value, DATE_RANGE[0])
    ok &= _case("ACCU header AA5 number_format", ws['AA5'].number_format, 'm/d')
    ok &= _case("ACCU header AJ5 value", ws['AJ5'].value, DATE_RANGE[9])
    ok &= _case("ACCU Z7 value", ws['Z7'].value, 2)
    ok &= _case("ACCU Z13 value", ws['Z13'].value, 2)
    ok &= _case("ACCU AA7 formula", ws['AA7'].value, _formula(7, 'O'))
    ok &= _case("ACCU AJ13 formula", ws['AJ13'].value, _formula(13, 'X'))
    ok &= _case("ACCU AA6 untouched (above header)", ws['AA6'].value, None)
    ok &= _case("ACCU Z14 untouched (below body)", ws['Z14'].value, None)

    # DECU-shaped block: count_row=16, n=4 -> header row 18, body rows 20-23.
    _write_status_grid(ws, count_row=16, n=4, date_range=DATE_RANGE)

    ok &= _case("DECU header AA18 value", ws['AA18'].value, DATE_RANGE[0])
    ok &= _case("DECU Z20 value", ws['Z20'].value, 2)
    ok &= _case("DECU Z23 value", ws['Z23'].value, 2)
    ok &= _case("DECU AA20 formula", ws['AA20'].value, _formula(20, 'O'))

    # n=0 -> nothing written at all.
    _write_status_grid(ws, count_row=50, n=0, date_range=DATE_RANGE)
    ok &= _case("n=0 header untouched", ws['AA52'].value, None)
    ok &= _case("n=0 body untouched", ws['Z54'].value, None)

    return ok


def check_live():
    app = create_app()
    app.config["DATABASE"] = os.path.join(SERVER, "instance", "LTV Stocks.db")
    ctx = app.app_context()
    ctx.push()
    try:
        db = get_db()
        buf = build_workbook(db, date(2026, 7, 10), ['DBPe'])
    finally:
        ctx.pop()

    wb = openpyxl.load_workbook(buf)
    if 'DBPe-HKD' not in wb.sheetnames:
        print("  live DBPe-HKD: FAIL  sheet not found")
        return False
    ws = wb['DBPe-HKD']
    ok = True
    ok &= _case("live DBPe-HKD Z7", ws['Z7'].value, 2)
    ok &= _case("live DBPe-HKD Z13", ws['Z13'].value, 2)
    ok &= _case("live DBPe-HKD AA7 formula", ws['AA7'].value, _formula(7, 'O'))
    ok &= _case("live DBPe-HKD Z20", ws['Z20'].value, 2)
    ok &= _case("live DBPe-HKD Z23", ws['Z23'].value, 2)
    ok &= _case("live DBPe-HKD AA20 formula", ws['AA20'].value, _formula(20, 'O'))
    return ok


def main():
    print("Isolated checks:")
    ok1 = check_isolated()
    print("Live DBPe-HKD sanity check:")
    ok2 = check_live()
    ok = ok1 and ok2
    print("RESULT:", "ALL PASS" if ok else "FAIL")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
